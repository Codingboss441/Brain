import os
import requests
import json
import time
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
from io import BytesIO
from PIL import Image, ImageEnhance, ImageFilter
from bs4 import BeautifulSoup
from dotenv import load_dotenv
import anthropic
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from concurrent.futures import ThreadPoolExecutor, as_completed
import re
from typing import Dict, List, Tuple, Optional
import smtplib
import boto3
from botocore.exceptions import ClientError, NoCredentialsError
import io
import re
import time
import uuid
from typing import Dict, Any, Optional, Tuple
import boto3
import requests
from PIL import Image
import pytesseract
from enum import Enum
from dataclasses import dataclass
import logging
import fitz

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# Replace the DocumentAnalyzer class with this version:
class DocumentType(Enum):
    """Enum for document types with proper categorization"""
    # Identity Documents
    AADHAAR_CARD = "aadhaar_card"
    PAN_CARD = "pan_card"
    PASSPORT = "passport"
    DRIVING_LICENSE = "driving_license"
    VOTER_ID = "voter_id"
    
    # Financial Documents
    BANK_STATEMENT = "bank_statement"
    CANCELLED_CHEQUE = "cancelled_cheque"
    SALARY_SLIP = "salary_slip"
    INVOICE = "invoice"
    RECEIPT = "receipt"
    
    # Insurance Documents
    POLICY_DOCUMENT = "policy_document"
    CLAIM_FORM = "claim_form"
    COVERAGE_NOTE = "coverage_note"
    ENDORSEMENT = "endorsement"
    RENEWAL_NOTICE = "renewal_notice"
    
    # Vehicle Documents
    RC_BOOK = "rc_book"
    POLLUTION_CERTIFICATE = "pollution_certificate"
    VEHICLE_INSURANCE = "vehicle_insurance"
    
    # Medical Documents
    MEDICAL_REPORT = "medical_report"
    PRESCRIPTION = "prescription"
    DISCHARGE_SUMMARY = "discharge_summary"
    TEST_RESULTS = "test_results"
    HOSPITAL_BILL = "hospital_bill"
    
    # Legal Documents
    AGREEMENT = "agreement"
    AFFIDAVIT = "affidavit"
    POWER_OF_ATTORNEY = "power_of_attorney"
    COURT_ORDER = "court_order"
    POLICE_REPORT = "police_report"
    FIR_COPY = "fir_copy"
    
    # Supporting Documents
    SURVEY_REPORT = "survey_report"
    REPAIR_ESTIMATE = "repair_estimate"
    PHOTOGRAPHS = "photographs"
    VIDEO_EVIDENCE = "video_evidence"
    
    # Unknown
    UNKNOWN = "unknown"

@dataclass
class DocumentMetadata:
    """Metadata for processed documents"""
    file_name: str
    file_size: int
    upload_time: datetime
    processing_time: float
    s3_key: Optional[str] = None
    error_message: Optional[str] = None

# Replace the existing DocumentAnalyzer class with this updated version

class DocumentAnalyzer:
    """
    Document analyzer using A Image Reader API instead of AWS Textract.
    """
    
    def __init__(self, api_endpoint="https://apiimagereader.insurancedekho.com/document/reader"):
        self.api_endpoint = api_endpoint
        self.cache_enabled = True
        self.max_workers = 5
        self._cache = {}
        
        # Initialize document mappings
        self._init_document_mappings()
        
        logger.info(f"DocumentAnalyzer initialized with API endpoint: {api_endpoint}")
    
    def _init_document_mappings(self):
        """Initialize comprehensive document mappings"""
        # Keep all the existing mappings as they are
        self.document_name_mapping = {
            # Identity Documents
            DocumentType.AADHAAR_CARD: 'Aadhaar Card',
            DocumentType.PAN_CARD: 'PAN Card',
            DocumentType.PASSPORT: 'Passport',
            DocumentType.DRIVING_LICENSE: 'Driving License',
            DocumentType.VOTER_ID: 'Voter ID Card',
            
            # Financial Documents
            DocumentType.BANK_STATEMENT: 'Bank Statement',
            DocumentType.CANCELLED_CHEQUE: 'Cancelled Cheque',
            DocumentType.SALARY_SLIP: 'Salary Slip',
            DocumentType.INVOICE: 'Invoice',
            DocumentType.RECEIPT: 'Receipt',
            
            # Insurance Documents
            DocumentType.POLICY_DOCUMENT: 'Insurance Policy Document',
            DocumentType.CLAIM_FORM: 'Claim Form',
            DocumentType.COVERAGE_NOTE: 'Coverage Note',
            DocumentType.ENDORSEMENT: 'Policy Endorsement',
            DocumentType.RENEWAL_NOTICE: 'Renewal Notice',
            
            # Vehicle Documents
            DocumentType.RC_BOOK: 'Vehicle Registration Certificate (RC)',
            DocumentType.POLLUTION_CERTIFICATE: 'Pollution Certificate',
            DocumentType.VEHICLE_INSURANCE: 'Vehicle Insurance',
            
            # Medical Documents
            DocumentType.MEDICAL_REPORT: 'Medical Report',
            DocumentType.PRESCRIPTION: 'Medical Prescription',
            DocumentType.DISCHARGE_SUMMARY: 'Hospital Discharge Summary',
            DocumentType.TEST_RESULTS: 'Medical Test Results',
            DocumentType.HOSPITAL_BILL: 'Hospital Bill',
            
            # Legal Documents
            DocumentType.AGREEMENT: 'Legal Agreement',
            DocumentType.AFFIDAVIT: 'Affidavit',
            DocumentType.POWER_OF_ATTORNEY: 'Power of Attorney',
            DocumentType.COURT_ORDER: 'Court Order',
            DocumentType.POLICE_REPORT: 'Police Report',
            DocumentType.FIR_COPY: 'FIR Copy',
            
            # Supporting Documents
            DocumentType.SURVEY_REPORT: 'Survey Report',
            DocumentType.REPAIR_ESTIMATE: 'Repair Estimate',
            DocumentType.PHOTOGRAPHS: 'Incident Photographs',
            DocumentType.VIDEO_EVIDENCE: 'Video Evidence'
        }
        
        # Keep all other mappings (document_categories, document_patterns) as they are
        self.document_categories = {
            'identity': {
                'types': [DocumentType.AADHAAR_CARD, DocumentType.PAN_CARD, 
                         DocumentType.PASSPORT, DocumentType.DRIVING_LICENSE, DocumentType.VOTER_ID],
                'keywords': ['identity', 'identification', 'government', 'official', 'photo id'],
                'priority': 10
            },
            'financial': {
                'types': [DocumentType.BANK_STATEMENT, DocumentType.CANCELLED_CHEQUE,
                         DocumentType.SALARY_SLIP, DocumentType.INVOICE, DocumentType.RECEIPT],
                'keywords': ['bank', 'financial', 'transaction', 'payment', 'account'],
                'priority': 8
            },
            'insurance': {
                'types': [DocumentType.POLICY_DOCUMENT, DocumentType.CLAIM_FORM,
                         DocumentType.COVERAGE_NOTE, DocumentType.ENDORSEMENT, DocumentType.RENEWAL_NOTICE],
                'keywords': ['insurance', 'policy', 'claim', 'coverage', 'premium'],
                'priority': 9
            },
            'vehicle': {
                'types': [DocumentType.RC_BOOK, DocumentType.POLLUTION_CERTIFICATE,
                         DocumentType.VEHICLE_INSURANCE],
                'keywords': ['vehicle', 'car', 'bike', 'registration', 'transport'],
                'priority': 7
            },
            'medical': {
                'types': [DocumentType.MEDICAL_REPORT, DocumentType.PRESCRIPTION,
                         DocumentType.DISCHARGE_SUMMARY, DocumentType.TEST_RESULTS, DocumentType.HOSPITAL_BILL],
                'keywords': ['medical', 'hospital', 'doctor', 'patient', 'treatment'],
                'priority': 8
            },
            'legal': {
                'types': [DocumentType.AGREEMENT, DocumentType.AFFIDAVIT,
                         DocumentType.POWER_OF_ATTORNEY, DocumentType.COURT_ORDER,
                         DocumentType.POLICE_REPORT, DocumentType.FIR_COPY],
                'keywords': ['legal', 'court', 'police', 'agreement', 'affidavit'],
                'priority': 7
            },
            'claim_supporting': {
                'types': [DocumentType.SURVEY_REPORT, DocumentType.REPAIR_ESTIMATE,
                         DocumentType.PHOTOGRAPHS, DocumentType.VIDEO_EVIDENCE],
                'keywords': ['survey', 'damage', 'repair', 'incident', 'evidence'],
                'priority': 6
            }
        }
        
        # Keep document patterns as they are
        self.document_patterns = {
            DocumentType.AADHAAR_CARD: {
                'patterns': [
                    r'\b\d{4}\s?\d{4}\s?\d{4}\b',  # Aadhaar number
                    r'(?i)aadhaar|uid|uidai|आधार'
                ],
                'required_fields': ['aadhaar_number'],
                'confidence_boost': 0.3
            },
            DocumentType.PAN_CARD: {
                'patterns': [
                    r'\b[A-Z]{5}[0-9]{4}[A-Z]\b',  # PAN number
                    r'(?i)permanent\s+account\s+number|income\s+tax'
                ],
                'required_fields': ['pan_number'],
                'confidence_boost': 0.3
            },
            # ... keep all other patterns as they are
        }

    def _call_image_reader_api(self, image_url=None, image_content=None, doc_type=None):
        """Call the InsuranceDekho Image Reader API"""
        try:
            # Prepare the form data
            form_data = {}
            files = {}
            
            if image_url:
                form_data['image_url'] = image_url
            elif image_content:
                files['image'] = ('document.jpg', image_content, 'image/jpeg')
            
            if doc_type:
                form_data['doc_type'] = doc_type
            
            # Make the API call
            response = requests.post(
                self.api_endpoint,
                data=form_data,
                files=files if files else None,
                timeout=30
            )
            
            response.raise_for_status()
            return response.json()
            
        except requests.exceptions.RequestException as e:
            logger.error(f"API call failed: {str(e)}")
            raise
    
    def _analyze_image(self, content: bytes) -> Tuple[str, Dict[str, str], float]:
        """Analyze image using InsuranceDekho Image Reader API"""
        logger.info("Analyzing image with InsuranceDekho API")
        
        try:
            # Call the API with image content
            result = self._call_image_reader_api(image_content=content)
            
            # Parse the API response
            text = result.get('extracted_text', '')
            data = result.get('extracted_data', {})
            confidence = result.get('confidence', 0.8)
            
            # Convert API response to our format
            extracted_data = {}
            
            # Map common fields from API response
            field_mapping = {
                'name': 'name',
                'aadhaar_number': 'aadhaar_number',
                'pan_number': 'pan_number',
                'date_of_birth': 'date_of_birth',
                'address': 'address',
                'mobile_number': 'mobile_number',
                'email': 'email_address',
                'vehicle_number': 'vehicle_registration_number',
                'policy_number': 'policy_number',
                'claim_number': 'claim_number'
            }
            
            for api_field, our_field in field_mapping.items():
                if api_field in data:
                    extracted_data[our_field] = data[api_field]
            
            return text, extracted_data, confidence
            
        except Exception as e:
            logger.error(f"Image analysis failed: {str(e)}")
            # Fallback to basic extraction
            return "", {}, 0.5
    
    def _analyze_pdf(self, content: bytes, file_name: str) -> Tuple[str, Dict[str, str], float, int]:
        """Analyze PDF by converting pages to images and using the API"""
        logger.info(f"Analyzing PDF: {file_name}")
        
        try:
            # Convert PDF to images
            import fitz  # PyMuPDF
            pdf_document = fitz.open(stream=content, filetype="pdf")
            
            all_text = []
            all_data = {}
            confidence_scores = []
            page_count = len(pdf_document)
            
            for page_num in range(page_count):
                page = pdf_document[page_num]
                
                # Convert page to image
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x zoom for better quality
                img_data = pix.tobytes("png")
                
                # Analyze the page image
                text, data, confidence = self._analyze_image(img_data)
                
                all_text.append(text)
                all_data.update(data)
                confidence_scores.append(confidence)
            
            # Combine results
            combined_text = '\n'.join(all_text)
            avg_confidence = sum(confidence_scores) / len(confidence_scores) if confidence_scores else 0.5
            
            return combined_text, all_data, avg_confidence, page_count
            
        except ImportError:
            logger.error("PyMuPDF not installed. Install with: pip install PyMuPDF")
            # Fallback: treat as single image
            return self._analyze_image(content) + (1,)
        except Exception as e:
            logger.error(f"PDF analysis failed: {str(e)}")
            return "", {}, 0.5, 1
    

import os
import requests
import json
import time
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
from io import BytesIO
from PIL import Image, ImageEnhance, ImageFilter
from bs4 import BeautifulSoup
from dotenv import load_dotenv
import anthropic
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from concurrent.futures import ThreadPoolExecutor, as_completed
import re
from typing import Dict, List, Tuple, Optional
import smtplib
import boto3
from botocore.exceptions import ClientError, NoCredentialsError
import io
import re
import time
import uuid
from typing import Dict, Any, Optional, Tuple
import boto3
import requests
from PIL import Image
import pytesseract
from enum import Enum
from dataclasses import dataclass
import logging
import fitz

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# --- Load Environment Variables ---
def load_config():
    """Load configuration for both desktop and Android"""
    config = {}
    
    # Try .env first (for desktop development)
    try:
        load_dotenv()
        
        config = {
            'FRESHDESK_API_KEY': os.getenv('FRESHDESK_API_KEY'),
            'FRESHDESK_DOMAIN': os.getenv('FRESHDESK_DOMAIN'),
            'CLAUDE_API_KEY': os.getenv('CLAUDE_API_KEY') or os.getenv('ANTHROPIC_API_KEY'),
            'AWS_ACCESS_KEY_ID': os.getenv('AWS_ACCESS_KEY_ID'),
            'AWS_SECRET_ACCESS_KEY': os.getenv('AWS_SECRET_ACCESS_KEY'),
            'AWS_REGION': os.getenv('AWS_REGION', 'us-east-1')
        }
        
        # If we got valid config from .env, use it
        if config['FRESHDESK_API_KEY']:
            return config
    except:
        pass
    
    # Try to load from JSON file (for Android)
    try:
        # Try different paths
        script_dir = os.path.dirname(os.path.abspath(__file__))
        possible_paths = [
            'assets/config.json',
            os.path.join(script_dir, 'assets', 'config.json'),
            os.path.join(os.path.dirname(script_dir), 'assets', 'config.json'),
            '/data/data/com.mycompany.id_brain/files/flet/app/assets/config.json'
        ]
        
        for path in possible_paths:
            if os.path.exists(path):
                print(f"Loading config from: {path}")
                with open(path, 'r') as f:
                    config = json.load(f)
                    # Handle both CLAUDE_API_KEY and ANTHROPIC_API_KEY
                    if 'ANTHROPIC_API_KEY' in config and 'CLAUDE_API_KEY' not in config:
                        config['CLAUDE_API_KEY'] = config['ANTHROPIC_API_KEY']
                    
                    # Set default AWS region if not provided
                    if 'AWS_REGION' not in config:
                        config['AWS_REGION'] = 'us-east-1'
                    
                    return config
    except Exception as e:
        print(f"Error loading config: {e}")
    
    # Fallback - return empty strings
    return {
        'FRESHDESK_API_KEY': '',
        'FRESHDESK_DOMAIN': '',
        'CLAUDE_API_KEY': '',
        'AWS_ACCESS_KEY_ID': '',
        'AWS_SECRET_ACCESS_KEY': '',
        'AWS_REGION': 'us-east-1'
    }

# Load configuration
config = load_config()
FRESHDESK_DOMAIN = config['FRESHDESK_DOMAIN']
FRESHDESK_API_KEY = config['FRESHDESK_API_KEY']
CLAUDE_API_KEY = config['CLAUDE_API_KEY']
AWS_ACCESS_KEY_ID = config.get('AWS_ACCESS_KEY_ID', '')
AWS_SECRET_ACCESS_KEY = config.get('AWS_SECRET_ACCESS_KEY', '')
AWS_REGION = config.get('AWS_REGION', 'us-east-1')

# Check if keys are loaded
if not FRESHDESK_API_KEY or not FRESHDESK_DOMAIN:
    print("Warning: Freshdesk API keys not configured!")
if not CLAUDE_API_KEY:
    print("Warning: Claude/Anthropic API key not configured!")
if not AWS_ACCESS_KEY_ID or not AWS_SECRET_ACCESS_KEY:
    print("Warning: AWS credentials not configured! Document analysis features will be limited.")



# Replace the DocumentAnalyzer class with this version:
class DocumentType(Enum):
    """Enum for document types with proper categorization"""
    # Identity Documents
    AADHAAR_CARD = "aadhaar_card"
    PAN_CARD = "pan_card"
    PASSPORT = "passport"
    DRIVING_LICENSE = "driving_license"
    VOTER_ID = "voter_id"
    
    # Financial Documents
    BANK_STATEMENT = "bank_statement"
    CANCELLED_CHEQUE = "cancelled_cheque"
    SALARY_SLIP = "salary_slip"
    INVOICE = "invoice"
    RECEIPT = "receipt"
    
    # Insurance Documents
    POLICY_DOCUMENT = "policy_document"
    CLAIM_FORM = "claim_form"
    COVERAGE_NOTE = "coverage_note"
    ENDORSEMENT = "endorsement"
    RENEWAL_NOTICE = "renewal_notice"
    
    # Vehicle Documents
    RC_BOOK = "rc_book"
    POLLUTION_CERTIFICATE = "pollution_certificate"
    VEHICLE_INSURANCE = "vehicle_insurance"
    
    # Medical Documents
    MEDICAL_REPORT = "medical_report"
    PRESCRIPTION = "prescription"
    DISCHARGE_SUMMARY = "discharge_summary"
    TEST_RESULTS = "test_results"
    HOSPITAL_BILL = "hospital_bill"
    
    # Legal Documents
    AGREEMENT = "agreement"
    AFFIDAVIT = "affidavit"
    POWER_OF_ATTORNEY = "power_of_attorney"
    COURT_ORDER = "court_order"
    POLICE_REPORT = "police_report"
    FIR_COPY = "fir_copy"
    
    # Supporting Documents
    SURVEY_REPORT = "survey_report"
    REPAIR_ESTIMATE = "repair_estimate"
    PHOTOGRAPHS = "photographs"
    VIDEO_EVIDENCE = "video_evidence"
    
    # Unknown
    UNKNOWN = "unknown"

@dataclass
class DocumentMetadata:
    """Metadata for processed documents"""
    file_name: str
    file_size: int
    upload_time: datetime
    processing_time: float
    s3_key: Optional[str] = None
    error_message: Optional[str] = None

# Replace the existing DocumentAnalyzer class with this updated version

class DocumentAnalyzer:
    """
    Document analyzer using A Image Reader API instead of AWS Textract.
    """
    
    def __init__(self, api_endpoint="https://apiimagereader.insurancedekho.com/document/reader"):
        self.api_endpoint = api_endpoint
        self.cache_enabled = True
        self.max_workers = 5
        self._cache = {}
        
        # Initialize document mappings
        self._init_document_mappings()
        
        logger.info(f"DocumentAnalyzer initialized with API endpoint: {api_endpoint}")
    
    def _init_document_mappings(self):
        """Initialize comprehensive document mappings"""
        # Keep all the existing mappings as they are
        self.document_name_mapping = {
            # Identity Documents
            DocumentType.AADHAAR_CARD: 'Aadhaar Card',
            DocumentType.PAN_CARD: 'PAN Card',
            DocumentType.PASSPORT: 'Passport',
            DocumentType.DRIVING_LICENSE: 'Driving License',
            DocumentType.VOTER_ID: 'Voter ID Card',
            
            # Financial Documents
            DocumentType.BANK_STATEMENT: 'Bank Statement',
            DocumentType.CANCELLED_CHEQUE: 'Cancelled Cheque',
            DocumentType.SALARY_SLIP: 'Salary Slip',
            DocumentType.INVOICE: 'Invoice',
            DocumentType.RECEIPT: 'Receipt',
            
            # Insurance Documents
            DocumentType.POLICY_DOCUMENT: 'Insurance Policy Document',
            DocumentType.CLAIM_FORM: 'Claim Form',
            DocumentType.COVERAGE_NOTE: 'Coverage Note',
            DocumentType.ENDORSEMENT: 'Policy Endorsement',
            DocumentType.RENEWAL_NOTICE: 'Renewal Notice',
            
            # Vehicle Documents
            DocumentType.RC_BOOK: 'Vehicle Registration Certificate (RC)',
            DocumentType.POLLUTION_CERTIFICATE: 'Pollution Certificate',
            DocumentType.VEHICLE_INSURANCE: 'Vehicle Insurance',
            
            # Medical Documents
            DocumentType.MEDICAL_REPORT: 'Medical Report',
            DocumentType.PRESCRIPTION: 'Medical Prescription',
            DocumentType.DISCHARGE_SUMMARY: 'Hospital Discharge Summary',
            DocumentType.TEST_RESULTS: 'Medical Test Results',
            DocumentType.HOSPITAL_BILL: 'Hospital Bill',
            
            # Legal Documents
            DocumentType.AGREEMENT: 'Legal Agreement',
            DocumentType.AFFIDAVIT: 'Affidavit',
            DocumentType.POWER_OF_ATTORNEY: 'Power of Attorney',
            DocumentType.COURT_ORDER: 'Court Order',
            DocumentType.POLICE_REPORT: 'Police Report',
            DocumentType.FIR_COPY: 'FIR Copy',
            
            # Supporting Documents
            DocumentType.SURVEY_REPORT: 'Survey Report',
            DocumentType.REPAIR_ESTIMATE: 'Repair Estimate',
            DocumentType.PHOTOGRAPHS: 'Incident Photographs',
            DocumentType.VIDEO_EVIDENCE: 'Video Evidence'
        }
        
        # Keep all other mappings (document_categories, document_patterns) as they are
        self.document_categories = {
            'identity': {
                'types': [DocumentType.AADHAAR_CARD, DocumentType.PAN_CARD, 
                         DocumentType.PASSPORT, DocumentType.DRIVING_LICENSE, DocumentType.VOTER_ID],
                'keywords': ['identity', 'identification', 'government', 'official', 'photo id'],
                'priority': 10
            },
            'financial': {
                'types': [DocumentType.BANK_STATEMENT, DocumentType.CANCELLED_CHEQUE,
                         DocumentType.SALARY_SLIP, DocumentType.INVOICE, DocumentType.RECEIPT],
                'keywords': ['bank', 'financial', 'transaction', 'payment', 'account'],
                'priority': 8
            },
            'insurance': {
                'types': [DocumentType.POLICY_DOCUMENT, DocumentType.CLAIM_FORM,
                         DocumentType.COVERAGE_NOTE, DocumentType.ENDORSEMENT, DocumentType.RENEWAL_NOTICE],
                'keywords': ['insurance', 'policy', 'claim', 'coverage', 'premium'],
                'priority': 9
            },
            'vehicle': {
                'types': [DocumentType.RC_BOOK, DocumentType.POLLUTION_CERTIFICATE,
                         DocumentType.VEHICLE_INSURANCE],
                'keywords': ['vehicle', 'car', 'bike', 'registration', 'transport'],
                'priority': 7
            },
            'medical': {
                'types': [DocumentType.MEDICAL_REPORT, DocumentType.PRESCRIPTION,
                         DocumentType.DISCHARGE_SUMMARY, DocumentType.TEST_RESULTS, DocumentType.HOSPITAL_BILL],
                'keywords': ['medical', 'hospital', 'doctor', 'patient', 'treatment'],
                'priority': 8
            },
            'legal': {
                'types': [DocumentType.AGREEMENT, DocumentType.AFFIDAVIT,
                         DocumentType.POWER_OF_ATTORNEY, DocumentType.COURT_ORDER,
                         DocumentType.POLICE_REPORT, DocumentType.FIR_COPY],
                'keywords': ['legal', 'court', 'police', 'agreement', 'affidavit'],
                'priority': 7
            },
            'claim_supporting': {
                'types': [DocumentType.SURVEY_REPORT, DocumentType.REPAIR_ESTIMATE,
                         DocumentType.PHOTOGRAPHS, DocumentType.VIDEO_EVIDENCE],
                'keywords': ['survey', 'damage', 'repair', 'incident', 'evidence'],
                'priority': 6
            }
        }
        
        # Keep document patterns as they are
        self.document_patterns = {
            DocumentType.AADHAAR_CARD: {
                'patterns': [
                    r'\b\d{4}\s?\d{4}\s?\d{4}\b',  # Aadhaar number
                    r'(?i)aadhaar|uid|uidai|आधार'
                ],
                'required_fields': ['aadhaar_number'],
                'confidence_boost': 0.3
            },
            DocumentType.PAN_CARD: {
                'patterns': [
                    r'\b[A-Z]{5}[0-9]{4}[A-Z]\b',  # PAN number
                    r'(?i)permanent\s+account\s+number|income\s+tax'
                ],
                'required_fields': ['pan_number'],
                'confidence_boost': 0.3
            },
            # ... keep all other patterns as they are
        }

    def _call_image_reader_api(self, image_url=None, image_content=None, doc_type=None):
        """Call the InsuranceDekho Image Reader API"""
        try:
            # Prepare the form data
            form_data = {}
            files = {}
            
            if image_url:
                form_data['image_url'] = image_url
            elif image_content:
                files['image'] = ('document.jpg', image_content, 'image/jpeg')
            
            if doc_type:
                form_data['doc_type'] = doc_type
            
            # Make the API call
            response = requests.post(
                self.api_endpoint,
                data=form_data,
                files=files if files else None,
                timeout=30
            )
            
            response.raise_for_status()
            return response.json()
            
        except requests.exceptions.RequestException as e:
            logger.error(f"API call failed: {str(e)}")
            raise
    
    def _analyze_image(self, content: bytes) -> Tuple[str, Dict[str, str], float]:
        """Analyze image using InsuranceDekho Image Reader API"""
        logger.info("Analyzing image with InsuranceDekho API")
        
        try:
            # Call the API with image content
            result = self._call_image_reader_api(image_content=content)
            
            # Parse the API response
            text = result.get('extracted_text', '')
            data = result.get('extracted_data', {})
            confidence = result.get('confidence', 0.8)
            
            # Convert API response to our format
            extracted_data = {}
            
            # Map common fields from API response
            field_mapping = {
                'name': 'name',
                'aadhaar_number': 'aadhaar_number',
                'pan_number': 'pan_number',
                'date_of_birth': 'date_of_birth',
                'address': 'address',
                'mobile_number': 'mobile_number',
                'email': 'email_address',
                'vehicle_number': 'vehicle_registration_number',
                'policy_number': 'policy_number',
                'claim_number': 'claim_number'
            }
            
            for api_field, our_field in field_mapping.items():
                if api_field in data:
                    extracted_data[our_field] = data[api_field]
            
            return text, extracted_data, confidence
            
        except Exception as e:
            logger.error(f"Image analysis failed: {str(e)}")
            # Fallback to basic extraction
            return "", {}, 0.5
    
    def _analyze_pdf(self, content: bytes, file_name: str) -> Tuple[str, Dict[str, str], float, int]:
        """Analyze PDF by converting pages to images and using the API"""
        logger.info(f"Analyzing PDF: {file_name}")
        
        try:
            # Convert PDF to images
            import fitz  # PyMuPDF
            pdf_document = fitz.open(stream=content, filetype="pdf")
            
            all_text = []
            all_data = {}
            confidence_scores = []
            page_count = len(pdf_document)
            
            for page_num in range(page_count):
                page = pdf_document[page_num]
                
                # Convert page to image
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x zoom for better quality
                img_data = pix.tobytes("png")
                
                # Analyze the page image
                text, data, confidence = self._analyze_image(img_data)
                
                all_text.append(text)
                all_data.update(data)
                confidence_scores.append(confidence)
            
            # Combine results
            combined_text = '\n'.join(all_text)
            avg_confidence = sum(confidence_scores) / len(confidence_scores) if confidence_scores else 0.5
            
            return combined_text, all_data, avg_confidence, page_count
            
        except ImportError:
            logger.error("PyMuPDF not installed. Install with: pip install PyMuPDF")
            # Fallback: treat as single image
            return self._analyze_image(content) + (1,)
        except Exception as e:
            logger.error(f"PDF analysis failed: {str(e)}")
            return "", {}, 0.5, 1
    
    def analyze_document(self, file_path: str = None, file_url: str = None, 
                        use_cache: bool = True) -> Dict[str, Any]:
        """
        Comprehensive document analysis using InsuranceDekho API
        
        Args:
            file_path: Local file path
            file_url: Remote file URL
            use_cache: Whether to use cached results
            
        Returns:
            Detailed analysis results
        """
        try:
            # Validate inputs
            if not file_path and not file_url:
                raise ValueError("Either file_path or file_url must be provided")
            
            # Generate cache key
            cache_key = None
            if use_cache and self.cache_enabled:
                cache_key = self._generate_cache_key(file_path or file_url)
                if cache_key in self._cache:
                    logger.info(f"Returning cached result for {cache_key}")
                    return self._cache[cache_key]
            
            # Start processing timer
            start_time = time.time()
            
            # Read file content
            content, file_name, metadata = self._read_file_content(file_path, file_url)
            
            # Detect file type
            file_type = self._detect_file_type(content, file_name)
            
            # Determine document type hint for API
            doc_type_hint = self._guess_doc_type_from_filename(file_name)
            
            # Perform analysis
            if file_type == 'pdf':
                text, data, confidence, page_count = self._analyze_pdf(content, file_name)
            else:
                # For images, we can pass doc_type hint to API
                if file_url:
                    # Use URL directly if available
                    result = self._call_image_reader_api(image_url=file_url, doc_type=doc_type_hint)
                    text = result.get('extracted_text', '')
                    data = result.get('extracted_data', {})
                    confidence = result.get('confidence', 0.8)
                else:
                    text, data, confidence = self._analyze_image(content)
                page_count = 1
            
            # Extract additional structured data using regex patterns
            structured_data = self._extract_structured_data_from_text(text)
            data.update(structured_data)
            
            # Advanced classification
            classification = self._classify_document_advanced(text, data)
            
            # Determine specific document type and name
            doc_type = classification['document_type']
            doc_name = self.document_name_mapping.get(doc_type, 'Unknown Document')
            
            # Quality assessment
            quality = self._assess_document_quality(text, confidence, data)
            
            # Build result
            result = {
                'document_name': doc_name,
                'document_type': doc_type,
                'category': classification['category'],
                'confidence': classification['confidence'],
                'extracted_text': text,
                'extracted_data': data,
                'page_count': page_count,
                'quality_assessment': quality,
                'processing_time': time.time() - start_time,
                'metadata': metadata.__dict__ if metadata else {},
                'validation': self._validate_document(doc_type, data),
                'suggestions': self._generate_document_suggestions(doc_type, quality, data),
                'api_used': 'InsuranceDekho Image Reader'
            }
            
            # Cache result
            if cache_key and self.cache_enabled:
                self._cache[cache_key] = result
            
            logger.info(f"Successfully analyzed document: {doc_name} (confidence: {classification['confidence']:.2f})")
            
            return result
            
        except Exception as e:
            logger.error(f"Document analysis failed: {str(e)}")
            return {
                'error': str(e),
                'document_name': 'Unknown Document',
                'document_type': DocumentType.UNKNOWN,
                'category': 'unknown',
                'confidence': 0.0,
                'extracted_text': '',
                'extracted_data': {}
            }
    
    def _guess_doc_type_from_filename(self, filename: str) -> str:
        """Guess document type from filename for API hint"""
        filename_lower = filename.lower()
        
        if 'aadhaar' in filename_lower or 'aadhar' in filename_lower:
            return 'aadhar'
        elif 'pan' in filename_lower:
            return 'pan'
        elif 'license' in filename_lower or 'dl' in filename_lower:
            return 'driving_license'
        elif 'rc' in filename_lower or 'registration' in filename_lower:
            return 'rc'
        elif 'passport' in filename_lower:
            return 'passport'
        elif 'voter' in filename_lower:
            return 'voter_id'
        
        return None  # Let API auto-detect
    def _process_single_attachment(self, attachment: Dict[str, Any]) -> Dict[str, Any]:
        """Process a single attachment and return analysis results"""
        try:
            attachment_url = attachment.get('attachment_url')
            filename = attachment.get('name', 'unknown_file')
            file_size = attachment.get('size', 0)
            
            if not attachment_url:
                return {
                    'error': 'No attachment URL provided',
                    'filename': filename,
                    'document_name': 'Unknown Document',
                    'document_type': DocumentType.UNKNOWN,
                    'category': 'error',
                    'confidence': 0.0
                }
            
            # Analyze using the attachment URL
            analysis_result = self.analyze_document(
                file_path=None,
                file_url=attachment_url,
                use_cache=True
            )
            
            # Add metadata
            analysis_result['original_filename'] = filename
            analysis_result['file_size'] = file_size
            analysis_result['attachment_url'] = attachment_url
            
            # Ensure required fields
            if 'document_name' not in analysis_result:
                analysis_result['document_name'] = 'Unknown Document'
            if 'document_type' not in analysis_result:
                analysis_result['document_type'] = DocumentType.UNKNOWN
            if 'category' not in analysis_result:
                analysis_result['category'] = 'unknown'
            if 'confidence' not in analysis_result:
                analysis_result['confidence'] = 0.0
                
            return analysis_result
            
        except Exception as e:
            return {
                'error': str(e),
                'filename': filename,
                'document_name': 'Processing Failed',
                'document_type': DocumentType.UNKNOWN,
                'category': 'error',
                'confidence': 0.0
            }

    def _read_file_content(self, file_path: str = None, file_url: str = None):
        """Read file content from local path or URL"""
        try:
            if file_path:
                with open(file_path, 'rb') as f:
                    content = f.read()
                filename = os.path.basename(file_path)
                metadata = DocumentMetadata(
                    file_name=filename,
                    file_size=len(content),
                    upload_time=datetime.now(),
                    processing_time=0.0
                )
                return content, filename, metadata
                
            elif file_url:
                response = requests.get(file_url, auth=(FRESHDESK_API_KEY, "X"), timeout=30)
                response.raise_for_status()
                content = response.content
                
                # Get filename from headers or use default
                filename = 'downloaded_file'
                content_disposition = response.headers.get('Content-Disposition', '')
                if 'filename=' in content_disposition:
                    filename = content_disposition.split('filename=')[1].strip('"')
                    
                metadata = DocumentMetadata(
                    file_name=filename,
                    file_size=len(content),
                    upload_time=datetime.now(),
                    processing_time=0.0
                )
                return content, filename, metadata
            else:
                raise ValueError("Either file_path or file_url must be provided")
        except Exception as e:
            raise

    def _detect_file_type(self, content: bytes, filename: str) -> str:
        """Detect file type from content and filename"""
        filename_lower = filename.lower()
        
        if filename_lower.endswith('.pdf'):
            return 'pdf'
        elif filename_lower.endswith(('.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.gif')):
            return 'image'
        elif content.startswith(b'%PDF'):
            return 'pdf'
        elif content.startswith((b'\xff\xd8\xff', b'\x89PNG', b'GIF8')):
            return 'image'
        
        return 'image'  # Default to image

    def _generate_cache_key(self, identifier: str) -> str:
        """Generate cache key"""
        import hashlib
        return hashlib.md5(identifier.encode()).hexdigest()

    def _extract_structured_data_from_text(self, text: str) -> Dict[str, str]:
        """Extract structured data using regex"""
        extracted_data = {}
        patterns = {
            'aadhaar_number': r'\b\d{4}\s?\d{4}\s?\d{4}\b',
            'pan_number': r'\b[A-Z]{5}[0-9]{4}[A-Z]\b',
            'mobile_number': r'\b(?:\+91[\s-]?)?[6-9]\d{9}\b',
            'policy_number': r'(?:policy|pol)[\s#:]*([A-Z0-9/-]{6,})',
            'amount': r'₹\s*(\d+(?:,\d+)*(?:\.\d+)?)'
        }
        
        for field_name, pattern in patterns.items():
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                extracted_data[field_name] = matches[0] if len(matches) == 1 else matches
        
        return extracted_data
        # Add this method to the DocumentAnalyzer class

    def _generate_recommendations(self, results: Dict[str, Any]) -> List[Dict[str, str]]:
        """
        Generate recommendations based on document analysis results
        
        Args:
            results: Dictionary containing analysis results with keys like:
                    - 'analyzed': List of analyzed documents
                    - 'missing_documents': List of missing documents
                    - 'quality_issues': List of quality issues
                    - 'statistics': Statistics about the analysis
                    - 'document_inventory': Inventory of found documents
                    - 'category_summary': Summary by category
        
        Returns:
            List of recommendation dictionaries with 'type', 'message', 'priority', 'action'
        """
        recommendations = []
        
        try:
            # Get statistics
            stats = results.get('statistics', {})
            total_processed = stats.get('total_processed', 0)
            successful = stats.get('successful', 0)
            failed = stats.get('failed', 0)
            avg_confidence = stats.get('average_confidence', 0.0)
            
            # Get other analysis results
            analyzed_docs = results.get('analyzed', [])
            missing_docs = results.get('missing_documents', [])
            quality_issues = results.get('quality_issues', [])
            document_inventory = results.get('document_inventory', {})
            category_summary = results.get('category_summary', {})
            
            # 1. Recommendations based on missing documents
            if missing_docs and len(missing_docs) > 0:
                if len(missing_docs) == 1:
                    recommendations.append({
                        'type': 'MISSING_DOCUMENT',
                        'message': f"1 document could not be processed: {missing_docs[0].get('file', 'Unknown')}",
                        'priority': 'MEDIUM',
                        'action': 'Review and resubmit the failed document with better quality'
                    })
                else:
                    recommendations.append({
                        'type': 'MISSING_DOCUMENTS',
                        'message': f"{len(missing_docs)} documents could not be processed",
                        'priority': 'HIGH',
                        'action': 'Review and resubmit failed documents with better quality'
                    })
            
            # 2. Recommendations based on quality issues
            if quality_issues and len(quality_issues) > 0:
                low_confidence_docs = [doc for doc in quality_issues if doc.get('confidence', 1.0) < 0.6]
                
                if low_confidence_docs:
                    recommendations.append({
                        'type': 'QUALITY_IMPROVEMENT',
                        'message': f"{len(low_confidence_docs)} documents have low confidence scores",
                        'priority': 'MEDIUM',
                        'action': 'Request clearer images or rescanned documents for better accuracy'
                    })
            
            # 3. Recommendations based on overall confidence
            if avg_confidence < 0.7 and successful > 0:
                recommendations.append({
                    'type': 'OVERALL_QUALITY',
                    'message': f"Overall document quality is below optimal (avg confidence: {avg_confidence:.1%})",
                    'priority': 'MEDIUM',
                    'action': 'Consider requesting higher quality document images for better processing'
                })
            
            # 4. Recommendations based on document types found
            identity_docs = []
            financial_docs = []
            insurance_docs = []
            
            for doc_name, doc_list in document_inventory.items():
                doc_name_lower = doc_name.lower()
                if any(keyword in doc_name_lower for keyword in ['aadhaar', 'pan', 'license', 'passport']):
                    identity_docs.extend(doc_list)
                elif any(keyword in doc_name_lower for keyword in ['bank', 'statement', 'cheque', 'invoice']):
                    financial_docs.extend(doc_list)
                elif any(keyword in doc_name_lower for keyword in ['policy', 'insurance', 'claim']):
                    insurance_docs.extend(doc_list)
            
            # 5. Recommendations for document completeness
            if total_processed > 0:
                if not identity_docs:
                    recommendations.append({
                        'type': 'MISSING_CATEGORY',
                        'message': 'No identity documents detected',
                        'priority': 'HIGH',
                        'action': 'Verify if identity documents (Aadhaar, PAN, etc.) are required and submitted'
                    })
                
                if not financial_docs and 'claims' in str(category_summary).lower():
                    recommendations.append({
                        'type': 'MISSING_CATEGORY',
                        'message': 'No financial documents detected for claims processing',
                        'priority': 'MEDIUM',
                        'action': 'Check if bank statements or financial documents are needed'
                    })
            
            # 6. Recommendations based on processing success rate
            if total_processed > 0:
                success_rate = successful / total_processed
                if success_rate < 0.8:
                    recommendations.append({
                        'type': 'PROCESSING_ISSUES',
                        'message': f"Low processing success rate: {success_rate:.1%}",
                        'priority': 'HIGH',
                        'action': 'Review document formats and quality. Consider manual review for failed documents'
                    })
            
            # 7. Recommendations for workflow optimization
            if successful > 5:  # Only for batches with multiple documents
                unknown_docs = [doc for doc in analyzed_docs if doc.get('document_type') == 'unknown']
                if len(unknown_docs) > 0:
                    recommendations.append({
                        'type': 'CLASSIFICATION_IMPROVEMENT',
                        'message': f"{len(unknown_docs)} documents could not be classified",
                        'priority': 'LOW',
                        'action': 'Consider manual classification for unidentified documents'
                    })
            
            # 8. Positive recommendations
            if avg_confidence > 0.9 and successful > 0:
                recommendations.append({
                    'type': 'QUALITY_EXCELLENT',
                    'message': 'Excellent document quality detected',
                    'priority': 'INFO',
                    'action': 'Documents are of high quality and ready for processing'
                })
            
            # 9. Security recommendations
            sensitive_docs = [doc for doc in analyzed_docs 
                             if any(keyword in doc.get('document_name', '').lower() 
                                   for keyword in ['aadhaar', 'pan', 'passport', 'bank'])]
            
            if sensitive_docs:
                recommendations.append({
                    'type': 'SECURITY',
                    'message': f"Sensitive documents detected ({len(sensitive_docs)} documents)",
                    'priority': 'INFO',
                    'action': 'Ensure secure handling and storage of sensitive personal information'
                })
            
            # 10. Default recommendation if no specific issues
            if not recommendations:
                recommendations.append({
                    'type': 'STATUS',
                    'message': 'Document analysis completed successfully',
                    'priority': 'INFO',
                    'action': 'No immediate action required - proceed with normal processing'
                })
            
            # Sort by priority (HIGH > MEDIUM > LOW > INFO)
            priority_order = {'HIGH': 1, 'MEDIUM': 2, 'LOW': 3, 'INFO': 4}
            recommendations.sort(key=lambda x: priority_order.get(x.get('priority', 'INFO'), 4))
            
            return recommendations
            
        except Exception as e:
            logger.error(f"Error generating recommendations: {str(e)}")
            # Return a basic recommendation in case of error
            return [{
                'type': 'ERROR',
                'message': 'Could not generate recommendations due to processing error',
                'priority': 'LOW',
                'action': 'Manual review may be required'
            }]

    def _classify_document_advanced(self, text: str, data: Dict[str, str]) -> Dict[str, Any]:
        """Advanced document classification"""
        text_lower = text.lower()
        
        if 'aadhaar_number' in data or 'aadhaar' in text_lower:
            return {
                'document_type': DocumentType.AADHAAR_CARD,
                'category': 'identity',
                'confidence': 0.9
            }
        elif 'pan_number' in data or 'permanent account' in text_lower:
            return {
                'document_type': DocumentType.PAN_CARD,
                'category': 'identity', 
                'confidence': 0.9
            }
        elif 'policy_number' in data or 'insurance policy' in text_lower:
            return {
                'document_type': DocumentType.POLICY_DOCUMENT,
                'category': 'insurance',
                'confidence': 0.8
            }
        else:
            return {
                'document_type': DocumentType.UNKNOWN,
                'category': 'unknown',
                'confidence': 0.0
            }

    def _assess_document_quality(self, text: str, confidence: float, data: Dict[str, str]) -> Dict[str, Any]:
        """Assess document quality"""
        return {
            'is_readable': len(text.strip()) > 50,
            'confidence_score': confidence,
            'overall_score': min(confidence + (len(data) * 0.1), 1.0),
            'issues': ['Low text content'] if len(text.strip()) < 50 else []
        }

    def _validate_document(self, doc_type: DocumentType, data: Dict[str, str]) -> Dict[str, Any]:
        """Validate document"""
        return {
            'is_valid': True,
            'missing_fields': [],
            'warnings': []
        }

    def _generate_document_suggestions(self, doc_type: DocumentType, quality: Dict[str, Any], data: Dict[str, str]) -> List[str]:
        """Generate suggestions"""
        suggestions = []
        if quality['overall_score'] < 0.6:
            suggestions.append("Consider uploading a clearer image")
        return suggestions
    
    def analyze_all_attachments(self, attachments: List[Dict[str, Any]], 
                               progress_callback: Optional[callable] = None) -> Dict[str, Any]:
        """
        Analyze multiple attachments with parallel processing
        """
        start_time = time.time()
        total_attachments = len(attachments)
        
        results = {
            'analyzed': [],
            'missing_documents': [],
            'quality_issues': [],
            'statistics': {
                'total_processed': 0,
                'successful': 0,
                'failed': 0,
                'average_confidence': 0.0,
                'processing_time': 0.0
            },
            'document_inventory': {},
            'category_summary': {}
        }
        
        if not attachments:
            logger.warning("No attachments provided for analysis")
            return results
        
        # Process attachments in parallel
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_attachment = {}
            
            for idx, att in enumerate(attachments):
                future = executor.submit(self._process_single_attachment, att)
                future_to_attachment[future] = (idx, att)
            
            completed = 0
            confidence_sum = 0.0
            
            for future in as_completed(future_to_attachment):
                idx, att = future_to_attachment[future]
                completed += 1
                
                if progress_callback:
                    progress_callback(completed / total_attachments * 100)
                
                try:
                    result = future.result()
                    results['analyzed'].append(result)
                    
                    # Update statistics
                    if 'error' not in result:
                        results['statistics']['successful'] += 1
                        confidence = result.get('confidence', 0)
                        confidence_sum += confidence
                        
                        # Categorize document
                        doc_name = result.get('document_name', 'Unknown')
                        category = result.get('category', 'unknown')
                        
                        # Update inventory
                        if doc_name not in results['document_inventory']:
                            results['document_inventory'][doc_name] = []
                        results['document_inventory'][doc_name].append({
                            'file_name': att.get('filename', att.get('name', 'unknown')),
                            'confidence': confidence,
                            'page_count': result.get('page_count', 1)
                        })
                        
                        # Update category summary
                        if category not in results['category_summary']:
                            results['category_summary'][category] = 0
                        results['category_summary'][category] += 1
                        
                        # Check for issues
                        if confidence < 0.5:
                            results['quality_issues'].append({
                                'document': doc_name,
                                'file': att.get('filename', 'unknown'),
                                'confidence': confidence,
                                'issue': 'Low confidence score'
                            })
                        
                        if result.get('document_type') == DocumentType.UNKNOWN:
                            results['missing_documents'].append({
                                'file': att.get('filename', 'unknown'),
                                'reason': 'Could not identify document type'
                            })
                    else:
                        results['statistics']['failed'] += 1
                        results['missing_documents'].append({
                            'file': att.get('filename', 'unknown'),
                            'reason': result.get('error', 'Processing failed')
                        })
                
                except Exception as e:
                    logger.error(f"Error processing attachment {idx}: {str(e)}")
                    results['statistics']['failed'] += 1
                
                results['statistics']['total_processed'] = completed
        
        # Calculate final statistics
        if results['statistics']['successful'] > 0:
            results['statistics']['average_confidence'] = (
                confidence_sum / results['statistics']['successful']
            )
        
        results['statistics']['processing_time'] = time.time() - start_time
        
        # Add recommendations based on analysis
        results['recommendations'] = self._generate_recommendations(results)
        
        logger.info(f"Processed {total_attachments} attachments in {results['statistics']['processing_time']:.2f}s")
        
        return results
    
    # Keep all other methods as they are (_read_file_content, _detect_file_type, 
    # _extract_structured_data_from_text, _classify_document_advanced, etc.)
    # They remain unchanged as they work with the extracted text and data

# Update the process_ticket_attachments_enhanced function to use the new API

def process_ticket_attachments_enhanced(ticket_data: dict) -> dict:
    """Enhanced attachment processing with InsuranceDekho API"""
    print("DEBUG: process_ticket_attachments_enhanced called")
    print(f"DEBUG: Input ticket_data type: {type(ticket_data)}")
    print(f"DEBUG: Input ticket_data keys: {list(ticket_data.keys()) if isinstance(ticket_data, dict) else 'Not a dict'}")
    
    try:
        # 1) Fetch attachments if they're missing
        if not ticket_data.get('attachments'):
            ticket_id = ticket_data.get('Ticket ID')
            if ticket_id:
                print(f"DEBUG: No attachments in ticket_data, fetching for ticket {ticket_id}")
                fresh_data = fetch_ticket_by_id(ticket_id)
                if fresh_data and 'attachments' in fresh_data:
                    ticket_data['attachments'] = fresh_data['attachments']
                    print(f"DEBUG: Added {len(fresh_data['attachments'])} attachments to ticket_data")

        # 2) Create analyzer with InsuranceDekho API
        print("DEBUG: Creating DocumentAnalyzer with InsuranceDekho API...")
        analyzer = DocumentAnalyzer()
        print("DEBUG: DocumentAnalyzer created successfully")

        # 3) Analyze attachments
        print("DEBUG: Calling analyze_all_attachments...")
        attachment_analysis = analyzer.analyze_all_attachments(ticket_data['attachments'])
        print(f"DEBUG: analyze_all_attachments completed, result type: {type(attachment_analysis)}")
        print(f"DEBUG: Attachment analysis keys: {list(attachment_analysis.keys()) if isinstance(attachment_analysis, dict) else 'Not a dict'}")

        # 4) Merge results back into ticket_data
        print("DEBUG: Updating ticket_data with attachment_analysis...")
        ticket_data['attachment_analysis'] = attachment_analysis

        # 5) Build insights based on analysis
        print("DEBUG: Generating document-based insights...")
        insights = []

        # Missing documents?
        if attachment_analysis.get('missing_documents'):
            missing_docs_list = [
                doc.get('file', 'unknown') for doc in attachment_analysis['missing_documents']
                if isinstance(doc, dict)
            ]
            print(f"DEBUG: Found {len(missing_docs_list)} missing documents")
            insights.append({
                'type': 'MISSING_DOCUMENTS',
                'priority': 'HIGH',
                'message': f"Missing required documents: {', '.join(missing_docs_list)}",
                'action': 'Request missing documents from customer'
            })

        # Quality issues?
        if attachment_analysis.get('quality_issues'):
            print(f"DEBUG: Found {len(attachment_analysis['quality_issues'])} quality issues")
            insights.append({
                'type': 'QUALITY_ISSUES',
                'priority': 'MEDIUM',
                'message': f"{len(attachment_analysis['quality_issues'])} documents have quality issues",
                'action': 'Request clearer copies of affected documents'
            })

        # 6) Extract & validate structured data
        print("DEBUG: Extracting key information...")
        extracted_info = {}
        for doc in attachment_analysis.get('analyzed', []):
            if doc.get('extracted_data'):
                extracted_info.update(doc['extracted_data'])
        print(f"DEBUG: Extracted {len(extracted_info)} pieces of information")

        print("DEBUG: Validating extracted information...")
        validation_results = validate_extracted_info(extracted_info, ticket_data)
        if validation_results.get('mismatches'):
            print(f"DEBUG: Found {len(validation_results['mismatches'])} data mismatches")
            insights.append({
                'type': 'DATA_MISMATCH',
                'priority': 'HIGH',
                'message': "Document data doesn't match ticket information",
                'details': validation_results['mismatches']
            })

        # 7) Attach insights & extracted data back to ticket
        print(f"DEBUG: Generated {len(insights)} insights")
        ticket_data['document_insights'] = insights
        ticket_data['extracted_document_data'] = extracted_info

        print("DEBUG: process_ticket_attachments_enhanced completed successfully")
        return ticket_data

    except Exception as e:
        print(f"DEBUG: Exception in process_ticket_attachments_enhanced: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()

        ticket_data['attachment_analysis'] = {
            'error': str(e),
            'message': 'Document analysis failed'
        }
        return ticket_data

# Update analyze_downloaded_attachments to use the new API

def analyze_downloaded_attachments(ticket_id):
    """
    Download and analyze all attachments for a ticket using InsuranceDekho API
    """
    # Download attachments to temp directory
    temp_dir = f"temp_attachments_{ticket_id}"
    downloaded = download_all_ticket_attachments(ticket_id, temp_dir)
    
    # Initialize analyzer with InsuranceDekho API
    analyzer = DocumentAnalyzer()
    
    # Analyze each downloaded file
    results = []
    for attachment in downloaded:
        if 'saved_to' in attachment:
            # Analyze the saved file
            analysis = analyzer.analyze_document(
                attachment['saved_to'], 
                None  # No URL needed since we have local file
            )
            analysis['filename'] = attachment['original_name']
            analysis['attachment_id'] = attachment['attachment_id']
            results.append(analysis)
    
    # Clean up temp files
    import shutil
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
    
    return results
    
class DocumentRequirementEngine:
    """Engine to determine document requirements based on insurer and claim type"""
    
    def __init__(self):
        self.document_rules = self._load_document_rules()
        self.insurer_mapping = self._load_insurer_mapping()
    
    def _load_document_rules(self) -> Dict:
        """Load document requirement rules from your data"""
        return {
            'CASHLESS_GARAGE_REQUIRED': {
                'Policy Doc': {'required': 1, 'insurers': [0]},
                'Rc Page 1': {'required': 1, 'insurers': [0]},
                'Drvng Linc 1': {'required': 1, 'insurers': [0]},
                'Invoice Copy': {'required': 0, 'insurers': [0], 'invoice_required': 1},
                'Aadhar Page 1': {'required': 1, 'insurers': [0], 'kyc_required': 1},
                'Pan': {'required': 0, 'insurers': [1,2,3,4,5,6,7,8,9,10,11,12,15,16,17,20,21,24,26,27,28,22,25,31,32,30], 'kyc_required': 1},
                'Aadhar Page 2': {'required': 0, 'insurers': [0], 'kyc_required': 1},
                'Drvng Linc 2': {'required': 0, 'insurers': [0]},
                'Rc Page 2': {'required': 0, 'insurers': [0]},
            },
            'CLAIM_INTIMATION': {
                'Policy Doc': {'required': 1, 'insurers': [0]},
                'Rc Page 1': {'required': 1, 'insurers': [0]},
                'Drvng Linc 1': {'required': 1, 'insurers': [0]},
                'Invoice Copy': {'required': 0, 'insurers': [0], 'invoice_required': 1},
                'Aadhar Page 1': {'required': 1, 'insurers': [0], 'kyc_required': 1},
                'Pan': {
                    'required': 0, 
                    'insurers': [1,2,3,4,5,6,7,8,9,10,11,12,15,16,17,20,21,24,26,27,28,22,25,31,32,30], 
                    'kyc_required': 1,
                    'special_insurers': {
                        14: {'required': 1},  # New India
                        18: {'required': 1},  # United India
                        19: {'required': 1},  # Not in mapping, needs verification
                        23: {'required': 1}   # Not in mapping, needs verification
                    }
                },
                'Repair Estimates': {'required': 1, 'insurers': [14,18,19,23]},
                'Claim Form': {'required': 1, 'insurers': [14,18,19,23]},
                'Aadhar Page 2': {'required': 0, 'insurers': [0], 'kyc_required': 1},
                'Drvng Linc 2': {'required': 0, 'insurers': [0]},
                'Rc Page 2': {'required': 0, 'insurers': [0]},
                'Pan_special': {'required': 1, 'insurers': [14,18,19,23], 'kyc_required': 1}
            },
            'SURVEY_PENDING': {
                'Policy Doc': {'required': 1, 'insurers': [0]},
                'Rc Page 1': {'required': 1, 'insurers': [0]},
                'Drvng Linc 1': {'required': 1, 'insurers': [0]},
                'Invoice Copy': {'required': 0, 'insurers': [0], 'invoice_required': 1},
                'Aadhar Page 1': {'required': 1, 'insurers': [0], 'kyc_required': 1},
                'Pan': {'required': 0, 'insurers': [1,2,3,4,5,6,7,8,9,10,11,12,15,16,17,20,21,24,26,27,28,22,25,31,32,30], 'kyc_required': 1},
                'Aadhar Page 2': {'required': 0, 'insurers': [0], 'kyc_required': 1},
                'Drvng Linc 2': {'required': 0, 'insurers': [0]},
                'Rc Page 2': {'required': 0, 'insurers': [0]},
            },
            'DELIVERY_ORDER_PENDING': {
                'Policy Doc': {'required': 1, 'insurers': [0]},
                'Rc Page 1': {'required': 1, 'insurers': [0]},
                'Drvng Linc 1': {'required': 1, 'insurers': [0]},
                'Invoice Copy': {'required': 0, 'insurers': [0], 'invoice_required': 1},
                'Aadhar Page 1': {'required': 1, 'insurers': [0], 'kyc_required': 1},
                'Pan': {'required': 0, 'insurers': [1,2,3,4,5,6,7,8,9,10,11,12,15,16,17,20,21,24,26,27,28,22,25,31,32,30], 'kyc_required': 1},
                'Aadhar Page 2': {'required': 0, 'insurers': [0], 'kyc_required': 1},
                'Drvng Linc 2': {'required': 0, 'insurers': [0]},
                'Rc Page 2': {'required': 0, 'insurers': [0]},
            },
            'REIMBURSEMENT_PENDING': {
                'Policy Doc': {'required': 1, 'insurers': [0]},
                'Rc Page 1': {'required': 1, 'insurers': [0]},
                'Drvng Linc 1': {'required': 1, 'insurers': [0]},
                'Invoice Copy': {'required': 0, 'insurers': [0], 'invoice_required': 1},
                'Aadhar Page 1': {'required': 1, 'insurers': [0], 'kyc_required': 1},
                'Pan': {'required': 0, 'insurers': [1,2,3,4,5,6,7,8,9,10,11,12,15,16,17,20,21,24,26,27,28,22,25,31,32,30], 'kyc_required': 1},
                'Aadhar Page 2': {'required': 0, 'insurers': [0], 'kyc_required': 1},
                'Drvng Linc 2': {'required': 0, 'insurers': [0]},
                'Rc Page 2': {'required': 0, 'insurers': [0]},
            },
            'RAISE_QUERY_ON_THE_SETTLED_CLAIM_AMOUNT': {
                'Policy Doc': {'required': 1, 'insurers': [0]},
                'Rc Page 1': {'required': 1, 'insurers': [0]},
                'Drvng Linc 1': {'required': 1, 'insurers': [0]},
                'Invoice Copy': {'required': 0, 'insurers': [0], 'invoice_required': 1},
                'Aadhar Page 1': {'required': 1, 'insurers': [0], 'kyc_required': 1},
                'Pan': {'required': 0, 'insurers': [1,2,3,4,5,6,7,8,9,10,11,12,15,16,17,20,21,24,26,27,28,22,25,31,32,30], 'kyc_required': 1},
                'Aadhar Page 2': {'required': 0, 'insurers': [0], 'kyc_required': 1},
                'Drvng Linc 2': {'required': 0, 'insurers': [0]},
                'Rc Page 2': {'required': 0, 'insurers': [0]},
            },
            'WORK_APPROVAL_PENDING': {
                'Policy Doc': {'required': 1, 'insurers': [0]},
                'Rc Page 1': {'required': 1, 'insurers': [0]},
                'Drvng Linc 1': {'required': 1, 'insurers': [0]},
                'Drvng Linc 2': {'required': 0, 'insurers': [0]},
                'Rc Page 2': {'required': 0, 'insurers': [0]},
            }
        }
    
    def _load_insurer_mapping(self) -> Dict:
        """Load insurer ID to name mapping"""
        return {
            1: "HDFC Ergo",
            6: "Royal Sundaram",
            10: "Bajaj Allianz",
            18: "United India Insurance Company Limited",
            2: "Reliance General Insurance",
            14: "New India",
            4: "ICICI Lombard",
            5: "Kotak Mahindra General Insurance Limited",
            28: "Edelweiss General Insurance",
            20: "Liberty General Insurance Company Limited",
            9: "Future Generali",
            22: "Magma HDI General Insurance Company Ltd",
            16: "Shriram General Insurance",
            12: "Digit",
            11: "Universal Sompo",
            21: "Cholamandalam MS General Insurance Company Ltd",
            15: "SBI"
        }
    
    def get_required_documents(self, claim_type: str, insurer_id: int) -> List[Dict]:
        """Get list of required documents for given claim type and insurer"""
        required_docs = []
        optional_docs = []
        
        if claim_type not in self.document_rules:
            # Default to CLAIM_INTIMATION if type not found
            claim_type = 'CLAIM_INTIMATION'
        
        rules = self.document_rules[claim_type]
        
        for doc_name, doc_rules in rules.items():
            # Skip special entries
            if '_special' in doc_name:
                continue
                
            is_required = False
            is_optional = False
            
            # Check base requirement
            base_required = doc_rules.get('required', 0)
            
            # Check if this insurer is in the insurers list
            if insurer_id in doc_rules.get('insurers', []):
                if base_required == 1:
                    is_required = True
                else:
                    # Special case: document is generally optional but required for this insurer
                    is_required = True
            elif base_required == 1 and 0 in doc_rules.get('insurers', []):
                # Document is required for all insurers (0 means all)
                is_required = True
            
            # Check special insurer rules (like PAN for specific insurers in CLAIM_INTIMATION)
            if claim_type == 'CLAIM_INTIMATION' and doc_name == 'Pan':
                if insurer_id in [14, 18, 19, 23]:
                    is_required = True
                elif insurer_id in doc_rules.get('insurers', []):
                    is_optional = True
            
            doc_info = {
                'name': doc_name,
                'type': self._get_document_type(doc_name),
                'description': self._get_document_description(doc_name),
                'kyc_document': doc_rules.get('kyc_required', 0) == 1,
                'invoice_document': doc_rules.get('invoice_required', 0) == 1
            }
            
            if is_required:
                required_docs.append(doc_info)
            elif is_optional:
                optional_docs.append(doc_info)
        
        return required_docs, optional_docs
    
    def _get_document_type(self, doc_name: str) -> str:
        """Map document name to type"""
        mapping = {
            'Policy Doc': 'policy',
            'Rc Page 1': 'vehicle_registration',
            'Rc Page 2': 'vehicle_registration',
            'Drvng Linc 1': 'driving_license',
            'Drvng Linc 2': 'driving_license',
            'Invoice Copy': 'invoice',
            'Aadhar Page 1': 'identity_proof',
            'Aadhar Page 2': 'identity_proof',
            'Pan': 'identity_proof',
            'Repair Estimates': 'claim_document',
            'Claim Form': 'claim_document'
        }
        return mapping.get(doc_name, 'other')
    
    def _get_document_description(self, doc_name: str) -> str:
        """Get user-friendly description for document"""
        descriptions = {
            'Policy Doc': 'Insurance Policy Document',
            'Rc Page 1': 'Vehicle Registration Certificate (RC) - Front Page',
            'Rc Page 2': 'Vehicle Registration Certificate (RC) - Back Page (if available)',
            'Drvng Linc 1': 'Driving License - Front Side',
            'Drvng Linc 2': 'Driving License - Back Side (if available)',
            'Invoice Copy': 'Purchase Invoice or Bill',
            'Aadhar Page 1': 'Aadhaar Card - Front Side',
            'Aadhar Page 2': 'Aadhaar Card - Back Side (if available)',
            'Pan': 'PAN Card',
            'Repair Estimates': 'Repair Cost Estimates from Garage',
            'Claim Form': 'Duly Filled and Signed Claim Form'
        }
        return descriptions.get(doc_name, doc_name)

    
# Enhanced Contextual Routing Analyzer with Full Ticket Understanding
class EnhancedContextualRoutingAnalyzer:
    """
    Comprehensive analyzer that understands ticket routing, current actions,
    parent-child relationships, and generates intelligent responses.
    """
    
    def __init__(self, freshdesk_domain, freshdesk_api_key):
        self.freshdesk_domain = freshdesk_domain
        self.freshdesk_api_key = freshdesk_api_key
        self.status_mappings = self._initialize_status_mappings()
        self.routing_patterns = self._initialize_routing_patterns()
        self.action_patterns = self._initialize_action_patterns()
        
    def _initialize_status_mappings(self):
        """Initialize Freshdesk status mappings with routing and action context"""
        return {
            2: {
                'name': 'Open',
                'display': 'Being Processed',
                'routing': None,
                'action': 'Initial processing',
                'category': 'active'
            },
            3: {
                'name': 'Pending',
                'display': 'Awaiting your Reply',
                'routing': 'customer',
                'action': 'Waiting for customer response',
                'category': 'waiting'
            },
            4: {
                'name': 'Resolved',
                'display': 'This ticket has been Resolved',
                'routing': None,
                'action': 'Solution provided',
                'category': 'complete'
            },
            5: {
                'name': 'Closed',
                'display': 'This ticket has been Closed',
                'routing': None,
                'action': 'Ticket closed',
                'category': 'complete'
            },
            10: {
                'name': 'Parent-Pending/BM-Dealer',
                'display': 'Parent- Pending / BM-Dealer',
                'routing': 'dealer',
                'action': 'Awaiting dealer/channel partner action',
                'category': 'waiting'
            },
            11: {
                'name': 'Parent-Pending/Insurer',
                'display': 'Parent- Pending / Insurer',
                'routing': 'insurer',
                'action': 'Awaiting insurer response',
                'category': 'waiting'
            },
            12: {
                'name': 'Parent-Pending/Customer',
                'display': 'Parent- Pending /Customer',
                'routing': 'customer',
                'action': 'Awaiting customer input',
                'category': 'waiting'
            },
            13: {
                'name': 'Child-Pending/BM-Dealer',
                'display': 'Child- Pending / BM-Dealer',
                'routing': 'dealer',
                'action': 'Child ticket with dealer',
                'category': 'waiting'
            },
            14: {
                'name': 'Child-Pending/Insurer',
                'display': 'Child- Pending / Insurer',
                'routing': 'insurer',
                'action': 'Child ticket with insurer',
                'category': 'waiting'
            },
            15: {
                'name': 'Child-Pending/Customer',
                'display': 'Child- Pending / Customer',
                'routing': 'customer',
                'action': 'Child ticket awaiting customer',
                'category': 'waiting'
            },
            16: {
                'name': 'CODERED-WIP',
                'display': 'CODERED - Work in Progress',
                'routing': 'internal',
                'action': 'High priority - actively working',
                'category': 'critical'
            },
            17: {
                'name': 'CODERED-PENDING-SPOC',
                'display': 'CODERED - PENDING WITH SPOC',
                'routing': 'internal_spoc',
                'action': 'Critical escalation with SPOC',
                'category': 'critical'
            },
            18: {
                'name': 'CODERED-Pending-Partner',
                'display': 'CODERED - Pending with Partner',
                'routing': 'partner',
                'action': 'Critical issue with partner',
                'category': 'critical'
            },
            21: {
                'name': 'Customer-Responded',
                'display': 'Customer Responded',
                'routing': 'internal',
                'action': 'Processing customer response',
                'category': 'active'
            },
            22: {
                'name': 'CODERED-NOC-HOLD',
                'display': 'CODERED NOC HOLD',
                'routing': 'noc',
                'action': 'On hold for NOC clearance',
                'category': 'critical'
            },
            23: {
                'name': 'KYC-Pending-Requester',
                'display': 'KYC Pending with Requester',
                'routing': 'customer',
                'action': 'KYC documents needed from customer',
                'category': 'compliance'
            },
            24: {
                'name': 'KYC-WIP',
                'display': 'KYC WIP',
                'routing': 'internal',
                'action': 'KYC verification in progress',
                'category': 'compliance'
            },
            25: {
                'name': 'Pending-Tech-Support',
                'display': 'Pending with Tech Support',
                'routing': 'tech_support',
                'action': 'Technical issue resolution',
                'category': 'technical'
            },
            26: {
                'name': 'Reopen',
                'display': 'Reopen',
                'routing': 'internal',
                'action': 'Ticket reopened for review',
                'category': 'active'
            }
        }
    
    def _initialize_routing_patterns(self):
        """Initialize comprehensive routing detection patterns"""
        return {
            'insurer': {
                'keywords': [
                    'forwarded to insurer', 'raised to insurer', 'submitted to insurance',
                    'claim intimated', 'policy servicing team', 'underwriting',
                    'insurer portal', 'carrier', 'insurance company'
                ],
                'entities': [
                    'hdfc ergo', 'icici lombard', 'bajaj allianz', 'tata aig',
                    'reliance', 'sbi general', 'kotak', 'max life', 'star health',
                    'united india', 'national insurance', 'new india', 'oriental',
                    'iffco tokio', 'digit', 'acko', 'go digit'
                ],
                'patterns': [
                    r'claim\s+(?:no|number|id)[\s:]*([A-Z0-9/-]+)',
                    r'(?:forwarded?|sent?|raised?)\s+to\s+(\w+)\s+(?:insurance|insurer)',
                    r'child\s+ticket.*?insurer'
                ]
            },
            'customer': {
                'keywords': [
                    'waiting for customer', 'customer to provide', 'pending from customer',
                    'requested documents', 'customer response', 'requester pending',
                    'awaiting clarification', 'customer action required'
                ],
                'patterns': [
                    r'pending\s+(?:from|with)\s+customer',
                    r'waiting\s+for\s+(?:customer|requester)',
                    r'documents?\s+(?:pending|required)\s+from\s+customer'
                ]
            },
            'dealer': {
                'keywords': [
                    'channel partner', 'dealer', 'agent', 'pos', 'broker',
                    'intermediary', 'franchise', 'branch', 'bm dealer'
                ],
                'patterns': [
                    r'(?:forwarded?|sent?)\s+to\s+(?:dealer|partner|agent)',
                    r'pending\s+(?:from|with)\s+(?:dealer|partner)',
                    r'channel\s+partner\s+(?:action|response)'
                ]
            },
            'inspection': {
                'keywords': [
                    'inspection pending', 'surveyor', 'survey report', 'inspection scheduled',
                    'vehicle inspection', 'pre-inspection', 'post inspection',
                    'inspection completed', 'inspection agency'
                ],
                'patterns': [
                    r'inspection\s+(?:scheduled|pending|completed)',
                    r'surveyor\s+(?:appointed|assigned|visited)',
                    r'(?:pre|post)\s*-?\s*inspection'
                ]
            },
            'tech_support': {
                'keywords': [
                    'technical issue', 'system error', 'portal issue', 'api failure',
                    'payment gateway', 'tech team', 'it support', 'system down'
                ],
                'patterns': [
                    r'technical\s+(?:issue|problem|error)',
                    r'system\s+(?:error|down|failure)',
                    r'(?:api|portal|gateway)\s+(?:issue|error|failure)'
                ]
            },
            'internal': {
                'keywords': [
                    'internal review', 'processing', 'verification', 'quality check',
                    'supervisor review', 'escalation', 'team lead', 'manager'
                ],
                'patterns': [
                    r'(?:internal|team)\s+(?:review|processing)',
                    r'escalated?\s+to\s+(?:supervisor|manager|team lead)'
                ]
            }
        }
    
    def _initialize_action_patterns(self):
        """Initialize patterns to understand what's happening in the ticket"""
        return {
            'document_collection': {
                'patterns': [
                    r'documents?\s+(?:received|submitted|uploaded)',
                    r'(?:kyc|claim form|invoice|rc|policy)\s+(?:received|pending)',
                    r'waiting\s+for\s+documents?'
                ],
                'status': 'Collecting required documents'
            },
            'claim_processing': {
                'patterns': [
                    r'claim\s+(?:initiated|submitted|processing|under review)',
                    r'claim\s+(?:no|number|id)[\s:]*[A-Z0-9/-]+',
                    r'surveyor\s+(?:appointed|report\s+pending)'
                ],
                'status': 'Claim being processed'
            },
            'verification': {
                'patterns': [
                    r'(?:kyc|document|policy)\s+verification',
                    r'under\s+(?:review|verification|validation)',
                    r'(?:checking|verifying)\s+details'
                ],
                'status': 'Verification in progress'
            },
            'escalation': {
                'patterns': [
                    r'escalated?\s+to',
                    r'(?:high|critical)\s+priority',
                    r'codered|urgent|immediate\s+attention'
                ],
                'status': 'Escalated for priority handling'
            },
            'resolution': {
                'patterns': [
                    r'(?:issue|query|problem)\s+(?:resolved|fixed|addressed)',
                    r'solution\s+provided',
                    r'claim\s+(?:approved|settled|paid)'
                ],
                'status': 'Resolution provided'
            },
            'pending_response': {
                'patterns': [
                    r'waiting\s+for\s+(?:response|reply|revert)',
                    r'pending\s+(?:from|with)',
                    r'awaiting\s+(?:response|clarification|documents)'
                ],
                'status': 'Awaiting response'
            }
        }
    
    def analyze_complete_ticket_context(self, ticket_id):
        """
        Analyze complete ticket including parent-child relationships
        Returns comprehensive understanding of ticket state
        """
        # Fetch main ticket
        ticket_data = fetch_ticket_by_id(ticket_id)
        if not ticket_data:
            return None
        
        # Get all conversations
        conversations = fetch_all_ticket_conversations(ticket_id)
        
        # Analyze parent-child relationships
        parent_child_analysis = self._analyze_parent_child_tickets(ticket_data)
        
        # Extract routing history
        routing_history = self._extract_routing_history(ticket_data, conversations)
        
        # Understand current action
        current_action = self._understand_current_action(ticket_data, conversations)
        
        # Analyze status progression
        status_timeline = self._analyze_status_progression(conversations)
        
        # Build UI-friendly analysis sections
        primary_intent = self._determine_primary_routing(ticket_data, routing_history)
        complete_context = {
            'ticket_id': ticket_id,
            'current_status': self._get_status_details(ticket_data.get('status')),
            'routing_intent': primary_intent,
            'routing_history': routing_history,
            'current_action_raw': current_action,
            'routing_analysis': {
                'primary_intent': primary_intent,
                'current_routing': (routing_history[-1]['routed_to']
                                    if routing_history else primary_intent),
                'routing_history': routing_history
            },
            'action_analysis': {
                'current_action': current_action.get('primary_action', 'N/A'),
                'next_steps': current_action.get('next_steps', []),
                'details': current_action.get('details', [])
            },
            'parent_child_analysis': parent_child_analysis,
            'status_timeline': status_timeline,
            'ticket_age_hours': self._calculate_ticket_age(ticket_data.get('created_at')),
            'last_update_hours': self._calculate_last_update(ticket_data.get('updated_at')),
            'priority_indicators': self._identify_priority_indicators(ticket_data, conversations),
            'pending_items': self._extract_pending_items(conversations),
            'key_entities': self._extract_key_entities(ticket_data, conversations)
        }
        
        return complete_context

    
    def _analyze_parent_child_tickets(self, ticket_data):
        """Analyze parent-child ticket relationships with actual ticket fetching"""
        analysis = {
            'is_parent': False,
            'is_child': False,
            'child_tickets': [],
            'parent_ticket': None,
            'child_ticket_summary': {},
            'consolidated_routing': None,
            'child_actions': []
        }

        # Check if this is a parent ticket
        status = ticket_data.get('status')
        ticket_id = ticket_data.get('id')

        if status in [10, 11, 12]:  # Parent ticket statuses
            analysis['is_parent'] = True

            # Fetch actual child tickets
            child_tickets = fetch_child_tickets(ticket_id)
            analysis['child_tickets'] = child_tickets

            # Analyze each child ticket
            for child in child_tickets:
                child_id = child.get('id')
                child_status = child.get('status')

                # Get child conversations for detailed analysis
                child_conversations = fetch_all_ticket_conversations(child_id)

                # Extract routing from child
                child_routing = self._extract_routing_history(child, child_conversations)

                # Summarize child ticket
                analysis['child_ticket_summary'][child_id] = {
                    'subject': child.get('subject'),
                    'status': self._get_status_details(child_status),
                    'routing': child_routing,
                    'created_at': child.get('created_at'),
                    'updated_at': child.get('updated_at')
                }

                # Consolidate routing information
                if child_routing:
                    latest_route = child_routing[-1]
                    analysis['child_actions'].append({
                        'ticket_id': child_id,
                        'action': latest_route.get('reason'),
                        'routed_to': latest_route.get('routed_to'),
                        'timestamp': latest_route.get('timestamp')
                    })
            
            # Determine consolidated routing based on child tickets
            if analysis['child_actions']:
                # Get the most recent routing across all child tickets
                latest_action = max(analysis['child_actions'],
                                   key=lambda x: x.get('timestamp', ''))
                analysis['consolidated_routing'] = latest_action.get('routed_to')

        # Check if this is a child ticket
        elif status in [13, 14, 15]:  # Child ticket statuses
            analysis['is_child'] = True

            # Try to find parent ticket
            parent_ticket = fetch_parent_ticket(ticket_data)
            if parent_ticket:
                analysis['parent_ticket'] = parent_ticket.get('id')
                analysis['parent_subject'] = parent_ticket.get('subject')
                analysis['parent_status'] = self._get_status_details(parent_ticket.get('status'))
                
        return analysis
    
    def _extract_routing_history(self, ticket_data, conversations):
        """Extract complete routing history from ticket lifecycle"""
        routing_events = []
        
        # Check current status routing
        current_status = ticket_data.get('status')
        if current_status in self.status_mappings:
            status_info = self.status_mappings[current_status]
            if status_info['routing']:
                routing_events.append({
                    'timestamp': ticket_data.get('updated_at'),
                    'routed_to': status_info['routing'],
                    'reason': status_info['action'],
                    'source': 'status'
                })
        
        # Analyze conversations for routing events
        for conv in conversations:
            conv_text = conv.get('body_text', '') or clean_html(conv.get('body', ''))
            conv_lower = conv_text.lower()
            
            for route_type, route_config in self.routing_patterns.items():
                # Check keywords
                if any(keyword in conv_lower for keyword in route_config['keywords']):
                    routing_events.append({
                        'timestamp': conv.get('created_at'),
                        'routed_to': route_type,
                        'reason': self._extract_routing_reason(conv_text, route_type),
                        'source': 'conversation',
                        'actor': conv.get('from_email', 'System')
                    })
                    
                # Check patterns
                for pattern in route_config.get('patterns', []):
                    if re.search(pattern, conv_text, re.IGNORECASE):
                        routing_events.append({
                            'timestamp': conv.get('created_at'),
                            'routed_to': route_type,
                            'reason': self._extract_routing_reason(conv_text, route_type),
                            'source': 'conversation',
                            'actor': conv.get('from_email', 'System')
                        })
        
        # Sort by timestamp
        routing_events.sort(key=lambda x: x['timestamp'] or '')
        
        return routing_events
    
    def _understand_current_action(self, ticket_data, conversations):
        """Understand what's currently happening with the ticket"""
        current_action = {
            'primary_action': 'Unknown',
            'details': [],
            'blockers': [],
            'next_steps': []
        }
        
        # Get latest conversation
        if conversations:
            latest_conv = conversations[-1]
            conv_text = latest_conv.get('body_text', '') or clean_html(latest_conv.get('body', ''))
            
            # Check action patterns
            for action_type, config in self.action_patterns.items():
                for pattern in config['patterns']:
                    if re.search(pattern, conv_text, re.IGNORECASE):
                        current_action['primary_action'] = config['status']
                        current_action['details'].append({
                            'action': action_type,
                            'evidence': re.search(pattern, conv_text, re.IGNORECASE).group(0)
                        })
        
        # Check status-based action
        status_info = self.status_mappings.get(ticket_data.get('status'), {})
        if status_info.get('action'):
            current_action['primary_action'] = status_info['action']
        
        # Identify blockers
        if 'pending' in current_action['primary_action'].lower():
            current_action['blockers'].append({
                'type': 'waiting',
                'description': current_action['primary_action']
            })
        
        return current_action
    
    def _determine_primary_routing(self, ticket_data, routing_history):
        """Determine the primary routing intent"""
        if not routing_history:
            return 'internal'
        
        # Get the most recent routing
        latest_routing = routing_history[-1]
        
        # Check if it's a critical status
        status = ticket_data.get('status')
        if status in [16, 17, 18, 22]:  # CODERED statuses
            return 'critical_escalation'
        
        return latest_routing['routed_to']
    
    def _get_status_details(self, status_id):
        """Get detailed status information"""
        if status_id in self.status_mappings:
            return self.status_mappings[status_id]
        return {
            'name': 'Unknown',
            'display': 'Unknown Status',
            'routing': None,
            'action': 'Status not mapped',
            'category': 'unknown'
        }
    
    def _calculate_ticket_age(self, created_at):
        """Calculate ticket age in hours"""
        if not created_at:
            return 0
        try:
            created_time = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            return (datetime.now(created_time.tzinfo) - created_time).total_seconds() / 3600
        except:
            return 0
    
    def _calculate_last_update(self, updated_at):
        """Calculate hours since last update"""
        if not updated_at:
            return 0
        try:
            updated_time = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
            return (datetime.now(updated_time.tzinfo) - updated_time).total_seconds() / 3600
        except:
            return 0
    
    def _identify_priority_indicators(self, ticket_data, conversations):
        """Identify priority indicators in the ticket"""
        indicators = []
        
        # Check for CODERED status
        if ticket_data.get('status') in [16, 17, 18, 22]:
            indicators.append({
                'type': 'critical_status',
                'description': 'CODERED priority ticket'
            })
        
        # Check for escalation keywords
        content = ticket_data.get('description', '') + ' '.join([
            conv.get('body_text', '') for conv in conversations
        ])
        
        priority_keywords = ['urgent', 'critical', 'escalate', 'asap', 'immediately', 
                           'high priority', 'emergency', 'codered']
        
        for keyword in priority_keywords:
            if keyword in content.lower():
                indicators.append({
                    'type': 'keyword',
                    'description': f'Contains priority keyword: {keyword}'
                })
        
        # Check ticket age
        age_hours = self._calculate_ticket_age(ticket_data.get('created_at'))
        if age_hours > 48:
            indicators.append({
                'type': 'age',
                'description': f'Ticket open for {age_hours:.1f} hours'
            })
        
        return indicators
    
    def _extract_pending_items(self, conversations):
        """Extract what's currently pending"""
        pending_items = []
        
        for conv in conversations:
            text = conv.get('body_text', '') or clean_html(conv.get('body', ''))
            
            # Patterns for pending items
            patterns = [
                r'pending\s+(?:for|from)\s+([^.,]+)',
                r'waiting\s+for\s+([^.,]+)',
                r'require[sd]?\s+([^.,]+)',
                r'need\s+([^.,]+)\s+from'
            ]
            
            for pattern in patterns:
                matches = re.findall(pattern, text, re.IGNORECASE)
                for match in matches:
                    if len(match) < 100:  # Reasonable length
                        pending_items.append({
                            'item': match.strip(),
                            'source': conv.get('from_email', 'Unknown'),
                            'timestamp': conv.get('created_at')
                        })
        
        return pending_items
    
    def _extract_key_entities(self, ticket_data, conversations):
        """Extract key entities mentioned in the ticket"""
        entities = {
            'claim_numbers': set(),
            'policy_numbers': set(),
            'insurer_names': set(),
            'amounts': set(),
            'dates': set(),
            'people': set()
        }
        
        # Combine all text
        all_text = ticket_data.get('subject', '') + ' ' + \
                   ticket_data.get('description', '') + ' ' + \
                   ' '.join([conv.get('body_text', '') for conv in conversations])
        
        # Extract patterns
        import re
        
        # Claim numbers
        claim_patterns = re.findall(r'claim\s*#?\s*([A-Z0-9/-]+)', all_text, re.IGNORECASE)
        entities['claim_numbers'].update(claim_patterns)
        
        # Policy numbers
        policy_patterns = re.findall(r'policy\s*#?\s*([A-Z0-9/-]+)', all_text, re.IGNORECASE)
        entities['policy_numbers'].update(policy_patterns)
        
        # Amounts
        amount_patterns = re.findall(r'₹\s*(\d+(?:,\d+)*(?:\.\d+)?)', all_text)
        entities['amounts'].update(amount_patterns)
        
        # Dates
        date_patterns = re.findall(r'\b\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\b', all_text)
        entities['dates'].update(date_patterns)
        
        # Insurer names
        for insurer in self.routing_patterns['insurer']['entities']:
            if insurer in all_text.lower():
                entities['insurer_names'].add(insurer.title())
        
        # Convert sets to lists for JSON serialization
        return {k: list(v) for k, v in entities.items()}
    
    def _extract_routing_reason(self, text, route_type):
        """Extract the reason for routing"""
        # Simple extraction - can be enhanced
        sentences = text.split('.')
        for sentence in sentences:
            if route_type in sentence.lower():
                return sentence.strip()
        return f"Routed to {route_type}"
    
    def _analyze_status_progression(self, conversations):
        """Analyze how status has progressed over time"""
        status_changes = []
        
        for conv in conversations:
            # Look for status change indicators in conversation
            text = conv.get('body_text', '') or clean_html(conv.get('body', ''))
            
            status_patterns = [
                r'status\s+changed?\s+to\s+([^.,]+)',
                r'marked?\s+as\s+([^.,]+)',
                r'updated?\s+to\s+([^.,]+)'
            ]
            
            for pattern in status_patterns:
                matches = re.findall(pattern, text, re.IGNORECASE)
                for match in matches:
                    status_changes.append({
                        'timestamp': conv.get('created_at'),
                        'new_status': match.strip(),
                        'actor': conv.get('from_email', 'System')
                    })
        
        return status_changes

# Enhanced Child Ticket Reading and Pending Status Detection

class EnhancedTicketAnalyzer:
    """Enhanced analyzer for reading child tickets and determining pending status"""
    
    def __init__(self, freshdesk_domain, freshdesk_api_key):
        self.freshdesk_domain = freshdesk_domain
        self.freshdesk_api_key = freshdesk_api_key
        self.status_mappings = self._initialize_enhanced_status_mappings()
    
    def _initialize_enhanced_status_mappings(self):
        """Enhanced status mappings with pending from information"""
        return {
            2: {
                'name': 'Open',
                'display': 'Being Processed',
                'pending_from': 'internal_team',
                'action': 'Initial processing',
                'category': 'active'
            },
            3: {
                'name': 'Pending',
                'display': 'Awaiting your Reply',
                'pending_from': 'customer',
                'action': 'Waiting for customer response',
                'category': 'waiting'
            },
            4: {
                'name': 'Resolved',
                'display': 'This ticket has been Resolved',
                'pending_from': 'none',
                'action': 'Solution provided',
                'category': 'complete'
            },
            5: {
                'name': 'Closed',
                'display': 'This ticket has been Closed',
                'pending_from': 'none',
                'action': 'Ticket closed',
                'category': 'complete'
            },
            10: {
                'name': 'Parent-Pending/BM-Dealer',
                'display': 'Parent- Pending / BM-Dealer',
                'pending_from': 'dealer',
                'action': 'Awaiting dealer/channel partner action',
                'category': 'waiting'
            },
            11: {
                'name': 'Parent-Pending/Insurer',
                'display': 'Parent- Pending / Insurer',
                'pending_from': 'insurer',
                'action': 'Awaiting insurer response',
                'category': 'waiting'
            },
            12: {
                'name': 'Parent-Pending/Customer',
                'display': 'Parent- Pending /Customer',
                'pending_from': 'customer',
                'action': 'Awaiting customer input',
                'category': 'waiting'
            },
            13: {
                'name': 'Child-Pending/BM-Dealer',
                'display': 'Child- Pending / BM-Dealer',
                'pending_from': 'dealer',
                'action': 'Child ticket with dealer',
                'category': 'waiting'
            },
            14: {
                'name': 'Child-Pending/Insurer',
                'display': 'Child- Pending / Insurer',
                'pending_from': 'insurer',
                'action': 'Child ticket with insurer',
                'category': 'waiting'
            },
            15: {
                'name': 'Child-Pending/Customer',
                'display': 'Child- Pending / Customer',
                'pending_from': 'customer',
                'action': 'Child ticket awaiting customer',
                'category': 'waiting'
            },
            16: {
                'name': 'CODERED-WIP',
                'display': 'CODERED - Work in Progress',
                'pending_from': 'internal_team',
                'action': 'High priority - actively working',
                'category': 'critical'
            },
            17: {
                'name': 'CODERED-PENDING-SPOC',
                'display': 'CODERED - PENDING WITH SPOC',
                'pending_from': 'spoc',
                'action': 'Critical escalation with SPOC',
                'category': 'critical'
            },
            18: {
                'name': 'CODERED-Pending-Partner',
                'display': 'CODERED - Pending with Partner',
                'pending_from': 'partner',
                'action': 'Critical issue with partner',
                'category': 'critical'
            },
            21: {
                'name': 'Customer-Responded',
                'display': 'Customer Responded',
                'pending_from': 'internal_team',
                'action': 'Processing customer response',
                'category': 'active'
            },
            22: {
                'name': 'CODERED-NOC-HOLD',
                'display': 'CODERED NOC HOLD',
                'pending_from': 'noc',
                'action': 'On hold for NOC clearance',
                'category': 'critical'
            },
            23: {
                'name': 'KYC-Pending-Requester',
                'display': 'KYC Pending with Requester',
                'pending_from': 'customer',
                'action': 'KYC documents needed from customer',
                'category': 'compliance'
            },
            24: {
                'name': 'KYC-WIP',
                'display': 'KYC WIP',
                'pending_from': 'internal_team',
                'action': 'KYC verification in progress',
                'category': 'compliance'
            },
            25: {
                'name': 'Pending-Tech-Support',
                'display': 'Pending with Tech Support',
                'pending_from': 'tech_support',
                'action': 'Technical issue resolution',
                'category': 'technical'
            },
            26: {
                'name': 'Reopen',
                'display': 'Reopen',
                'pending_from': 'internal_team',
                'action': 'Ticket reopened for review',
                'category': 'active'
            }
        }

    def analyze_ticket_with_children(self, ticket_id):
        """Complete analysis of ticket including all child tickets"""
        print(f"Starting complete analysis for ticket {ticket_id}")
        
        # Get main ticket
        main_ticket = fetch_ticket_by_id(ticket_id)
        if not main_ticket:
            return {'error': f'Could not fetch ticket {ticket_id}'}
        
        # Get main ticket conversations
        main_conversations = fetch_all_ticket_conversations(ticket_id)
        
        # Analyze main ticket
        main_analysis = self.analyze_single_ticket(main_ticket, main_conversations, is_main=True)
        
        # Initialize result structure
        result = {
            'main_ticket': main_analysis,
            'child_tickets': [],
            'overall_status': self.determine_overall_status(main_analysis),
            'pending_summary': self.extract_pending_summary(main_analysis),
            'relationship_type': self.determine_relationship_type(main_ticket.get('status'))
        }
        
        # Check if this is a parent ticket
        if main_ticket.get('status') in [10, 11, 12]:  # Parent statuses
            print(f"Detected parent ticket, fetching child tickets...")
            
            # Fetch child tickets
            child_tickets = fetch_child_tickets(ticket_id)
            print(f"Found {len(child_tickets)} child tickets")
            
            # Analyze each child ticket
            for child in child_tickets:
                child_id = child.get('id')
                print(f"Analyzing child ticket {child_id}")
                
                # Get full child ticket data and conversations
                child_data = fetch_ticket_by_id(child_id)
                if child_data:
                    child_conversations = fetch_all_ticket_conversations(child_id)
                    child_analysis = self.analyze_single_ticket(
                        child_data, 
                        child_conversations, 
                        is_main=False, 
                        parent_id=ticket_id
                    )
                    result['child_tickets'].append(child_analysis)
            
            # Update overall status based on child tickets
            result['overall_status'] = self.determine_overall_status_with_children(
                main_analysis, 
                result['child_tickets']
            )
            
            # Create consolidated pending summary
            result['pending_summary'] = self.create_consolidated_pending_summary(
                main_analysis, 
                result['child_tickets']
            )
        
        # Add relationship analysis
        result['relationship_analysis'] = self.analyze_ticket_relationships(result)
        
        return result

    def analyze_single_ticket(self, ticket_data, conversations, is_main=True, parent_id=None):
        """Analyze a single ticket (parent or child) comprehensively"""
        ticket_id = ticket_data.get('id')
        status = ticket_data.get('status')
        status_info = self.status_mappings.get(status, {
            'name': 'Unknown',
            'display': 'Unknown Status',
            'pending_from': 'unknown',
            'action': 'Unknown action',
            'category': 'unknown'
        })
        
        # Extract content and actions
        raw_content, actions_taken = extract_email_content_and_attachments(ticket_data, conversations)
        
        # Determine where it's actually pending from (content analysis)
        actual_pending_from = self.determine_actual_pending_from(
            raw_content, 
            conversations, 
            status_info['pending_from']
        )
        
        # Extract key information from content
        key_info = self.extract_key_information(raw_content, conversations)
        
        # Calculate timing information
        timing_info = self.calculate_timing_info(ticket_data, conversations)
        
        analysis = {
            'ticket_id': ticket_id,
            'is_main_ticket': is_main,
            'parent_id': parent_id,
            'subject': ticket_data.get('subject', ''),
            'status': {
                'id': status,
                'name': status_info['name'],
                'display': status_info['display'],
                'category': status_info['category'],
                'action': status_info['action']
            },
            'pending_from': {
                'status_based': status_info['pending_from'],
                'content_based': actual_pending_from['primary'],
                'confidence': actual_pending_from['confidence'],
                'evidence': actual_pending_from['evidence']
            },
            'key_information': key_info,
            'timing': timing_info,
            'actions_taken': actions_taken,
            'last_activity': self.get_last_activity(conversations),
            'next_expected_action': self.determine_next_action(
                status_info, 
                actual_pending_from, 
                key_info
            )
        }
        
        return analysis

    def determine_actual_pending_from(self, content, conversations, status_pending):
        """Determine where ticket is actually pending from based on content analysis"""
        content_lower = content.lower()
        
        # Define patterns for different pending sources
        pending_patterns = {
            'insurer': {
                'keywords': [
                    'waiting for insurer', 'pending with insurance', 'insurer to respond',
                    'submitted to insurer', 'forwarded to insurance', 'claim intimated',
                    'awaiting insurer response', 'insurance company response',
                    'hdfc ergo', 'icici lombard', 'bajaj allianz', 'reliance general',
                    'new india', 'united india', 'tata aig'
                ],
                'patterns': [
                    r'submitted\s+to\s+(\w+)\s+(?:insurance|insurer)',
                    r'forwarded\s+to\s+(\w+)',
                    r'claim\s+intimated\s+to\s+(\w+)',
                    r'pending\s+(?:with|from)\s+(\w+)\s+(?:insurance|insurer)'
                ]
            },
            'customer': {
                'keywords': [
                    'waiting for customer', 'customer to provide', 'pending from customer',
                    'documents required from customer', 'customer response needed',
                    'awaiting customer confirmation', 'customer to submit',
                    'pending documents', 'kyc pending', 'forms to be filled'
                ],
                'patterns': [
                    r'waiting\s+for\s+customer',
                    r'customer\s+(?:to|needs to)\s+(?:provide|submit|send)',
                    r'pending\s+(?:from|with)\s+customer',
                    r'documents?\s+(?:required|needed|pending)\s+from\s+customer'
                ]
            },
            'dealer': {
                'keywords': [
                    'channel partner', 'dealer action', 'pos to update',
                    'broker to provide', 'agent response', 'partner pending',
                    'bm dealer', 'franchise to respond'
                ],
                'patterns': [
                    r'pending\s+(?:with|from)\s+(?:dealer|partner|agent)',
                    r'(?:channel\s+partner|dealer|agent)\s+(?:to|needs to)',
                    r'forwarded\s+to\s+(?:dealer|partner)'
                ]
            },
            'surveyor': {
                'keywords': [
                    'surveyor appointed', 'survey pending', 'inspection scheduled',
                    'surveyor to visit', 'survey report pending', 'assessment pending'
                ],
                'patterns': [
                    r'surveyor\s+(?:appointed|assigned|to\s+visit)',
                    r'survey\s+(?:pending|scheduled|in\s+progress)',
                    r'inspection\s+(?:pending|due)'
                ]
            },
            'garage': {
                'keywords': [
                    'garage estimate', 'workshop quote', 'repair estimate pending',
                    'garage to provide', 'workshop assessment'
                ],
                'patterns': [
                    r'garage\s+(?:to\s+provide|estimate)',
                    r'workshop\s+(?:quote|assessment)',
                    r'repair\s+estimate\s+pending'
                ]
            },
            'internal_team': {
                'keywords': [
                    'processing internally', 'under review', 'team working',
                    'internal verification', 'quality check', 'supervisor review'
                ],
                'patterns': [
                    r'(?:internal|team)\s+(?:processing|review|working)',
                    r'under\s+(?:review|verification)',
                    r'quality\s+check\s+in\s+progress'
                ]
            }
        }
        
        # Score each potential source
        scores = {}
        evidence = {}
        
        for source, config in pending_patterns.items():
            score = 0
            source_evidence = []
            
            # Check keywords
            for keyword in config['keywords']:
                if keyword in content_lower:
                    score += 2
                    source_evidence.append(f"Keyword: {keyword}")
            
            # Check patterns
            for pattern in config.get('patterns', []):
                matches = re.findall(pattern, content_lower)
                if matches:
                    score += 3
                    source_evidence.append(f"Pattern match: {pattern}")
            
            if score > 0:
                scores[source] = score
                evidence[source] = source_evidence
        
        # Get the latest conversation for recent context
        if conversations:
            latest_conv = conversations[-1]
            latest_text = latest_conv.get('body_text', '') or clean_html(latest_conv.get('body', ''))
            latest_lower = latest_text.lower()
            
            # Boost scores for mentions in latest conversation
            for source in scores:
                for keyword in pending_patterns[source]['keywords']:
                    if keyword in latest_lower:
                        scores[source] += 1
        
        # Determine primary pending source
        if scores:
            primary_source = max(scores.items(), key=lambda x: x[1])
            confidence = min(primary_source[1] / 5.0, 1.0)  # Normalize to 0-1
            
            return {
                'primary': primary_source[0],
                'confidence': confidence,
                'all_scores': scores,
                'evidence': evidence.get(primary_source[0], [])
            }
        else:
            # Fall back to status-based determination
            return {
                'primary': status_pending,
                'confidence': 0.3,
                'all_scores': {},
                'evidence': ['Based on ticket status only']
            }

    def extract_key_information(self, content, conversations):
        """Extract key information from ticket content"""
        import re
        
        key_info = {
            'claim_numbers': [],
            'policy_numbers': [],
            'insurer_names': [],
            'amounts': [],
            'dates': [],
            'reference_numbers': [],
            'contact_numbers': [],
            'document_types': []
        }
        
        # Extract patterns
        patterns = {
            'claim_numbers': r'claim\s*#?\s*([A-Z0-9/-]{6,})',
            'policy_numbers': r'policy\s*#?\s*([A-Z0-9/-]{6,})',
            'amounts': r'₹\s*(\d+(?:,\d+)*(?:\.\d+)?)',
            'dates': r'\b\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\b',
            'reference_numbers': r'(?:ref|reference)\s*#?\s*([A-Z0-9/-]{4,})',
            'contact_numbers': r'\b(?:\+91[\s-]?)?[6-9]\d{9}\b'
        }
        
        for key, pattern in patterns.items():
            matches = re.findall(pattern, content, re.IGNORECASE)
            key_info[key] = list(set(matches))  # Remove duplicates
        
        # Extract insurer names
        insurer_keywords = [
            'hdfc ergo', 'icici lombard', 'bajaj allianz', 'reliance general',
            'tata aig', 'new india', 'united india', 'kotak mahindra',
            'digit insurance', 'go digit', 'acko', 'sbi general'
        ]
        
        content_lower = content.lower()
        for insurer in insurer_keywords:
            if insurer in content_lower:
                key_info['insurer_names'].append(insurer.title())
        
        # Extract document types mentioned
        document_keywords = [
            'policy copy', 'rc copy', 'driving license', 'claim form',
            'survey report', 'estimate', 'invoice', 'kyc documents',
            'pan card', 'aadhaar', 'bank statement', 'cancelled cheque'
        ]
        
        for doc_type in document_keywords:
            if doc_type in content_lower:
                key_info['document_types'].append(doc_type)
        
        return key_info

    def calculate_timing_info(self, ticket_data, conversations):
        """Calculate various timing metrics"""
        created_at = ticket_data.get('created_at')
        updated_at = ticket_data.get('updated_at')
        
        if not created_at:
            return {'error': 'No creation date available'}
        
        try:
            created_time = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            current_time = datetime.now(created_time.tzinfo)
            
            age_hours = (current_time - created_time).total_seconds() / 3600
            
            timing_info = {
                'created_at': created_at,
                'updated_at': updated_at,
                'age_hours': round(age_hours, 2),
                'age_days': round(age_hours / 24, 1),
                'business_hours_age': self.calculate_business_hours(created_time, current_time)
            }
            
            # Calculate time since last update
            if updated_at:
                updated_time = datetime.fromisoformat(updated_at.replace('Z', '+00:00'))
                hours_since_update = (current_time - updated_time).total_seconds() / 3600
                timing_info['hours_since_last_update'] = round(hours_since_update, 2)
            
            # Find last customer interaction
            last_customer_interaction = None
            for conv in reversed(conversations):
                if conv.get('incoming') == True:  # Customer message
                    last_customer_interaction = conv.get('created_at')
                    break
            
            if last_customer_interaction:
                last_customer_time = datetime.fromisoformat(last_customer_interaction.replace('Z', '+00:00'))
                hours_since_customer = (current_time - last_customer_time).total_seconds() / 3600
                timing_info['hours_since_last_customer_message'] = round(hours_since_customer, 2)
            
            return timing_info
            
        except Exception as e:
            return {'error': f'Error calculating timing: {str(e)}'}

    def calculate_business_hours(self, start_time, end_time):
        """Calculate business hours between two times"""
        # Simplified calculation - assumes 9 AM to 6 PM, Monday to Friday
        business_hours = 0
        current = start_time.replace(hour=9, minute=0, second=0, microsecond=0)
        
        while current < end_time:
            # Check if it's a weekday
            if current.weekday() < 5:  # Monday = 0, Friday = 4
                # Check if within business hours
                if 9 <= current.hour < 18:
                    business_hours += 1
            current += timedelta(hours=1)
        
        return business_hours

    def get_last_activity(self, conversations):
        """Get information about the last activity"""
        if not conversations:
            return {'type': 'none', 'description': 'No conversations found'}
        
        last_conv = conversations[-1]
        
        return {
            'type': 'conversation',
            'timestamp': last_conv.get('created_at'),
            'from': last_conv.get('from_email', 'Unknown'),
            'is_customer': last_conv.get('incoming', False),
            'is_private': last_conv.get('private', False),
            'preview': (last_conv.get('body_text', '') or clean_html(last_conv.get('body', '')))[:100] + '...'
        }

    def determine_next_action(self, status_info, pending_analysis, key_info):
        """Determine what the next action should be"""
        pending_from = pending_analysis['primary']
        confidence = pending_analysis['confidence']
        
        # Define next actions based on pending source
        next_actions = {
            'customer': {
                'action': 'Follow up with customer',
                'details': 'Send reminder or make call to customer for pending requirements',
                'priority': 'high' if confidence > 0.7 else 'medium',
                'timeline': '24 hours'
            },
            'insurer': {
                'action': 'Follow up with insurer',
                'details': 'Check insurer portal or send reminder email to insurer',
                'priority': 'medium',
                'timeline': '48 hours'
            },
            'dealer': {
                'action': 'Contact channel partner',
                'details': 'Reach out to dealer/agent for required action',
                'priority': 'medium',
                'timeline': '24 hours'
            },
            'surveyor': {
                'action': 'Coordinate survey',
                'details': 'Contact surveyor for schedule or report status',
                'priority': 'high',
                'timeline': '12 hours'
            },
            'internal_team': {
                'action': 'Internal processing',
                'details': 'Continue internal review/processing of the request',
                'priority': 'medium',
                'timeline': '4 hours'
            }
        }
        
        base_action = next_actions.get(pending_from, {
            'action': 'Review and assess',
            'details': 'Determine appropriate next step based on ticket content',
            'priority': 'low',
            'timeline': '4 hours'
        })
        
        # Enhance with specific information if available
        if key_info['claim_numbers']:
            base_action['details'] += f" (Claim: {key_info['claim_numbers'][0]})"
        
        if key_info['insurer_names']:
            base_action['details'] += f" (Insurer: {key_info['insurer_names'][0]})"
        
        return base_action

    def determine_overall_status(self, main_analysis):
        """Determine overall status of the ticket"""
        status_category = main_analysis['status']['category']
        pending_from = main_analysis['pending_from']['content_based']
        
        return {
            'summary': f"Ticket is {status_category}, pending from {pending_from}",
            'category': status_category,
            'pending_from': pending_from,
            'confidence': main_analysis['pending_from']['confidence']
        }

    def determine_overall_status_with_children(self, main_analysis, child_analyses):
        """Determine overall status considering child tickets"""
        if not child_analyses:
            return self.determine_overall_status(main_analysis)
        
        # Analyze all child ticket statuses
        child_statuses = []
        for child in child_analyses:
            child_statuses.append({
                'id': child['ticket_id'],
                'category': child['status']['category'],
                'pending_from': child['pending_from']['content_based']
            })
        
        # Determine consolidated status
        active_children = [c for c in child_statuses if c['category'] in ['active', 'waiting']]
        
        if active_children:
            # Find the most critical pending source
            pending_sources = [c['pending_from'] for c in active_children]
            most_common_pending = max(set(pending_sources), key=pending_sources.count)
            
            return {
                'summary': f"Parent ticket with {len(active_children)} active child tickets, pending from {most_common_pending}",
                'category': 'waiting',
                'pending_from': most_common_pending,
                'child_statuses': child_statuses,
                'active_children_count': len(active_children)
            }
        else:
            return {
                'summary': "All child tickets completed or closed",
                'category': 'complete',
                'pending_from': 'none',
                'child_statuses': child_statuses,
                'active_children_count': 0
            }

    def extract_pending_summary(self, analysis):
        """Extract a concise summary of what's pending"""
        pending_from = analysis['pending_from']['content_based']
        evidence = analysis['pending_from']['evidence']
        key_info = analysis['key_information']
        
        summary = {
            'primary_pending_from': pending_from,
            'evidence': evidence[:3],  # Top 3 pieces of evidence
            'key_items': []
        }
        
        # Add key items that might be pending
        if key_info['document_types']:
            summary['key_items'].extend([f"Document: {doc}" for doc in key_info['document_types'][:3]])
        
        if key_info['claim_numbers']:
            summary['key_items'].append(f"Claim: {key_info['claim_numbers'][0]}")
        
        return summary

    def create_consolidated_pending_summary(self, main_analysis, child_analyses):
        """Create consolidated pending summary for parent-child relationship"""
        main_pending = self.extract_pending_summary(main_analysis)
        
        if not child_analyses:
            return main_pending
        
        # Collect all pending sources from children
        child_pending_sources = []
        all_evidence = main_pending['evidence'].copy()
        all_key_items = main_pending['key_items'].copy()
        
        for child in child_analyses:
            child_pending = self.extract_pending_summary(child)
            child_pending_sources.append(child_pending['primary_pending_from'])
            all_evidence.extend(child_pending['evidence'])
            all_key_items.extend(child_pending['key_items'])
        
        # Find most common pending source
        if child_pending_sources:
            most_common = max(set(child_pending_sources), key=child_pending_sources.count)
        else:
            most_common = main_pending['primary_pending_from']
        
        return {
            'primary_pending_from': most_common,
            'all_pending_sources': list(set(child_pending_sources + [main_pending['primary_pending_from']])),
            'evidence': list(set(all_evidence))[:5],  # Top 5 unique pieces of evidence
            'key_items': list(set(all_key_items))[:5],  # Top 5 unique key items
            'child_count': len(child_analyses)
        }

    def determine_relationship_type(self, status):
        """Determine the type of ticket relationship"""
        if status in [10, 11, 12]:
            return 'parent'
        elif status in [13, 14, 15]:
            return 'child'
        else:
            return 'standalone'

    def analyze_ticket_relationships(self, result):
        """Analyze relationships between parent and child tickets"""
        relationship_analysis = {
            'type': result['relationship_type'],
            'ticket_count': 1 + len(result['child_tickets']),
            'complexity_score': self.calculate_complexity_score(result)
        }
        
        if result['relationship_type'] == 'parent':
            # Analyze child ticket distribution
            child_statuses = {}
            child_pending = {}
            
            for child in result['child_tickets']:
                status = child['status']['category']
                pending = child['pending_from']['content_based']
                
                child_statuses[status] = child_statuses.get(status, 0) + 1
                child_pending[pending] = child_pending.get(pending, 0) + 1
            
            relationship_analysis.update({
                'child_status_distribution': child_statuses,
                'child_pending_distribution': child_pending,
                'coordination_needed': len(set(child_pending.keys())) > 1,
                'workflow_complexity': 'high' if len(child_pending) > 2 else 'medium' if len(child_pending) > 1 else 'low'
            })
        
        return relationship_analysis

    def calculate_complexity_score(self, result):
        """Calculate complexity score based on various factors"""
        score = 0
        
        # Base complexity
        main_ticket = result['main_ticket']
        
        # Age factor
        age_days = main_ticket['timing'].get('age_days', 0)
        if age_days > 7:
            score += 2
        elif age_days > 3:
            score += 1
        
        # Child ticket factor
        child_count = len(result['child_tickets'])
        score += min(child_count, 3)  # Max 3 points for child tickets
        
        # Pending sources diversity
        all_pending = set()
        all_pending.add(main_ticket['pending_from']['content_based'])
        for child in result['child_tickets']:
            all_pending.add(child['pending_from']['content_based'])
        
        score += len(all_pending) - 1  # More diverse pending sources = more complex
        
        # Key information complexity
        key_info = main_ticket['key_information']
        if len(key_info['claim_numbers']) > 1:
            score += 1
        if len(key_info['insurer_names']) > 1:
            score += 1
        
        return min(score, 10)  # Cap at 10

    def generate_actionable_insights(self, result):
        """Generate actionable insights based on the complete analysis"""
        insights = []
        main_ticket = result['main_ticket']
        
        # Time-based insights
        age_days = main_ticket['timing'].get('age_days', 0)
        if age_days > 7:
            insights.append({
                'type': 'urgent',
                'message': f"Ticket is {age_days} days old - requires immediate attention",
                'action': 'Escalate to supervisor'
            })
        
        # Pending-based insights
        pending_from = result['pending_summary']['primary_pending_from']
        confidence = main_ticket['pending_from']['confidence']
        
        if confidence > 0.8:
            if pending_from == 'customer':
                insights.append({
                    'type': 'action_required',
                    'message': 'High confidence that customer action is needed',
                    'action': 'Send follow-up communication to customer'
                })
            elif pending_from == 'insurer':
                insights.append({
                    'type': 'action_required',
                    'message': 'Ticket appears to be pending with insurer',
                    'action': 'Check insurer portal and send reminder if needed'
                })
        
        # Child ticket insights
        if result['relationship_type'] == 'parent' and result['child_tickets']:
            active_children = [c for c in result['child_tickets'] if c['status']['category'] in ['active', 'waiting']]
            if len(active_children) > 3:
                insights.append({
                    'type': 'complexity',
                    'message': f"Multiple active child tickets ({len(active_children)}) - coordination needed",
                    'action': 'Review child ticket priorities and consolidate actions'
                })
        
        # Stagnation insights
        hours_since_update = main_ticket['timing'].get('hours_since_last_update', 0)
        if hours_since_update > 48:
            insights.append({
                'type': 'stagnation',
                'message': f"No updates for {hours_since_update:.1f} hours",
                'action': 'Investigate current status and take proactive action'
            })
        
        return insights

    def format_analysis_for_display(self, result):
        """Format the analysis result for easy display/consumption"""
        main_ticket = result['main_ticket']
        
        formatted = {
            'summary': {
                'ticket_id': main_ticket['ticket_id'],
                'subject': main_ticket['subject'],
                'current_status': main_ticket['status']['display'],
                'pending_from': result['pending_summary']['primary_pending_from'],
                'age_days': main_ticket['timing'].get('age_days', 0),
                'relationship_type': result['relationship_type'],
                'child_count': len(result['child_tickets'])
            },
            'pending_analysis': {
                'primary_source': result['pending_summary']['primary_pending_from'],
                'confidence': main_ticket['pending_from']['confidence'],
                'evidence': result['pending_summary']['evidence'],
                'all_sources': result['pending_summary'].get('all_pending_sources', [])
            },
            'next_actions': [main_ticket['next_expected_action']],
            'key_information': main_ticket['key_information'],
            'child_summary': [],
            'insights': self.generate_actionable_insights(result)
        }
        
        # Add child summaries
        for child in result['child_tickets']:
            formatted['child_summary'].append({
                'ticket_id': child['ticket_id'],
                'subject': child['subject'][:50] + '...' if len(child['subject']) > 50 else child['subject'],
                'status': child['status']['display'],
                'pending_from': child['pending_from']['content_based'],
                'age_days': child['timing'].get('age_days', 0)
            })
            
            # Add child next actions
            formatted['next_actions'].append(child['next_expected_action'])
        
        return formatted


# Enhanced functions to integrate with your existing code

def analyze_ticket_comprehensively(ticket_id):
    """Main function to analyze ticket with children and pending status"""
    analyzer = EnhancedTicketAnalyzer(FRESHDESK_DOMAIN, FRESHDESK_API_KEY)
    
    try:
        # Perform comprehensive analysis
        result = analyzer.analyze_ticket_with_children(ticket_id)
        
        if 'error' in result:
            return result
        
        # Format for display
        formatted_result = analyzer.format_analysis_for_display(result)
        
        # Add the raw analysis for detailed access
        formatted_result['raw_analysis'] = result
        
        return formatted_result
        
    except Exception as e:
        return {
            'error': f'Error analyzing ticket {ticket_id}: {str(e)}',
            'ticket_id': ticket_id
        }

def get_pending_status_summary(ticket_id):
    """Quick function to get just the pending status information"""
    analyzer = EnhancedTicketAnalyzer(FRESHDESK_DOMAIN, FRESHDESK_API_KEY)
    
    # Get main ticket
    ticket_data = fetch_ticket_by_id(ticket_id)
    if not ticket_data:
        return {'error': f'Could not fetch ticket {ticket_id}'}
    
    conversations = fetch_all_ticket_conversations(ticket_id)
    raw_content, _ = extract_email_content_and_attachments(ticket_data, conversations)
    
    status = ticket_data.get('status')
    status_info = analyzer.status_mappings.get(status, {})
    
    # Determine actual pending from
    actual_pending = analyzer.determine_actual_pending_from(
        raw_content, 
        conversations, 
        status_info.get('pending_from', 'unknown')
    )
    
    return {
        'ticket_id': ticket_id,
        'status_name': status_info.get('display', 'Unknown'),
        'pending_from_status': status_info.get('pending_from', 'unknown'),
        'pending_from_analysis': actual_pending['primary'],
        'confidence': actual_pending['confidence'],
        'evidence': actual_pending['evidence'],
        'recommendation': analyzer.determine_next_action(
            status_info, 
            actual_pending, 
            analyzer.extract_key_information(raw_content, conversations)
        )
    }

def enhance_existing_ticket_processing(ticket_id):
    """Enhanced version of your existing process_ticket_id function"""
    # Get your existing analysis
    existing_result = process_ticket_id_orignal(ticket_id)
    
    if not existing_result or 'error' in existing_result:
        return existing_result
    
    # Add comprehensive child ticket and pending analysis
    comprehensive_analysis = analyze_ticket_comprehensively(ticket_id)
    
    if 'error' not in comprehensive_analysis:
        # Merge the analyses
        existing_result['child_ticket_analysis'] = comprehensive_analysis
        existing_result['pending_status'] = comprehensive_analysis['pending_analysis']
        existing_result['actionable_insights'] = comprehensive_analysis['insights']
        existing_result['next_actions'] = comprehensive_analysis['next_actions']
        
        # Update the main pending status
        existing_result['currently_pending_from'] = comprehensive_analysis['pending_analysis']['primary_source']
        existing_result['pending_confidence'] = comprehensive_analysis['pending_analysis']['confidence']
    
    return existing_result

# Usage examples:

def print_ticket_summary(ticket_id):
    """Print a nice summary of ticket status"""
    analysis = analyze_ticket_comprehensively(ticket_id)
    
    if 'error' in analysis:
        print(f"Error: {analysis['error']}")
        return
    
    summary = analysis['summary']
    pending = analysis['pending_analysis']
    
    print(f"\n=== TICKET {summary['ticket_id']} ANALYSIS ===")
    print(f"Subject: {summary['subject']}")
    print(f"Status: {summary['current_status']}")
    print(f"Age: {summary['age_days']} days")
    print(f"Type: {summary['relationship_type'].title()} ticket")
    
    if summary['child_count'] > 0:
        print(f"Child Tickets: {summary['child_count']}")
    
    print(f"\n--- PENDING STATUS ---")
    print(f"Currently Pending From: {pending['primary_source'].title()}")
    print(f"Confidence: {pending['confidence']:.2%}")
    print(f"Evidence: {', '.join(pending['evidence'][:3])}")
    
    if analysis['next_actions']:
        print(f"\n--- NEXT ACTIONS ---")
        for action in analysis['next_actions']:
            print(f"• {action['action']}: {action['details']}")
            print(f"  Priority: {action['priority']}, Timeline: {action['timeline']}")
    
    if analysis['child_summary']:
        print(f"\n--- CHILD TICKETS ---")
        for child in analysis['child_summary']:
            print(f"• #{child['ticket_id']}: {child['status']} (pending from {child['pending_from']})")
    
    if analysis['insights']:
        print(f"\n--- INSIGHTS ---")
        for insight in analysis['insights']:
            print(f"• [{insight['type'].upper()}] {insight['message']}")
            print(f"  Action: {insight['action']}")

# Example integration with your GUI
def get_enhanced_ticket_for_gui(ticket_id):
    """Enhanced version for GUI that includes all child ticket information"""
    result = enhance_existing_ticket_processing(ticket_id)
    
    # Format for GUI display
    if result and 'child_ticket_analysis' in result:
        gui_result = result.copy()
        
        # Add summary fields for easy GUI access
        analysis = result['child_ticket_analysis']
        gui_result['pending_from'] = analysis['pending_analysis']['primary_source']
        gui_result['pending_confidence'] = analysis['pending_analysis']['confidence']
        gui_result['child_tickets_count'] = analysis['summary']['child_count']
        gui_result['urgency_level'] = 'high' if analysis['summary']['age_days'] > 7 else 'medium' if analysis['summary']['age_days'] > 3 else 'low'
        
        # Format next actions for GUI
        gui_result['next_action_summary'] = []
        for action in analysis.get('next_actions', []):
            gui_result['next_action_summary'].append({
                'action': action['action'],
                'priority': action['priority'],
                'timeline': action['timeline']
            })
        
        return gui_result
    
    return result


def process_ticket_attachments_enhanced(ticket_data: dict) -> dict:
    """Enhanced attachment processing with Vision API - DEBUG VERSION"""
    print("DEBUG: process_ticket_attachments_enhanced called")
    print(f"DEBUG: Input ticket_data type: {type(ticket_data)}")
    print(f"DEBUG: Input ticket_data keys: {list(ticket_data.keys()) if isinstance(ticket_data, dict) else 'Not a dict'}")
    
    try:
        # 1) Fetch attachments if they're missing
        if not ticket_data.get('attachments'):
            ticket_id = ticket_data.get('Ticket ID')
            if ticket_id:
                print(f"DEBUG: No attachments in ticket_data, fetching for ticket {ticket_id}")
                fresh_data = fetch_ticket_by_id(ticket_id)
                if fresh_data and 'attachments' in fresh_data:
                    ticket_data['attachments'] = fresh_data['attachments']
                    print(f"DEBUG: Added {len(fresh_data['attachments'])} attachments to ticket_data")

        # 2) Create Textract client & analyzer in one go
        print("DEBUG: Creating DocumentAnalyzer with Textract + S3 settings...")
        analyzer = DocumentAnalyzer()
        print("DEBUG: DocumentAnalyzer created successfully")

        # 3) Analyze attachments
        print("DEBUG: Calling batch_analyze_attachments...")
        attachment_analysis = analyzer.analyze_all_attachments(ticket_data['attachments'])
        print(f"DEBUG: batch_analyze_attachments completed, result type: {type(attachment_analysis)}")
        print(f"DEBUG: Attachment analysis keys: {list(attachment_analysis.keys()) if isinstance(attachment_analysis, dict) else 'Not a dict'}")

        # 4) Merge results back into ticket_data
        print("DEBUG: Updating ticket_data with attachment_analysis...")
        ticket_data['attachment_analysis'] = attachment_analysis

        # 5) Build insights based on analysis
        print("DEBUG: Generating document-based insights...")
        insights = []

        # Missing documents?
        if attachment_analysis.get('missing_documents'):
            missing_docs_list = [
                doc.get('file', 'unknown') for doc in attachment_analysis['missing_documents']
                if isinstance(doc, dict)
            ]
            print(f"DEBUG: Found {len(missing_docs_list)} missing documents")
            insights.append({
                'type': 'MISSING_DOCUMENTS',
                'priority': 'HIGH',
                'message': f"Missing required documents: {', '.join(missing_docs_list)}",
                'action': 'Request missing documents from customer'
            })

        # Quality issues?
        if attachment_analysis.get('quality_issues'):
            print(f"DEBUG: Found {len(attachment_analysis['quality_issues'])} quality issues")
            insights.append({
                'type': 'QUALITY_ISSUES',
                'priority': 'MEDIUM',
                'message': f"{len(attachment_analysis['quality_issues'])} documents have quality issues",
                'action': 'Request clearer copies of affected documents'
            })

        # 6) Extract & validate structured data
        print("DEBUG: Extracting key information...")
        extracted_info = {}
        for doc in attachment_analysis.get('analyzed', []):
            if doc.get('extracted_data'):
                extracted_info.update(doc['extracted_data'])
        print(f"DEBUG: Extracted {len(extracted_info)} pieces of information")

        print("DEBUG: Validating extracted information...")
        validation_results = validate_extracted_info(extracted_info, ticket_data)
        if validation_results.get('mismatches'):
            print(f"DEBUG: Found {len(validation_results['mismatches'])} data mismatches")
            insights.append({
                'type': 'DATA_MISMATCH',
                'priority': 'HIGH',
                'message': "Document data doesn't match ticket information",
                'details': validation_results['mismatches']
            })

        # 7) Attach insights & extracted data back to ticket
        print(f"DEBUG: Generated {len(insights)} insights")
        ticket_data['document_insights'] = insights
        ticket_data['extracted_document_data'] = extracted_info

        print("DEBUG: process_ticket_attachments_enhanced completed successfully")
        return ticket_data

    except Exception as e:
        print(f"DEBUG: Exception in process_ticket_attachments_enhanced: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()

        ticket_data['attachment_analysis'] = {
            'error': str(e),
            'message': 'Document analysis failed'
        }
        return ticket_data

    
# You can also create a simple test version that doesn't do actual analysis:
def process_ticket_attachments_enhanced_mock(ticket_data: dict) -> dict:
    """Mock version for testing"""
    print("DEBUG: Mock document analysis called")
    
    # Add mock analysis results
    ticket_data['attachment_analysis'] = {
        'total_attachments': 2,
        'analyzed': [
            {
                'filename': 'mock_document_1.pdf',
                'document_type': 'insurance_policy',
                'category': 'insurance',
                'confidence': 0.85,
                'extracted_data': {'policy_number': 'MOCK123456'},
                'quality_assessment': {'is_readable': True, 'issues': []}
            }
        ],
        'document_summary': {'insurance': ['mock_document_1.pdf']},
        'missing_documents': ['id_proof', 'address_proof'],
        'quality_issues': []
    }
    
    ticket_data['document_insights'] = [
        {
            'type': 'MISSING_DOCUMENTS',
            'priority': 'HIGH',
            'message': 'Missing required documents: id_proof, address_proof',
            'action': 'Request missing documents from customer'
        }
    ]
    
    print("DEBUG: Mock analysis completed")
    return ticket_data

def validate_extracted_info(extracted_info: dict, ticket_data: dict) -> dict:
    """Validate extracted information against ticket data"""
    
    validation_results = {
        'valid': True,
        'mismatches': []
    }
    
    # Example validations
    if 'policy_number' in extracted_info:
        ticket_policy = ticket_data.get('custom_fields', {}).get('policy_number')
        if ticket_policy and extracted_info['policy_number'] != ticket_policy:
            validation_results['valid'] = False
            validation_results['mismatches'].append({
                'field': 'policy_number',
                'document_value': extracted_info['policy_number'],
                'ticket_value': ticket_policy
            })
    
    return validation_results

class DocumentWorkflowAutomation:
    """Automate document-based workflows"""
    
    def __init__(self, document_analyzer: DocumentAnalyzer):
        self.analyzer = document_analyzer
        self.automation_rules = self._load_automation_rules()
    
    def _load_automation_rules(self):
        """Load document automation rules"""
        return {
            'auto_approve_kyc': {
                'condition': lambda docs: self._has_valid_kyc_docs(docs),
                'action': 'approve_kyc',
                'priority': 'HIGH'
            },
            'auto_validate_claim': {
                'condition': lambda docs: self._has_complete_claim_docs(docs),
                'action': 'validate_claim_documents',
                'priority': 'MEDIUM'
            },
            'auto_extract_policy': {
                'condition': lambda docs: self._has_policy_document(docs),
                'action': 'extract_policy_details',
                'priority': 'HIGH'
            }
        }
    
    def process_document_workflow(self, ticket_data: dict) -> dict:
        """Process document-based workflow automation"""
        
        if 'attachment_analysis' not in ticket_data:
            return {'status': 'no_documents'}
        
        analyzed_docs = ticket_data['attachment_analysis']['analyzed']
        automation_results = []
        
        for rule_name, rule in self.automation_rules.items():
            if rule['condition'](analyzed_docs):
                result = self._execute_automation(rule['action'], ticket_data)
                automation_results.append({
                    'rule': rule_name,
                    'action': rule['action'],
                    'result': result,
                    'timestamp': datetime.now().isoformat()
                })
        
        return {
            'automations_executed': automation_results,
            'document_validation': self._validate_all_documents(analyzed_docs),
            'next_steps': self._determine_next_steps(ticket_data)
        }
    
    def _has_valid_kyc_docs(self, docs: list) -> bool:
        """Check if valid KYC documents are present"""
        required_kyc = {'identity', 'address_proof'}
        found_categories = {doc.get('category') for doc in docs if doc.get('confidence', 0) > 0.8}
        return required_kyc.issubset(found_categories)
    
    def _execute_automation(self, action: str, ticket_data: dict) -> dict:
        """Execute automated action based on document analysis"""
        
        if action == 'approve_kyc':
            # Auto-approve KYC if documents are valid
            return {
                'status': 'approved',
                'message': 'KYC documents verified automatically',
                'confidence': 0.95
            }
        elif action == 'validate_claim_documents':
            # Validate claim documents
            return self._validate_claim_documents(ticket_data)
        elif action == 'extract_policy_details':
            # Extract and update policy details
            return self._extract_and_update_policy(ticket_data)
        
        return {'status': 'unknown_action'}

# Integration functions with your existing code

def extract_claim_type_from_ticket(ticket_data: dict, classification: str) -> str:
    """Extract the specific claim type from ticket content"""
    content = ticket_data.get('raw_ticket_content', '').lower()
    subject = ticket_data.get('Subject', '').lower()
    
    # Map of keywords to claim types
    claim_type_keywords = {
        'CASHLESS_GARAGE_REQUIRED': ['cashless garage', 'network garage', 'preferred garage', 'cashless workshop'],
        'CLAIM_INTIMATION': ['claim intimation', 'intimate claim', 'new claim', 'fresh claim', 'report claim', 'file claim'],
        'SURVEY_PENDING': ['survey pending', 'surveyor', 'inspection pending', 'assessment pending'],
        'DELIVERY_ORDER_PENDING': ['delivery order', 'do pending', 'repair approval'],
        'REIMBURSEMENT_PENDING': ['reimbursement', 'reimburse', 'payment pending', 'claim payment'],
        'RAISE_QUERY_ON_THE_SETTLED_CLAIM_AMOUNT': ['settled amount', 'claim settled', 'settlement query', 'less settlement'],
        'WORK_APPROVAL_PENDING': ['work approval', 'repair approval', 'approval pending']
    }
    
    # Check in subject first (higher priority)
    for claim_type, keywords in claim_type_keywords.items():
        for keyword in keywords:
            if keyword in subject:
                return claim_type
    
    # Then check in content
    for claim_type, keywords in claim_type_keywords.items():
        for keyword in keywords:
            if keyword in content:
                return claim_type
    
    # Default based on classification
    if 'cashless' in classification.lower():
        return 'CASHLESS_GARAGE_REQUIRED'
    else:
        return 'CLAIM_INTIMATION'

def extract_insurer_from_ticket(ticket_data: dict) -> tuple[int, str]:
    """Extract insurer ID and name from ticket content"""
    content = ticket_data.get('raw_ticket_content', '').lower()
    subject = ticket_data.get('Subject', '').lower()
    combined = f"{subject} {content}"
    
    # Check custom fields first
    custom_fields = ticket_data.get('custom_fields', {})
    if 'insurer_id' in custom_fields:
        return int(custom_fields['insurer_id']), ''
    
    # Map of insurer names to IDs
    insurer_keywords = {
        1: ['hdfc ergo', 'hdfc', 'ergo'],
        6: ['royal sundaram', 'sundaram'],
        10: ['bajaj allianz', 'bajaj'],
        18: ['united india', 'uiic', 'united india insurance'],
        2: ['reliance general', 'reliance'],
        14: ['new india', 'new india assurance', 'nia'],
        4: ['icici lombard', 'icici'],
        5: ['kotak mahindra', 'kotak general', 'kotak'],
        28: ['edelweiss', 'edelweiss general'],
        20: ['liberty general', 'liberty'],
        9: ['future generali', 'generali'],
        22: ['magma hdi', 'magma'],
        16: ['shriram general', 'shriram'],
        12: ['digit', 'go digit'],
        11: ['universal sompo', 'sompo'],
        21: ['cholamandalam', 'chola ms', 'chola'],
        15: ['sbi general', 'sbi']
    }
    
    # Search for insurer mentions
    for insurer_id, keywords in insurer_keywords.items():
        for keyword in keywords:
            if keyword in combined:
                doc_engine = DocumentRequirementEngine()
                insurer_name = doc_engine.insurer_mapping.get(insurer_id, f"Insurer {insurer_id}")
                return insurer_id, insurer_name
    
    # Default to 0 (general requirements)
    return 0, "Unknown Insurer"

def check_existing_attachments(ticket_data: dict) -> List[str]:
    """Check which documents are already attached to the ticket"""
    attached_docs = []
    
    # Check attachments
    attachments = ticket_data.get('attachments', [])
    for attachment in attachments:
        filename = attachment.get('name', '').lower()
        
        # Map filenames to document types
        if 'policy' in filename:
            attached_docs.append('Policy Doc')
        elif 'rc' in filename or 'registration' in filename:
            if 'page1' in filename or 'front' in filename:
                attached_docs.append('Rc Page 1')
            elif 'page2' in filename or 'back' in filename:
                attached_docs.append('Rc Page 2')
            else:
                attached_docs.append('Rc Page 1')
        elif 'driving' in filename or 'dl' in filename or 'license' in filename:
            if 'page1' in filename or 'front' in filename:
                attached_docs.append('Drvng Linc 1')
            elif 'page2' in filename or 'back' in filename:
                attached_docs.append('Drvng Linc 2')
            else:
                attached_docs.append('Drvng Linc 1')
        elif 'aadhar' in filename or 'aadhaar' in filename:
            if 'page1' in filename or 'front' in filename:
                attached_docs.append('Aadhar Page 1')
            elif 'page2' in filename or 'back' in filename:
                attached_docs.append('Aadhar Page 2')
            else:
                attached_docs.append('Aadhar Page 1')
        elif 'pan' in filename:
            attached_docs.append('Pan')
        elif 'invoice' in filename or 'bill' in filename:
            attached_docs.append('Invoice Copy')
        elif 'estimate' in filename:
            attached_docs.append('Repair Estimates')
        elif 'claim' in filename and 'form' in filename:
            attached_docs.append('Claim Form')
    
    return attached_docs

def generate_document_request_response(required_docs: List[Dict], optional_docs: List[Dict], 
                                     insurer_name: str, claim_type: str, attached_docs: List[str]) -> str:
    """Generate a response asking for required documents"""
    
    # Filter out already attached documents
    missing_required = [doc for doc in required_docs if doc['name'] not in attached_docs]
    missing_optional = [doc for doc in optional_docs if doc['name'] not in attached_docs]
    
    if not missing_required and not missing_optional:
        return generate_all_documents_received_response(insurer_name, claim_type)
    
    response_parts = []
    
    # Greeting
    response_parts.append("Thank you for reaching out to us regarding your insurance claim.")
    
    # Claim type specific intro
    claim_intros = {
        'CASHLESS_GARAGE_REQUIRED': "To process your cashless garage request",
        'CLAIM_INTIMATION': "To initiate your claim with the insurer",
        'SURVEY_PENDING': "To proceed with the survey process",
        'DELIVERY_ORDER_PENDING': "To process the delivery order",
        'REIMBURSEMENT_PENDING': "To process your reimbursement claim",
        'RAISE_QUERY_ON_THE_SETTLED_CLAIM_AMOUNT': "To address your query on the settled claim amount",
        'WORK_APPROVAL_PENDING': "To get the repair work approved"
    }
    
    intro = claim_intros.get(claim_type, "To process your claim")
    if insurer_name != "Unknown Insurer":
        response_parts.append(f"{intro} with {insurer_name}, we require the following documents:")
    else:
        response_parts.append(f"{intro}, we require the following documents:")
    
    # Required documents
    if missing_required:
        response_parts.append("\n**Required Documents:**")
        for i, doc in enumerate(missing_required, 1):
            response_parts.append(f"{i}. {doc['description']}")
            if doc.get('kyc_document'):
                response_parts.append("   (KYC Document - Clear copy required)")
            if doc.get('invoice_document'):
                response_parts.append("   (Original invoice/bill required)")
    
    # Optional documents (if no required docs missing)
    if not missing_required and missing_optional:
        response_parts.append("\n**Additional Documents (if available):**")
        for i, doc in enumerate(missing_optional, 1):
            response_parts.append(f"{i}. {doc['description']}")
    
    # Special instructions based on insurer
    if insurer_name in ["New India", "United India Insurance Company Limited"]:
        response_parts.append("\n**Special Requirements:**")
        response_parts.append("- Claim form must be properly filled and signed")
        response_parts.append("- PAN card and cancelled cheque mandatory for this insurer")
    
    # Instructions
    response_parts.append("\n**Please ensure:**")
    response_parts.append("- All documents are clear and readable")
    response_parts.append("- Both sides of documents are uploaded where applicable")
    response_parts.append("- Documents are in PDF or image format (JPG/PNG)")
    
    # Closing
    response_parts.append("\nOnce we receive all required documents, we will immediately process your claim and keep you updated on the progress.")
    response_parts.append("\nFor any queries, please feel free to reach out to us.")
    
    return "\n".join(response_parts)

def generate_all_documents_received_response(insurer_name: str, claim_type: str) -> str:
    """Generate response when all documents are already received"""
    
    response_parts = []
    
    response_parts.append("Thank you for submitting your claim request.")
    response_parts.append("\nWe have received all the required documents for your claim.")
    
    # Next steps based on claim type
    next_steps = {
        'CASHLESS_GARAGE_REQUIRED': [
            "We will share the list of cashless garages in your area within 2 hours.",
            "You can visit any of these garages for cashless claim settlement."
        ],
        'CLAIM_INTIMATION': [
            f"We will now intimate your claim to {insurer_name if insurer_name != 'Unknown Insurer' else 'the insurance company'}.",
            "You will receive a claim number within 24 hours.",
            "Our team will follow up on your claim status regularly."
        ],
        'SURVEY_PENDING': [
            "We will coordinate with the surveyor for inspection.",
            "Survey will be scheduled within 48 hours.",
            "You will receive surveyor contact details shortly."
        ],
        'DELIVERY_ORDER_PENDING': [
            "We will process the delivery order with the insurer.",
            "Approval is expected within 24-48 hours.",
            "Garage will be notified once approved."
        ],
        'REIMBURSEMENT_PENDING': [
            "Your reimbursement claim will be processed.",
            "Expected settlement time is 7-10 working days.",
            "Amount will be credited to your registered bank account."
        ]
    }
    
    steps = next_steps.get(claim_type, [
        "We will process your request immediately.",
        "You will receive an update within 24 hours."
    ])
    
    response_parts.append("\n**Next Steps:**")
    for i, step in enumerate(steps, 1):
        response_parts.append(f"{i}. {step}")
    
    response_parts.append("\nWe will keep you updated at every stage of the claim process.")
    response_parts.append("\nFor any urgent queries, please feel free to contact us.")
    
    return "\n".join(response_parts)

def process_claims_ticket_with_documents(ticket_id: int) -> Dict:
    """
    Main function to process a claims ticket and generate appropriate response
    
    Args:
        ticket_id: Freshdesk ticket ID
        
    Returns:
        Dict with processing results and generated response
    """
    try:
        # Initialize document engine
        doc_engine = DocumentRequirementEngine()
        
        # Fetch ticket data
        ticket_data = fetch_ticket_by_id(ticket_id)
        if not ticket_data:
            return {
                'success': False,
                'error': 'Could not fetch ticket data',
                'ticket_id': ticket_id
            }
        
        # Get conversations
        conversations = fetch_all_ticket_conversations(ticket_id)
        raw_content, actions = extract_email_content_and_attachments(ticket_data, conversations)
        
        # Prepare enriched ticket data
        enriched_ticket = {
            'Ticket ID': ticket_id,
            'Subject': ticket_data.get('subject', ''),
            'raw_ticket_content': raw_content,
            'attachments': ticket_data.get('attachments', []),
            'custom_fields': ticket_data.get('custom_fields', {})
        }
        
        # Classify ticket
        classification, sop_details = classify_ticket_with_sop(raw_content)
        
        # Check if it's a claims ticket
        if not classification.startswith("Claims"):
            return {
                'success': True,
                'ticket_id': ticket_id,
                'classification': classification,
                'is_claims': False,
                'message': 'Not a claims ticket - standard processing applies',
                'suggested_action': 'Process according to ' + classification + ' SOP'
            }
        
        # Extract claim type
        claim_type = extract_claim_type_from_ticket(enriched_ticket, classification)
        
        # Extract insurer
        insurer_id, insurer_name = extract_insurer_from_ticket(enriched_ticket)
        
        # Get required documents
        required_docs, optional_docs = doc_engine.get_required_documents(claim_type, insurer_id)
        
        # Check existing attachments
        attached_docs = check_existing_attachments(enriched_ticket)
        
        # Generate response
        response = generate_document_request_response(
            required_docs, 
            optional_docs, 
            insurer_name, 
            claim_type, 
            attached_docs
        )
        
        # Prepare result
        result = {
            'success': True,
            'ticket_id': ticket_id,
            'classification': classification,
            'is_claims': True,
            'claim_type': claim_type,
            'insurer': {
                'id': insurer_id,
                'name': insurer_name
            },
            'documents': {
                'required': required_docs,
                'optional': optional_docs,
                'attached': attached_docs,
                'missing_required': [doc['name'] for doc in required_docs if doc['name'] not in attached_docs],
                'missing_optional': [doc['name'] for doc in optional_docs if doc['name'] not in attached_docs]
            },
            'generated_response': response,
            'next_action': 'Send response to customer requesting documents' if required_docs else 'Process claim with insurer',
            'sop_reference': sop_details
        }
        
        return result
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'ticket_id': ticket_id
        }

def send_automated_response(ticket_id: int, response_text: str) -> bool:
    """
    Send automated response to ticket
    
    Args:
        ticket_id: Freshdesk ticket ID
        response_text: Response text to send
        
    Returns:
        bool: Success status
    """
    try:
        url = f"https://{FRESHDESK_DOMAIN}.freshdesk.com/api/v2/tickets/{ticket_id}/reply"
        
        data = {
            "body": response_text,
            "status": 3  # Set to pending to wait for customer response
        }
        
        response = requests.post(
            url,
            json=data,
            auth=(FRESHDESK_API_KEY, "X"),
            headers={"Content-Type": "application/json"}
        )
        
        return response.status_code == 201
        
    except Exception as e:
        print(f"Error sending response: {e}")
        return False

# Example usage function
def process_incoming_ticket(ticket_id: int) -> Dict:
    """
    Complete flow for processing an incoming ticket
    
    Args:
        ticket_id: Freshdesk ticket ID
        
    Returns:
        Dict with complete processing results
    """
    # Process the ticket
    result = process_claims_ticket_with_documents(ticket_id)
    
    # If it's a claims ticket with missing documents, send automated response
    if result['success'] and result.get('is_claims') and result['documents']['missing_required']:
        # Send the generated response
        sent = send_automated_response(ticket_id, result['generated_response'])
        result['response_sent'] = sent
        
        # Update ticket tags
        if sent:
            update_ticket_tags(ticket_id, ['documents_requested', result['claim_type'].lower()])
    
    return result
def update_ticket_tags(ticket_id: int, tags: List[str]) -> bool:
   """Update ticket tags in Freshdesk"""
   try:
       url = f"https://{FRESHDESK_DOMAIN}.freshdesk.com/api/v2/tickets/{ticket_id}"
       
       data = {
           "tags": tags
       }
       
       response = requests.put(
           url,
           json=data,
           auth=(FRESHDESK_API_KEY, "X"),
           headers={"Content-Type": "application/json"}
       )
       
       return response.status_code == 200
       
   except Exception as e:
       print(f"Error updating tags: {e}")
       return False

def generate_follow_up_response(ticket_data: dict, days_since_request: int) -> str:
   """Generate follow-up response for pending documents"""
   
   missing_docs = ticket_data.get('documents', {}).get('missing_required', [])
   
   if days_since_request == 1:
       response = f"""Dear Customer,

This is a gentle reminder regarding your claim request.

We are still waiting for the following documents:
{chr(10).join(['- ' + doc for doc in missing_docs])}

Please upload these documents at your earliest convenience to avoid any delay in claim processing.

Thank you for your cooperation.

Best regards,
Claims Team"""

   elif days_since_request == 3:
       response = f"""Dear Customer,

We noticed that we haven't received the required documents for your claim yet.

Pending documents:
{chr(10).join(['- ' + doc for doc in missing_docs])}

**Important:** If documents are not received within the next 48 hours, we may have to close this request. You would need to raise a fresh claim request.

Please upload the documents immediately to ensure timely claim processing.

Best regards,
Claims Team"""

   else:
       response = f"""Dear Customer,

This is the final reminder for your pending claim documents.

Required documents not yet received:
{chr(10).join(['- ' + doc for doc in missing_docs])}

**Action Required:** Please upload documents within 24 hours to keep your claim active.

If you're facing any issues in arranging these documents, please let us know so we can assist you.

Best regards,
Claims Team"""
   
   return response

def check_document_completion(ticket_id: int) -> Dict:
   """
   Check if all required documents have been submitted after initial request
   
   Args:
       ticket_id: Freshdesk ticket ID
       
   Returns:
       Dict with document status
   """
   # Get latest ticket data
   ticket_data = fetch_ticket_by_id(ticket_id)
   conversations = fetch_all_ticket_conversations(ticket_id)
   
   # Get the initial requirements (you might want to store this in ticket custom fields)
   raw_content, _ = extract_email_content_and_attachments(ticket_data, conversations)
   
   enriched_ticket = {
       'Ticket ID': ticket_id,
       'Subject': ticket_data.get('subject', ''),
       'raw_ticket_content': raw_content,
       'attachments': ticket_data.get('attachments', []),
       'custom_fields': ticket_data.get('custom_fields', {})
   }
   
   # Re-check requirements
   doc_engine = DocumentRequirementEngine()
   classification, _ = classify_ticket_with_sop(raw_content)
   claim_type = extract_claim_type_from_ticket(enriched_ticket, classification)
   insurer_id, insurer_name = extract_insurer_from_ticket(enriched_ticket)
   
   required_docs, optional_docs = doc_engine.get_required_documents(claim_type, insurer_id)
   attached_docs = check_existing_attachments(enriched_ticket)
   
   missing_required = [doc['name'] for doc in required_docs if doc['name'] not in attached_docs]
   
   return {
       'ticket_id': ticket_id,
       'all_documents_received': len(missing_required) == 0,
       'missing_documents': missing_required,
       'attached_documents': attached_docs,
       'claim_type': claim_type,
       'insurer': insurer_name
   }

def process_document_submission(ticket_id: int) -> Dict:
   """
   Process when customer submits documents (triggered by conversation update)
   
   Args:
       ticket_id: Freshdesk ticket ID
       
   Returns:
       Dict with processing results
   """
   # Check document status
   doc_status = check_document_completion(ticket_id)
   
   if doc_status['all_documents_received']:
       # All documents received - proceed with claim
       response = f"""Dear Customer,

Thank you for submitting all the required documents.

We have verified and received:
{chr(10).join(['✓ ' + doc for doc in doc_status['attached_documents']])}

**Next Steps:**
1. We will now intimate your claim to {doc_status['insurer']}
2. You will receive the claim number within 24 hours
3. Our team will follow up regularly and keep you updated

Current Status: Claim Intimation in Progress

For any queries, please feel free to reach out.

Best regards,
Claims Team"""
       
       # Update ticket
       send_automated_response(ticket_id, response)
       update_ticket_status(ticket_id, 11)  # Parent-Pending/Insurer
       update_ticket_tags(ticket_id, ['all_documents_received', 'claim_intimation_pending'])
       
       # Trigger claim intimation process
       result = {
           'success': True,
           'action': 'proceed_with_claim',
           'next_step': 'intimate_claim_to_insurer',
           'documents_complete': True
       }
       
   else:
       # Still missing documents
       missing_docs_list = '\n'.join(['- ' + doc for doc in doc_status['missing_documents']])
       
       response = f"""Dear Customer,

Thank you for uploading the documents. However, we still need the following:

{missing_docs_list}

Please upload these remaining documents to proceed with your claim.

Best regards,
Claims Team"""
       
       send_automated_response(ticket_id, response)
       
       result = {
           'success': True,
           'action': 'request_remaining_documents',
           'documents_complete': False,
           'missing_documents': doc_status['missing_documents']
       }
   
   return result

def update_ticket_status(ticket_id: int, status: int) -> bool:
   """Update ticket status in Freshdesk"""
   try:
       url = f"https://{FRESHDESK_DOMAIN}.freshdesk.com/api/v2/tickets/{ticket_id}"
       
       data = {
           "status": status
       }
       
       response = requests.put(
           url,
           json=data,
           auth=(FRESHDESK_API_KEY, "X"),
           headers={"Content-Type": "application/json"}
       )
       
       return response.status_code == 200
       
   except Exception as e:
       print(f"Error updating status: {e}")
       return False

def create_claim_intimation_child_ticket(parent_ticket_id: int, insurer_name: str, documents: List[str]) -> Dict:
   """
   Create child ticket for claim intimation to insurer
   
   Args:
       parent_ticket_id: Parent ticket ID
       insurer_name: Name of the insurer
       documents: List of documents collected
       
   Returns:
       Dict with child ticket details
   """
   try:
       # Get parent ticket details
       parent_ticket = fetch_ticket_by_id(parent_ticket_id)
       
       # Create child ticket
       child_subject = f"Claim Intimation - {insurer_name} - Parent #{parent_ticket_id}"
       
       child_description = f"""Claim intimation request for Parent Ticket #{parent_ticket_id}

Customer Details:
- Name: {parent_ticket.get('requester', {}).get('name', 'N/A')}
- Email: {parent_ticket.get('requester', {}).get('email', 'N/A')}
- Phone: {parent_ticket.get('requester', {}).get('phone', 'N/A')}

Documents Received:
{chr(10).join(['- ' + doc for doc in documents])}

Original Request:
{parent_ticket.get('description', 'N/A')}

Please intimate this claim to {insurer_name} and provide claim number.
"""
       
       url = f"https://{FRESHDESK_DOMAIN}.freshdesk.com/api/v2/tickets"
       
       data = {
           "subject": child_subject,
           "description": child_description,
           "status": 14,  # Child-Pending/Insurer
           "priority": parent_ticket.get('priority', 1),
           "group_id": parent_ticket.get('group_id'),
           "type": "Claim Intimation",
           "custom_fields": {
               "cf_parent_ticket_id": str(parent_ticket_id)
           },
           "tags": ["child_ticket", "claim_intimation", insurer_name.lower().replace(" ", "_")]
       }
       
       response = requests.post(
           url,
           json=data,
           auth=(FRESHDESK_API_KEY, "X"),
           headers={"Content-Type": "application/json"}
       )
       
       if response.status_code == 201:
           child_ticket = response.json()
           return {
               'success': True,
               'child_ticket_id': child_ticket['id'],
               'child_ticket_url': f"https://{FRESHDESK_DOMAIN}.freshdesk.com/a/tickets/{child_ticket['id']}"
           }
       else:
           return {
               'success': False,
               'error': f"Failed to create child ticket: {response.text}"
           }
           
   except Exception as e:
       return {
           'success': False,
           'error': str(e)
       }

# Comprehensive workflow function
def automated_claims_workflow(ticket_id: int) -> Dict:
   """
   Complete automated workflow for claims processing
   
   This function handles:
   1. Initial classification
   2. Document requirement identification
   3. Automated response generation
   4. Document verification
   5. Claim intimation
   
   Args:
       ticket_id: Freshdesk ticket ID
       
   Returns:
       Dict with complete workflow results
   """
   workflow_log = []
   
   try:
       # Step 1: Initial Processing
       workflow_log.append({
           'step': 'initial_processing',
           'timestamp': datetime.now().isoformat()
       })
       
       initial_result = process_claims_ticket_with_documents(ticket_id)
       
       if not initial_result['success']:
           return {
               'success': False,
               'error': initial_result.get('error'),
               'workflow_log': workflow_log
           }
       
       # Step 2: Check if claims ticket
       if not initial_result.get('is_claims'):
           workflow_log.append({
               'step': 'non_claims_ticket',
               'classification': initial_result['classification'],
               'action': 'standard_processing'
           })
           return {
               'success': True,
               'is_claims': False,
               'result': initial_result,
               'workflow_log': workflow_log
           }
       
       # Step 3: Send document request if needed
       if initial_result['documents']['missing_required']:
           workflow_log.append({
               'step': 'document_request',
               'missing_documents': initial_result['documents']['missing_required'],
               'timestamp': datetime.now().isoformat()
           })
           
           # Send automated response
           sent = send_automated_response(ticket_id, initial_result['generated_response'])
           
           if sent:
               # Update ticket
               update_ticket_status(ticket_id, 3)  # Pending
               update_ticket_tags(ticket_id, [
                   'documents_pending',
                   initial_result['claim_type'].lower(),
                   f"insurer_{initial_result['insurer']['id']}"
               ])
               
               workflow_log.append({
                   'step': 'response_sent',
                   'status': 'success'
               })
           
           return {
               'success': True,
               'is_claims': True,
               'workflow_stage': 'documents_requested',
               'result': initial_result,
               'workflow_log': workflow_log
           }
       
       # Step 4: All documents received - proceed with claim
       else:
           workflow_log.append({
               'step': 'all_documents_available',
               'action': 'proceed_with_claim',
               'timestamp': datetime.now().isoformat()
           })
           
           # Send confirmation
           confirmation_response = generate_all_documents_received_response(
               initial_result['insurer']['name'],
               initial_result['claim_type']
           )
           
           send_automated_response(ticket_id, confirmation_response)
           
           # Create child ticket for insurer
           child_result = create_claim_intimation_child_ticket(
               ticket_id,
               initial_result['insurer']['name'],
               initial_result['documents']['attached']
           )
           
           if child_result['success']:
               workflow_log.append({
                   'step': 'child_ticket_created',
                   'child_ticket_id': child_result['child_ticket_id']
               })
               
               # Update parent ticket
               update_ticket_status(ticket_id, 11)  # Parent-Pending/Insurer
               update_ticket_tags(ticket_id, [
                   'claim_intimated',
                   'child_ticket_created',
                   initial_result['claim_type'].lower()
               ])
           
           return {
               'success': True,
               'is_claims': True,
               'workflow_stage': 'claim_intimated',
               'child_ticket': child_result,
               'result': initial_result,
               'workflow_log': workflow_log
           }
           
   except Exception as e:
       workflow_log.append({
           'step': 'error',
           'error': str(e),
           'timestamp': datetime.now().isoformat()
       })
       
       return {
           'success': False,
           'error': str(e),
           'workflow_log': workflow_log
       }
       
# Add intelligent document suggestions
class DocumentSuggestionEngine:
    """Suggest required documents based on ticket context"""
    
    def __init__(self):
        self.suggestion_rules = self._load_suggestion_rules()
    
    def suggest_documents(self, ticket_data: dict) -> list:
        """Suggest documents based on ticket type and content"""
        
        suggestions = []
        classification = ticket_data.get('Classification', '')
        content = ticket_data.get('raw_ticket_content', '').lower()
        
        # Get base suggestions from classification
        if classification in self.suggestion_rules:
            base_suggestions = self.suggestion_rules[classification]['required']
            suggestions.extend(base_suggestions)
            
            # Add conditional suggestions
            for condition, docs in self.suggestion_rules[classification].get('conditional', {}).items():
                if condition in content:
                    suggestions.extend(docs)
        
        # Remove already uploaded documents
        if 'attachment_analysis' in ticket_data:
            uploaded_types = {
                doc.get('document_type') for doc in 
                ticket_data['attachment_analysis'].get('analyzed', [])
            }
            suggestions = [s for s in suggestions if s['type'] not in uploaded_types]
        
        return suggestions
    
    def _load_suggestion_rules(self):
        """Load document suggestion rules"""
        return {
            'Claims-Motor': {
                'required': [
                    {
                        'type': 'claim_form',
                        'name': 'Claim Form',
                        'description': 'Duly filled and signed claim form',
                        'template_url': 'https://example.com/templates/motor_claim_form.pdf'
                    },
                    {
                        'type': 'rc_book',
                        'name': 'RC Book',
                        'description': 'Registration Certificate of the vehicle'
                    },
                    {
                        'type': 'driving_license',
                        'name': 'Driving License',
                        'description': 'Valid driving license of the driver'
                    }
                ],
                'conditional': {
                    'theft': [
                        {
                            'type': 'fir',
                            'name': 'FIR Copy',
                            'description': 'First Information Report from police'
                        }
                    ],
                    'third party': [
                        {
                            'type': 'legal_notice',
                            'name': 'Legal Notice',
                            'description': 'Legal notice served to third party'
                        }
                    ]
                }
            }
        }
    

status_map = {
    1: "New",
    2: "Open", 
    3: "Pending",
    4: "Resolved",
    5: "Closed"
}

# --- Global Agent Name Cache ---
agent_name_cache = {}
agent_cache_populated = False

# --- Initialize Claude Client ---
anthropic_client = None
if CLAUDE_API_KEY:
    try:
        anthropic_client = anthropic.Anthropic(api_key=CLAUDE_API_KEY)
        print("Claude client initialized successfully.")
    except Exception as e:
        print(f"Error initializing Claude client: {e}. Please check your API key.")
else:
    print("CLAUDE_API_KEY not found. Claude client will not be initialized.")

# --- Excel Configuration ---
# For Android compatibility, we need to handle file paths differently
def get_excel_path():
    """Get the path to the Excel file that works on both desktop and Android"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Try different possible locations
    possible_paths = [
        "ticket_summary.xlsx",
        os.path.join(script_dir, "ticket_summary.xlsx"),
        os.path.join(os.path.dirname(script_dir), "ticket_summary.xlsx"),
        "/data/data/com.mycompany.id_brain/files/flet/app/ticket_summary.xlsx"
    ]
    
    for path in possible_paths:
        if os.path.exists(path):
            return path
    
    # If file doesn't exist, return a default path where it should be created
    return os.path.join(script_dir, "ticket_summary.xlsx")

EXCEL_FILE = get_excel_path()
error_log_dir = "error_logs"
os.makedirs(error_log_dir, exist_ok=True)

# --- Placeholder for Clustering Models ---
cluster_model = None
embedding_model = None

def _sanitize_for_json(obj):
    """
    Recursively walk through `obj` and convert any datetime objects to ISO‐formatted strings.
    Works for dicts, lists, and primitives.
    """
    if isinstance(obj, dict):
        return {k: _sanitize_for_json(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_sanitize_for_json(v) for v in obj]
    elif isinstance(obj, datetime):
        return obj.isoformat()
    else:
        return obj


def fetch_freshdesk_tickets(page=1, per_page=30):
    """Fetches a page of tickets from Freshdesk."""
    url = f"https://{FRESHDESK_DOMAIN}.freshdesk.com/api/v2/tickets?page={page}&per_page={per_page}"
    response = requests.get(url, auth=(FRESHDESK_API_KEY, "X"))
    if response.status_code != 200:
        print(f"❌ Failed to fetch tickets from page {page}: Status {response.status_code}, Response: {response.text}")
        return []
    return response.json()

def fetch_ticket_by_id(ticket_id):
    """Fetches a single ticket by its ID from Freshdesk."""
    # Remove the include parameter - just fetch the ticket
    url = f"https://{FRESHDESK_DOMAIN}.freshdesk.com/api/v2/tickets/{ticket_id}"
    
    r = requests.get(url, auth=(FRESHDESK_API_KEY, "X"))
    
    if r.status_code != 200:
        return None
    
    ticket_data = r.json()
    
    # Fetch conversations separately (attachments are in conversations)
    conversations_url = f"https://{FRESHDESK_DOMAIN}.freshdesk.com/api/v2/tickets/{ticket_id}/conversations"
    conv_response = requests.get(conversations_url, auth=(FRESHDESK_API_KEY, "X"))
    
    if conv_response.status_code == 200:
        conversations = conv_response.json()
        # Extract attachments from conversations
        attachments = []
        for conv in conversations:
            if 'attachments' in conv:
                attachments.extend(conv['attachments'])
        ticket_data['attachments'] = attachments
    
    return ticket_data

def fetch_ticket_attachments_separately(ticket_id):
    """Fetch attachments using the dedicated attachments endpoint"""
    url = f"https://{FRESHDESK_DOMAIN}.freshdesk.com/api/v2/tickets/{ticket_id}/attachments"
    print(f"DEBUG: Fetching attachments from: {url}")
    
    try:
        r = requests.get(url, auth=(FRESHDESK_API_KEY, "X"))
        print(f"DEBUG: Attachments endpoint status: {r.status_code}")
        
        if r.status_code == 200:
            attachments = r.json()
            print(f"DEBUG: Found {len(attachments)} attachments via attachments endpoint")
            for i, att in enumerate(attachments):
                print(f"DEBUG: Attachment {i+1}: {att.get('name', 'Unknown')} - Size: {att.get('size', 'Unknown')} bytes")
            return attachments
        else:
            print(f"DEBUG: Attachments endpoint error: {r.text}")
            return []
    except Exception as e:
        print(f"DEBUG: Error fetching attachments: {e}")
        return []
# ========== ATTACHMENT DOWNLOAD FUNCTIONS ==========

def download_attachment(attachment_url, save_path=None):
    """
    Download an attachment from Freshdesk using its URL
    
    Args:
        attachment_url: The URL of the attachment
        save_path: Optional path to save the file locally
    
    Returns:
        Dict with file content and metadata
    """
    try:
        # Download the attachment
        response = requests.get(attachment_url, auth=(FRESHDESK_API_KEY, "X"))
        
        if response.status_code != 200:
            print(f"Failed to download attachment: {response.status_code}")
            return None
        
        # Get file info from headers
        content_type = response.headers.get('Content-Type', 'unknown')
        content_disposition = response.headers.get('Content-Disposition', '')
        
        # Extract filename if available
        filename = 'unknown'
        if 'filename=' in content_disposition:
            filename = content_disposition.split('filename=')[1].strip('"')
        
        result = {
            'content': response.content,
            'content_type': content_type,
            'filename': filename,
            'size': len(response.content)
        }
        
        # Save locally if path provided
        if save_path:
            with open(save_path, 'wb') as f:
                f.write(response.content)
            result['saved_to'] = save_path
        
        return result
        
    except Exception as e:
        print(f"Error downloading attachment: {e}")
        return None

def download_all_ticket_attachments(ticket_id, save_directory=None):
    """
    Download all attachments for a ticket
    
    Args:
        ticket_id: The ticket ID
        save_directory: Optional directory to save files
    
    Returns:
        List of downloaded attachments
    """
    # Fetch ticket with attachments
    ticket_data = fetch_ticket_by_id(ticket_id)
    if not ticket_data:
        return []
    
    attachments = ticket_data.get('attachments', [])
    downloaded = []
    
    # Create save directory if specified
    if save_directory and not os.path.exists(save_directory):
        os.makedirs(save_directory)
    
    for i, attachment in enumerate(attachments):
        url = attachment.get('attachment_url')
        name = attachment.get('name', f'attachment_{i}')
        
        if url:
            print(f"Downloading: {name}")
            
            # Determine save path
            save_path = None
            if save_directory:
                # Sanitize filename
                safe_name = "".join(c for c in name if c.isalnum() or c in (' ', '.', '_', '-')).rstrip()
                save_path = os.path.join(save_directory, safe_name)
            
            # Download
            result = download_attachment(url, save_path)
            
            if result:
                result['attachment_id'] = attachment.get('id')
                result['original_name'] = name
                downloaded.append(result)
                print(f"✓ Downloaded: {name} ({result['size']} bytes)")
            else:
                print(f"✗ Failed to download: {name}")
    
    return downloaded

def analyze_downloaded_attachments(ticket_id):
    """
    Download and analyze all attachments for a ticket
    """
    # Download attachments to temp directory
    temp_dir = f"temp_attachments_{ticket_id}"
    downloaded = download_all_ticket_attachments(ticket_id, temp_dir)
    
    # Initialize analyzer
    analyzer = DocumentAnalyzer(
    )
    
    # Analyze each downloaded file
    results = []
    for attachment in downloaded:
        if 'saved_to' in attachment:
            # Analyze the saved file
            analysis = analyzer.analyze_document(
                attachment['saved_to'], 
                None  # No URL needed since we have local file
            )
            analysis['filename'] = attachment['original_name']
            analysis['attachment_id'] = attachment['attachment_id']
            results.append(analysis)
    
    # Clean up temp files
    import shutil
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
    
    return results

# ========== END ATTACHMENT DOWNLOAD FUNCTIONS ==========

def fetch_ticket_conversations(ticket_id, page=1):
    """Fetches a single page of conversations for a given ticket."""
    url = f"https://{FRESHDESK_DOMAIN}.freshdesk.com/api/v2/tickets/{ticket_id}/conversations?page={page}"
    response = requests.get(url, auth=(FRESHDESK_API_KEY, "X"))
    print(f"DEBUG: Response Status Code for conversations of {ticket_id}, page {page}: {response.status_code}")
    if response.status_code != 200:
        print(f"⚠️ Could not fetch conversations for {ticket_id} on page {page}. Response: {response.text}")
        return []
    
    conversations = response.json()
    
    # Debug: Show that created_at and updated_at are included
    if conversations and len(conversations) > 0:
        first_conv = conversations[0]
        print(f"DEBUG: Sample conversation fields include - created_at: {first_conv.get('created_at')}, updated_at: {first_conv.get('updated_at')}")
    
    return conversations

def fetch_all_ticket_conversations(ticket_id):
    """Fetches all conversation pages for a given Freshdesk ticket by handling pagination.
    Each conversation includes created_at and updated_at timestamps."""
    all_conversations = []
    page = 1
    per_page_limit = 30 

    print(f"DEBUG: Starting to fetch all conversations for ticket ID: {ticket_id}")

    while True:
        url = f"https://{FRESHDESK_DOMAIN}.freshdesk.com/api/v2/tickets/{ticket_id}/conversations?page={page}"
        
        try:
            response = requests.get(
                url,
                auth=(FRESHDESK_API_KEY, 'X'),
                headers={"Content-Type": "application/json"}
            )
            response.raise_for_status()
            
            current_page_conversations = response.json()
            
            if not current_page_conversations:
                print(f"DEBUG: No more conversations found on page {page} for ticket {ticket_id}. Ending fetch.")
                break
            
            all_conversations.extend(current_page_conversations)
            print(f"DEBUG: Fetched page {page} for ticket {ticket_id}. Total conversations so far: {len(all_conversations)}")
            
            if len(current_page_conversations) < per_page_limit:
                print(f"DEBUG: Last page reached for ticket {ticket_id} (items < {per_page_limit}). Ending fetch.")
                break
            
            page += 1

        except requests.exceptions.RequestException as e:
            print(f"❌ Error fetching conversations for ticket {ticket_id}, page {page}: {e}")
            break
        except json.JSONDecodeError:
            print(f"❌ Error decoding JSON response for ticket {ticket_id}, page {page}. Response content: {response.text}")
            break
            
    print(f"DEBUG: Finished fetching all conversations for ticket {ticket_id}. Total: {len(all_conversations)}")
    
    # Debug: Show timestamp info for all conversations
    if all_conversations:
        print(f"DEBUG: Timestamp summary for ticket {ticket_id}:")
        for i, conv in enumerate(all_conversations[:3]):  # Show first 3 as sample
            print(f"  Conversation {i+1}: created_at={conv.get('created_at')}, updated_at={conv.get('updated_at')}")
        if len(all_conversations) > 3:
            print(f"  ... and {len(all_conversations) - 3} more conversations")
    
    return all_conversations

def process_conversation_timestamps(conversations):
    """Helper function to extract and process timestamp information from conversations."""
    from datetime import datetime
    
    timestamp_data = []
    for conv in conversations:
        timestamp_info = {
            'id': conv.get('id'),
            'created_at': conv.get('created_at'),
            'updated_at': conv.get('updated_at'),
            'created_at_parsed': None,
            'updated_at_parsed': None
        }
        
        # Parse timestamps if they exist
        try:
            if conv.get('created_at'):
                timestamp_info['created_at_parsed'] = datetime.fromisoformat(conv['created_at'].replace('Z', '+00:00'))
            if conv.get('updated_at'):
                timestamp_info['updated_at_parsed'] = datetime.fromisoformat(conv['updated_at'].replace('Z', '+00:00'))
        except ValueError as e:
            print(f"⚠️ Error parsing timestamp for conversation {conv.get('id')}: {e}")
        
        timestamp_data.append(timestamp_info)
    
    return timestamp_data

def fetch_child_tickets(parent_ticket_id):
    """
    Fetch all child tickets associated with a parent ticket
    using Freshdesk’s Association APIs.
    """
    try:
        # 1) Preferred: Call the dedicated endpoint
        url = f"https://{FRESHDESK_DOMAIN}.freshdesk.com/api/v2/tickets/{parent_ticket_id}/associated_tickets"
        resp = requests.get(url, auth=(FRESHDESK_API_KEY, "X"))
        resp.raise_for_status()
        children = resp.json().get("tickets", [])
        if children:
            print(f"Found {len(children)} child tickets via /associated_tickets")
            return children

        # 2) Fallback: Read the parent’s associated_tickets_list field
        url = f"https://{FRESHDESK_DOMAIN}.freshdesk.com/api/v2/tickets/{parent_ticket_id}"
        resp = requests.get(
            url,
            params={"include": "conversations"},   # optional includes
            auth=(FRESHDESK_API_KEY, "X"),
        )
        resp.raise_for_status()
        parent = resp.json()
        if parent.get("association_type") == 1:
            child_ids = parent.get("associated_tickets_list", [])
            children = []
            for cid in child_ids:
                c = requests.get(
                    f"https://{FRESHDESK_DOMAIN}.freshdesk.com/api/v2/tickets/{cid}",
                    auth=(FRESHDESK_API_KEY, "X"),
                )
                c.raise_for_status()
                children.append(c.json())
            print(f"Found {len(children)} child tickets via associated_tickets_list")
            return children

        # 3) No children found
        print("No child tickets found.")
        return []

    except requests.RequestException as e:
        print(f"Error fetching child tickets: {e}")
        return []

def fetch_parent_ticket(child_ticket_data):
    """
    Extract and fetch parent ticket from child ticket data
    """
    try:
        # Check subject for parent ticket reference
        subject = child_ticket_data.get('subject', '')
        parent_match = re.search(r'#(\d+)', subject)
        if parent_match:
            parent_id = parent_match.group(1)
            return fetch_ticket_by_id(parent_id)
        
        # Check custom fields
        custom_fields = child_ticket_data.get('custom_fields', {})
        parent_id = custom_fields.get('cf_parent_ticket_id')
        if parent_id:
            return fetch_ticket_by_id(parent_id)
        
        return None
        
    except Exception as e:
        print(f"Error fetching parent ticket: {e}")
        return None
    
def get_agent_name_from_id(agent_id):
    """Fetches the name of a Freshdesk agent given their ID."""
    if not agent_id:
        return "Unassigned"

    url = f"https://{FRESHDESK_DOMAIN}.freshdesk.com/api/v2/agents/{agent_id}"
    try:
        response = requests.get(url, auth=(FRESHDESK_API_KEY, "X"))
        if response.status_code == 200:
            agent_data = response.json()
            return f"{agent_data.get('first_name', '')} {agent_data.get('last_name', '')}".strip()
        else:
            print(f"⚠️ Error fetching agent {agent_id}: Status {response.status_code}, Response: {response.text}")
            return "Unknown Agent"
    except requests.exceptions.RequestException as e:
        print(f"❌ Network error fetching agent {agent_id}: {e}")
        return "Network Error Agent"

def classify_error_type(text):
    """Original classification function - kept as fallback."""
    text_lower = text.lower()

    financial_keywords = [
        "utr not shared", "payment pending", "negative balance", "refund",
        "redemption accepted but not paid", "not raised any redemption request",
        "invoice not received", "duplicate redemption request"
    ]
    technical_keywords = [
        "score not updated", "policy not visible", "technical issue",
        "add-ons not updated at itms", "score released", "grid/addendum appiled correct"
    ]
    operations_keywords = [
        "aadhaar", "pan", "kyc", "address proof", "document mismatch",
        "incorrect mmv", "calculation done on od instead of net"
    ]

    # Check categories
    if any(k in text_lower for k in financial_keywords):
        return "Financial"
    if any(k in text_lower for k in technical_keywords):
        return "Technical"
    if any(k in text_lower for k in operations_keywords):
        return "Operations"

    return "Uncategorized"

def initialize_excel_if_needed():
    """Creates the Excel file and header row if it doesn't exist."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Ticket ID", "Subject", "Problem", "Why", "Solution", "Classification", "Cluster"])
        wb.save(EXCEL_FILE)

def get_processed_ticket_ids():
    """Returns a set of ticket IDs already processed in the Excel file."""
    initialize_excel_if_needed()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    return {str(r[0].value) for r in ws.iter_rows(min_row=2) if r[0].value}

def append_to_excel(row):
    """Appends a row of data to the Excel summary file."""
    def sanitize(item):
        if isinstance(item, list):
            return ", ".join(str(i) for i in item)
        return str(item)
    row = [sanitize(cell) for cell in row]
    try:
        initialize_excel_if_needed()
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append(row)
        wb.save(EXCEL_FILE)
    except InvalidFileException:
        print("❌ Invalid Excel file path or corrupted file. Please check if the file is open or damaged.")
    except Exception as e:
        print(f"❌ Error appending to Excel: {e}")

def search_ticket_in_excel(ticket_id):
    """Searches for a ticket ID in the Excel file and returns its summary if found."""
    initialize_excel_if_needed()
    if not os.path.exists(EXCEL_FILE):
        print(f"Excel file not found at {EXCEL_FILE}. Cannot search.")
        return None
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for r in ws.iter_rows(min_row=2):
        if str(r[0].value) == str(ticket_id):
            return {
                "Ticket ID": r[0].value,
                "Subject": r[1].value,
                "Problem": r[2].value,
                "Why": r[3].value,
                "Solution": r[4].value,
                "Classification": r[5].value,
                "Cluster": r[6].value,
            }
    return None

def clean_html(raw):
    """Removes HTML tags from a string."""
    if not raw:
        return ""
    return BeautifulSoup(raw, "html.parser").get_text(separator=" ", strip=True)

def preprocess_image(img):
    """Preprocesses an image for OCR."""
    img = img.convert("L").filter(ImageFilter.MedianFilter())
    return ImageEnhance.Contrast(img).enhance(2)

def flatten(v):
    """Flattens a list into a comma-separated string."""
    return ", ".join(v) if isinstance(v, list) else str(v)

def process_attachment(att):
    """Processes an attachment."""
    url = att.get("attachment_url")
    r = requests.get(url, auth=(FRESHDESK_API_KEY, "X"))
    if r.status_code != 200:
        print(f"⚠️ Failed to download attachment from {url}: Status {r.status_code}")
        return ""
    if "image" in r.headers.get("Content-Type", ""):
        try:
            img = Image.open(BytesIO(r.content))
            img = preprocess_image(img)
            return "(OCR text extracted from image)"
        except Exception as e:
            print(f"❌ Error processing image attachment: {e}")
            return "(Error processing image)"
    elif "text" in r.headers.get("Content-Type", "") or "application/pdf" in r.headers.get("Content-Type", ""):
        return f"(Attachment: {att.get('name')})"
    return ""

def extract_subject_and_description(ticket_data):
    """
    Extracts and returns the subject and description from ticket data for classification.
    This provides the initial context for ticket classification.
    """
    subject = ticket_data.get('subject', '')
    description = clean_html(ticket_data.get('description_text', ticket_data.get('description', '')))
    
    # Combine subject and description for classification
    combined_text = f"{subject} {description}"
    
    return {
        'subject': subject,
        'description': description,
        'combined': combined_text
    }

def classify_ticket_with_subject_priority(ticket_data, conversations_data=None):
    """
    Enhanced classification that prioritizes subject line for initial classification
    and uses description and conversations for confirmation.
    """
    # Extract subject and description
    content = extract_subject_and_description(ticket_data)
    subject_lower = content['subject'].lower()
    
    # First, try to classify based on subject alone (high confidence)
    subject_classification, subject_sop = classify_ticket_with_sop(subject_lower)
    
    # If subject gives a clear classification (not default), use it
    if not subject_classification.startswith(("Uncategorized", "Financial", "Technical", "Operations")):
        print(f"Classification from subject: {subject_classification}")
        
        # Verify with description if available
        if content['description']:
            desc_classification, desc_sop = classify_ticket_with_sop(content['combined'])
            
            # If descriptions confirms or enhances the classification, use the more specific one
            if desc_classification.startswith(subject_classification.split("-")[0]):
                # Same main category, use the more specific classification
                if len(desc_classification.split("-")) > len(subject_classification.split("-")):
                    print(f"Enhanced classification from description: {desc_classification}")
                    return desc_classification, desc_sop
        
        return subject_classification, subject_sop
    
    # If subject doesn't give clear classification, use full content
    if conversations_data:
        # Use full ticket content including conversations
        full_content, _ = extract_email_content_and_attachments(ticket_data, conversations_data)
        return classify_ticket_with_sop(full_content)
    else:
        # Use combined subject and description
        return classify_ticket_with_sop(content['combined'])


def extract_email_content_and_attachments(ticket_data, conversations_data):
    """Combines ticket subject, description, and all conversation content into a single string."""
    full_content_parts = []

    full_content_parts.append(f"Subject: {ticket_data.get('subject', 'N/A')}")
    full_content_parts.append(f"Description: {clean_html(ticket_data.get('description_text', ticket_data.get('description', 'No description provided.')))}")

    actions_taken_list = []
    for conv in conversations_data:
        conv_text = conv.get('body_text') or clean_html(conv.get('body', ''))
        sender = conv.get('from_email', 'Unknown Sender')
        created_at = conv.get('created_at', 'N/A')
        
        message_type = conv.get('category', 'message')
        if not isinstance(message_type, str):
            message_type = str(message_type)
        
        if conv.get('user_id') == ticket_data.get('responder_id') or conv.get('user_id') == ticket_data.get('requester_id'):
            if conv_text.strip():
                if conv.get('private'):
                    actions_taken_list.append(f"Agent Note ({created_at}): {conv_text}")
                elif conv.get('incoming') == False:
                    actions_taken_list.append(f"Agent Reply ({created_at}): {conv_text}")
                elif conv.get('incoming') == True:
                    full_content_parts.append(f"\n--- CUSTOMER {message_type.upper()} from {sender} at {created_at} ---\n{conv_text}")
                else:
                    full_content_parts.append(f"\n--- {message_type.upper()} from {sender} at {created_at} ---\n{conv_text}")
        else:
            if conv_text.strip():
                full_content_parts.append(f"\n--- {message_type.upper()} from {sender} at {created_at} ---\n{conv_text}")

    raw_content = "\n\n".join(full_content_parts)

    MAX_RAW_CONTENT_LENGTH = 100000
    if len(raw_content) > MAX_RAW_CONTENT_LENGTH:
        raw_content = raw_content[:MAX_RAW_CONTENT_LENGTH] + "\n\n... (content truncated)"
        print(f"WARNING: Raw ticket content truncated for LLM input to {MAX_RAW_CONTENT_LENGTH} chars.")

    return raw_content, "\n".join(actions_taken_list)

def chunk_text(text, max_chunk_size=8000):
    """Splits text into chunks of specified word count for LLM processing."""
    words = text.split()
    chunks = []
    for i in range(0, len(words), max_chunk_size):
        chunks.append(" ".join(words[i : i + max_chunk_size]))
    return chunks

def get_claude_answer(ticket_content_text: str, user_question: str, ticket_data=None) -> str:
    """Uses Claude to answer a question about the given ticket content with SOP context."""
    if not anthropic_client:
        return "Error: Claude client not initialized. Cannot answer questions. Please check CLAUDE_API_KEY."

    # Classify the ticket to provide SOP context
    classification, sop_details = classify_ticket_with_sop(ticket_content_text)

    qa_system_prompt = (
        "You are an AI assistant for  Freshdesk ticket analysis with deep knowledge of company SOPs. "
        "You have access to detailed Standard Operating Procedures for: "
        "- Claims handling (Motor, Health, Life, MSME) "
        "- Endorsement processes (Financial and Non-financial) "
        "- Support ticket resolution "
        "- Escalation matrices and TATs "
        "When answering questions, always reference the relevant SOP and provide specific, actionable guidance based on company procedures."
    )

    qa_user_prompt = (
        f"Here is the Freshdesk ticket content:\n\n---\n{ticket_content_text}\n---\n\n"
        f"Ticket Classification: {classification}\n"
        f"Relevant SOP: {json.dumps(sop_details, indent=2) if sop_details else 'Standard procedures apply'}\n\n"
        f"User's Question: {user_question}\n\n"
        "Answer the question using the provided ticket content and applicable SOPs. "
        "If asking about process/procedures, refer to the specific SOP steps. "
        "If the information is not present, say so."
    )

    try:
        message = anthropic_client.messages.create(
            model="claude-3-haiku-20240307",
            max_tokens=1000,
            messages=[
                {"role": "user", "content": qa_user_prompt}
            ],
            system=qa_system_prompt,
            temperature=0.2,
        )
        return message.content[0].text.strip()
    except Exception as e:
        print(f"Error calling Claude for Q&A: {e}")
        return f"An error occurred while trying to answer with Claude: {e}"


def process_ticket_id_for_gui(ticket_id_input):
    """Entry point for GUI to process a ticket ID for the RCA summary."""
    try:
        ticket_id = int(ticket_id_input)
    except ValueError:
        return {"error": "Invalid Ticket ID. Please enter a number."}

    result = process_ticket_id_orignal(ticket_id)
    if result:
        return result
    else:
        return {"error": f"Failed to process or find ticket {ticket_id}. Check logs for details."}

# --- SOP Knowledge Base ---
# This dictionary contains structured SOP information extracted from the documents


# Define the system prompt with SOP context
CLAUDE_SYSTEM_PROMPT = """You are a skilled insurance analyst with access to  Standard Operating Procedures (SOPs). 
Your task is to accurately extract and summarize key information from customer support tickets and provide solutions based on the company's SOPs.

Key SOPs to consider:
1. Claims Process: Motor, Health, Life claims follow specific workflows with defined TATs
2. Endorsement Process: Financial and Non-financial endorsements have different procedures
3. Support Issues: PDPNR, PLNG, PFASP have specific resolution steps
4. Escalation Matrix: Each process has defined escalation levels

When analyzing tickets, identify the category (Claims/Endorsement/Support) and provide SOP-based solutions."""

CLAUDE_CHATBOT_SYSTEM_PROMPT = """You are an AI assistant for A Freshdesk ticket analysis with deep knowledge of company SOPs.
You have access to detailed Standard Operating Procedures for:
- Claims handling (Motor, Health, Life, MSME)
- Endorsement processes (Financial and Non-financial)
- Support ticket resolution
- Escalation matrices and TATs

When answering questions, always reference the relevant SOP and provide specific, actionable guidance based on company procedures.
If the ticket doesn't clearly fall into a known SOP category, analyze the content and suggest the most appropriate process to follow."""

# ========== ENHANCED CLASSIFICATION WITH SOP MAPPING ==========

def classify_ticket_with_sop(text: str) -> Tuple[str, Dict]:
    """
    Comprehensive classification that maps tickets to specific SOPs.
    Returns a tuple of (category, sop_details)
    """
    text_lower = text.lower()
    
    # Define comprehensive keyword mappings for all categories
    
    # 1. CLAIMS KEYWORDS
    claim_keywords = {
        "motor": {
            "primary": ["motor claim", "vehicle claim", "car claim", "bike claim", "two wheeler claim", "four wheeler claim"],
            "secondary": ["accident", "damage", "collision", "theft", "total loss", "own damage", "od claim", "third party", "tp claim"],
            "context": ["garage", "surveyor", "loss date", "claim number", "vehicle repair", "fnol", "first notification", "towing", "workshop", "body shop", "paint work", "denting", "parts replacement"],
            "psu_specific": ["new india", "uiic", "united india", "psu insurer"]
        },
        "health": {
            "primary": ["health claim", "medical claim", "hospitalization claim", "mediclaim"],
            "secondary": ["hospital", "treatment", "surgery", "medical emergency", "illness", "disease", "injury"],
            "context": ["cashless", "reimbursement", "pre-authorization", "discharge", "medical bills", "doctor", "diagnosis", "admission date", "discharge date", "room rent", "icu", "pharmacy bills", "diagnostic tests", "consultation fees"],
            "specific": ["pre-existing disease", "ped", "waiting period", "sub-limit", "co-pay", "deductible"]
        },
        "life": {
            "primary": ["life claim", "death claim", "maturity claim", "survival benefit"],
            "secondary": ["demise", "death certificate", "nominee claim", "policy maturity"],
            "context": ["death certificate", "nominee", "legal heir", "succession certificate", "will", "probate", "claim forms", "discharge voucher"]
        },
        "sme": {
            "primary": ["sme claim", "business claim", "commercial claim", "msme claim"],
            "secondary": ["shop claim", "office claim", "factory claim", "warehouse claim", "business interruption"],
            "context": ["fire", "burglary", "machinery breakdown", "stock damage", "business loss", "property damage", "liability claim"]
        }
    }
    
    # 2. ENDORSEMENT KEYWORDS
    endorsement_keywords = {
        "motor": {
            "financial": {
                "primary": ["ownership transfer", "policy cancellation", "refund request", "add coverage", "remove coverage"],
                "secondary": ["cng kit", "lpg kit", "ncb update", "ncb correction", "idv change", "increase idv", "decrease idv", "wrong ncb", "ncb falsification"],
                "context": ["make model change", "variant change", "pa cover", "personal accident", "add-on cover", "zero dep", "engine protect", "consumables", "invoice cover", "key protect", "tyre protect", "payment charged twice", "duplicate payment", "excess payment", "invalid pyp", "previous policy"]
            },
            "non_financial": {
                "primary": ["name correction", "address change", "contact update", "nominee change"],
                "secondary": ["gstin update", "gst number", "hypothecation add", "hypothecation remove", "bank finance", "loan clearance"],
                "context": ["engine number", "chassis number", "registration number", "vehicle number", "pin code", "email update", "mobile update", "salutation", "digital signature", "risk start date", "policy start date", "vaahan update", "vahan portal"]
            }
        },
        "health": {
            "financial": {
                "primary": ["add member", "delete member", "newborn addition", "sum insured change", "policy cancellation"],
                "secondary": ["ped update", "pre-existing disease", "medical condition update", "dob correction", "age correction"],
                "context": ["family addition", "spouse addition", "parent addition", "child addition", "member deletion", "height weight", "bmi update", "lifestyle change", "occupation change", "policy holder change", "proposer change", "free look cancellation", "mid-term cancellation"]
            },
            "non_financial": {
                "primary": ["name change", "address update", "contact details", "nominee update"],
                "secondary": ["gender correction", "relationship correction", "salutation change", "communication address"],
                "context": ["email id", "mobile number", "pincode", "city change", "state change", "kyc update", "id proof update", "gstin", "pan update", "aadhaar update", "effective date", "policy date correction", "auto debit", "si mandate", "ecs mandate"]
            },
            "miscellaneous": ["health checkup", "health card", "e-card", "policy copy", "premium receipt", "claim ratio", "network hospital", "policy features", "coverage details", "waiting period query", "sub-limit query", "room rent limit", "co-pay details", "policy status", "renewal status", "grace period", "portal issue", "website error", "app issue", "premium calculation", "loading details", "discount query"]
        },
        "life": {
            "financial": {
                "primary": ["rider addition", "rider deletion", "sum assured change", "premium change", "policy cancellation"],
                "secondary": ["premium frequency", "payment mode change", "fund switch", "partial withdrawal"],
                "context": ["term rider", "critical illness", "accidental death benefit", "waiver of premium", "income benefit", "loan against policy", "surrender value", "paid up value", "premium payment term", "policy term change", "revival", "reinstatement"]
            },
            "non_financial": {
                "primary": ["nominee change", "address change", "name correction", "contact update"],
                "secondary": ["assignee update", "beneficiary change", "communication preference"],
                "context": ["email update", "mobile update", "pan correction", "aadhaar update", "bank details", "ecs mandate", "auto debit", "standing instruction", "due date change", "premium due date", "annual mode", "half yearly", "quarterly", "monthly mode"]
            },
            "miscellaneous": ["policy copy", "premium paid receipt", "tax certificate", "80c certificate", "80d certificate", "loan eligibility", "surrender quote", "maturity amount", "bonus details", "nav details", "fund performance", "medical reports", "underwriting query", "revival quote", "grace period", "lapsed policy", "policy status", "premium holiday", "top up", "partial withdrawal status"]
        },
        "msme": {
            "financial": {
                "primary": ["add employee", "delete employee", "sum insured change", "coverage modification", "policy extension"],
                "secondary": ["gmc addition", "gmc deletion", "gpa coverage", "workmen compensation", "stock update"],
                "context": ["employee list", "salary update", "designation change", "location addition", "branch coverage", "risk location", "machinery addition", "stock value update", "building value", "contents update", "liability limit", "policy period extension", "short period cancellation"]
            },
            "non_financial": {
                "primary": ["company name", "address change", "contact person", "authorized signatory"],
                "secondary": ["gstin update", "pan update", "cin update", "registration details"],
                "context": ["bank details", "hypothecation", "mortgage details", "email id", "phone number", "branch address", "head office", "registered office", "factory address", "warehouse location", "directors details", "partners details", "proprietor details"]
            },
            "miscellaneous": ["employee cards", "id cards", "uhid list", "active employee list", "deleted employee list", "premium calculation", "experience report", "claim ratio", "renewal quote", "coverage certificate", "policy wordings", "endorsement copy", "debit note", "credit note"]
        }
    }
    
    # 3. SUPPORT QUERY KEYWORDS
    support_keywords = {
        "pdpnr": {
            "primary": ["payment done policy not received", "pdpnr", "payment successful no policy", "policy not generated"],
            "secondary": ["pdf not received", "policy document pending", "payment confirmed but", "transaction successful but"],
            "context": ["payment receipt", "transaction id", "payment reference", "utr number", "payment screenshot", "bank statement", "credit card statement", "debit confirmation"]
        },
        "plng": {
            "primary": ["payment link not generated", "plng", "link not received", "payment link issue"],
            "secondary": ["proposal not submitting", "proposal stuck", "payment page error", "unable to proceed payment"],
            "context": ["api down", "technical error", "underwriting issue", "kyc pending", "validation error", "system error", "timeout error", "session expired", "quote expired"]
        },
        "pfasp": {
            "primary": ["payment failed after successful", "pfasp", "payment deducted but failed", "amount debited but"],
            "secondary": ["transaction failed", "payment gateway error", "payment reversed", "refund pending"],
            "context": ["bank deducted", "amount debited", "pg error", "gateway timeout", "technical failure", "reconciliation", "payment status", "failed transaction"]
        },
        "kyc": {
            "primary": ["kyc issue", "kyc pending", "kyc verification", "vkyc problem"],
            "secondary": ["identity verification", "document verification", "aadhaar verification", "pan verification"],
            "context": ["video kyc", "ckyc", "ckycr", "otp issue", "biometric", "face match", "document upload", "unclear document", "verification failed", "mismatch error"]
        },
        "general_support": {
            "primary": ["portal issue", "website problem", "app not working", "login issue"],
            "secondary": ["technical problem", "system issue", "unable to access", "error message"],
            "context": ["pos portal", "broker portal", "customer portal", "mobile app", "payment gateway", "otp not received", "password reset", "forgot password", "account locked", "session timeout"]
        }
    }
    
    # 4. COMMON REQUEST KEYWORDS
    common_request_keywords = {
        "pi_request": {
            "primary": ["pi report", "pre inspection", "inspection report", "vehicle inspection"],
            "secondary": ["back documents", "supporting documents", "pi photos", "inspection photos"],
            "context": ["vehicle photos", "chassis photo", "engine photo", "odometer", "rc copy", "previous policy", "form 29", "form 30", "noc", "hypothecation letter", "finance noc"]
        },
        "cashless_garage": {
            "primary": ["cashless garage", "network garage", "preferred garage", "garage list"],
            "secondary": ["workshop list", "authorized garage", "panel garage", "tie up garage"],
            "context": ["near me", "in my area", "city wise", "location wise", "contact details", "garage address", "workshop number", "pickup drop", "towing facility"]
        },
        "cashless_hospital": {
            "primary": ["cashless hospital", "network hospital", "empanelled hospital", "hospital list"],
            "secondary": ["tpa hospital", "preferred hospital", "panel hospital", "tie up hospital"],
            "context": ["near me", "in my city", "specialty hospital", "super specialty", "clinic", "diagnostic center", "day care", "hospital address", "hospital contact", "pre authorization"]
        },
        "document_request": {
            "primary": ["policy copy", "soft copy", "policy document", "insurance copy"],
            "secondary": ["endorsement copy", "renewal notice", "debit note", "credit note"],
            "context": ["email policy", "download policy", "policy pdf", "coverage note", "certificate", "tax receipt", "gst invoice", "premium receipt", "payment receipt"]
        },
        "information_request": {
            "primary": ["policy details", "coverage details", "policy status", "premium details"],
            "secondary": ["benefits", "features", "terms conditions", "exclusions"],
            "context": ["sum insured", "policy period", "premium amount", "next due date", "claim history", "ncb details", "add on covers", "deductible", "waiting period", "sub limits"]
        }
    }
    
    # Enhanced classification logic with priority order
    
    # 1. Check for specific common requests first (highest priority)
    for request_type, keyword_groups in common_request_keywords.items():
        match_score = 0
        for priority, keywords in keyword_groups.items():
            if any(keyword in text_lower for keyword in keywords):
                match_score += 3 if priority == "primary" else 2 if priority == "secondary" else 1
        
        if match_score >= 3:  # Strong match
            if request_type == "pi_request":
                return "Claims-PI-Request", SOP_KNOWLEDGE_BASE["claims"]["common_requests"]["pi_request"]
            elif request_type == "cashless_garage":
                return "Claims-Cashless-Garage", SOP_KNOWLEDGE_BASE["claims"]["common_requests"]["cashless_garage"]
            elif request_type == "cashless_hospital":
                return "Claims-Cashless-Hospital", SOP_KNOWLEDGE_BASE["claims"]["common_requests"]["cashless_hospital"]
            elif request_type == "document_request":
                return "General-Document-Request", {"process": ["Check policy type", "Verify customer identity", "Send document via registered email", "Update ticket as resolved"], "tat": "2 hours"}
            elif request_type == "information_request":
                return "General-Information-Request", {"process": ["Identify query type", "Fetch policy details from ITMS", "Provide accurate information", "Offer additional assistance"], "tat": "2 hours"}
    
    # 2. Check for support queries (high priority)
    for issue_type, keyword_groups in support_keywords.items():
        match_score = 0
        for priority, keywords in keyword_groups.items():
            if any(keyword in text_lower for keyword in keywords):
                match_score += 3 if priority == "primary" else 2 if priority == "secondary" else 1
        
        if match_score >= 2:  # Medium-strong match
            if issue_type == "general_support":
                return "Support-General", {"description": "General technical support query", "solution": "Identify specific issue, provide troubleshooting steps, escalate to tech team if needed"}
            else:
                return f"Support-{issue_type.upper()}", SOP_KNOWLEDGE_BASE["support"].get(issue_type, {})
    
    # 3. Check for claims (medium priority)
    for claim_type, keyword_groups in claim_keywords.items():
        match_score = 0
        for priority, keywords in keyword_groups.items():
            if any(keyword in text_lower for keyword in keywords):
                match_score += 3 if priority == "primary" else 2 if priority == "secondary" else 1
        
        if match_score >= 2:  # Medium match sufficient for claims
            return f"Claims-{claim_type.title()}", SOP_KNOWLEDGE_BASE["claims"].get(claim_type, {})
    
    # 4. Check for endorsements (requires more specific matching)
    best_endorsement_match = None
    best_endorsement_score = 0
    best_endorsement_sub_type = None
    
    for endorsement_type, subcategories in endorsement_keywords.items():
        for sub_type, keyword_groups in subcategories.items():
            if sub_type == "miscellaneous":
                # Handle miscellaneous as a list
                if isinstance(keyword_groups, list):
                    if any(keyword in text_lower for keyword in keyword_groups):
                        score = 2
                        if score > best_endorsement_score:
                            best_endorsement_score = score
                            best_endorsement_match = endorsement_type
                            best_endorsement_sub_type = "miscellaneous"
            else:
                # Handle financial/non-financial with priority groups
                match_score = 0
                for priority, keywords in keyword_groups.items():
                    if any(keyword in text_lower for keyword in keywords):
                        match_score += 3 if priority == "primary" else 2 if priority == "secondary" else 1
                
                if match_score > best_endorsement_score:
                    best_endorsement_score = match_score
                    best_endorsement_match = endorsement_type
                    best_endorsement_sub_type = sub_type
    
    if best_endorsement_score >= 2:  # Medium match for endorsements
        return f"Endorsement-{best_endorsement_match.title()}-{best_endorsement_sub_type.title()}", \
               SOP_KNOWLEDGE_BASE["endorsement"].get(best_endorsement_match, {})
    
    # 5. Fallback classification based on general keywords
    general_classifications = {
        "payment": "Support-Payment-Issue",
        "policy": "General-Policy-Query", 
        "claim": "Claims-General",
        "document": "General-Document-Request",
        "endorse": "Endorsement-General",
        "cancel": "Endorsement-Cancellation",
        "refund": "Endorsement-Refund",
        "query": "General-Query",
        "complaint": "General-Complaint",
        "feedback": "General-Feedback"
    }
    
    for keyword, classification in general_classifications.items():
        if keyword in text_lower:
            if classification.startswith("Claims"):
                return classification, SOP_KNOWLEDGE_BASE.get("claims", {}).get("general_claim_handling_procedure", {})
            elif classification.startswith("Endorsement"):
                return classification, SOP_KNOWLEDGE_BASE.get("endorsement", {})
            else:
                return classification, {"process": ["Understand query", "Provide information", "Escalate if needed"], "tat": "2 hours"}
    
    # 6. Default classification using original error type function
    return classify_error_type(text), {"process": ["Follow standard SOP", "Identify correct category", "Route to appropriate team"], "tat": "As per SOP"}

def generate_sop_based_solution(classification: str, sop_details: Dict, problem_text: str) -> str:
    """
    Generates a solution based on the SOP for the classified ticket type.
    """
    solution_parts = []
    
    if classification.startswith("Claims"):
        if "Cashless" in classification:
            solution_parts.append(f"Follow SOP: {sop_details.get('process', ['Share details within 2 hours'])[0]}")
            if "resources" in sop_details and "list" in sop_details["resources"]:
                solution_parts.append(f"Resource: {sop_details['resources']['list']}")
        else:
            claim_type = classification.split("-")[1].lower()
            if claim_type in ["motor", "health", "life", "sme"]:
                solution_parts.append(f"Follow {claim_type.title()} Claims SOP:")
                if "process_flow" in sop_details:
                    # Show first 3 key steps
                    key_steps = sop_details["process_flow"][:3]
                    solution_parts.append("Process: " + " → ".join(key_steps))
                if "tat" in sop_details:
                    if "call_customer" in sop_details["tat"]:
                        solution_parts.append(f"TAT: Call within {sop_details['tat']['call_customer']}")
                    elif "manual_revert" in sop_details["tat"]:
                        solution_parts.append(f"TAT: Revert within {sop_details['tat']['manual_revert']}")
    
    elif classification.startswith("Endorsement"):
        parts = classification.split("-")
        if len(parts) >= 2:
            endorsement_type = parts[1].lower()
            solution_parts.append(f"Follow {endorsement_type.title()} Endorsement SOP:")
            
            # Add process information
            if "process" in sop_details:
                if isinstance(sop_details["process"], dict):
                    if "email" in sop_details["process"]:
                        solution_parts.append(f"Process: {sop_details['process']['email']}")
                elif isinstance(sop_details["process"], list):
                    solution_parts.append(f"Process: {' → '.join(sop_details['process'][:2])}")
                else:
                    solution_parts.append(f"Process: {sop_details['process']}")
            
            # Add specific TAT if available
            if len(parts) >= 3 and parts[2].lower() in sop_details:
                sub_details = sop_details[parts[2].lower()]
                if isinstance(sub_details, dict) and "tat" in sub_details:
                    # Find a relevant TAT from the ticket content
                    for tat_key, tat_value in sub_details["tat"].items():
                        if any(word in problem_text.lower() for word in tat_key.split("_")):
                            solution_parts.append(f"TAT: {tat_value}")
                            break
                    else:
                        # If no specific match, show first TAT
                        first_tat = list(sub_details["tat"].values())[0]
                        solution_parts.append(f"TAT varies by type: {first_tat}")
    
    elif classification.startswith("Support"):
        issue_type = classification.split("-")[1]
        if "solution" in sop_details:
            solution_parts.append(f"Support Issue ({issue_type}): {sop_details['solution']}")
        if "causes" in sop_details:
            solution_parts.append(f"Common causes: {', '.join(sop_details['causes'])}")
    
    # If no specific SOP found, provide general guidance
    if not solution_parts:
        solution_parts.append("Follow standard SOP. Email support@insurancedekho.com or escalate as per matrix.")
    
    return " ".join(solution_parts)


# ========== Modified Claude Summary Function ==========

def get_claude_summary(text):
    """
    Enhanced version that incorporates SOP knowledge into summaries.
    """
    if not anthropic_client:
        print("Claude client not initialized. Cannot generate summary.")
        return {"Problem": "Claude client not available.", "Why": "API key missing or invalid.", "Solution": "Check API key."}

    # First, classify the ticket to understand which SOP applies
    classification, sop_details = classify_ticket_with_sop(text)
    
    print(f"Ticket classified as: {classification}")
    print("Starting SOP-enhanced summary generation...")
    
    chunks = chunk_text(text)
    print(f"Split into {len(chunks)} chunks.")

    # Enhanced user prompt that includes SOP context
    user_prompt_tpl_chunk = (
        f"This appears to be a {classification} ticket. "
        "Extract the following information from the given text. Your output MUST be a JSON object with exactly these keys:\n"
        "- Problem (≤25 words): clearly stated main issue\n"
        "- Why (≤25 words): root cause based on the ticket content\n"
        "- Solution (≤25 words): specific steps based on InsuranceDekho SOPs\n\n"
        "For the Solution, consider the relevant SOP procedures and TATs.\n"
        "Return only the JSON. Do not explain anything.\n\n"
        "Chunk Text:\n\"\"\"{chunk}\"\"\""
    )

    def summarize_chunk(chunk, idx, max_retries=3):
        """Summarizes a single text chunk using Claude with SOP context."""
        out = ""
        for attempt in range(1, max_retries + 1):
            try:
                messages = [
                    {"role": "user", "content": user_prompt_tpl_chunk.format(chunk=chunk)}
                ]
                res = anthropic_client.messages.create(
                    model="claude-3-haiku-20240307",
                    max_tokens=1000,
                    messages=messages,
                    system=CLAUDE_SYSTEM_PROMPT,
                )
                
                out = res.content[0].text.strip()

                try:
                    result = json.loads(out)
                except json.JSONDecodeError:
                    if out.startswith("```json"):
                        out = out.strip("```json").strip("` \n")
                    result = json.loads(out)

                if not (isinstance(result, dict) and all(k in result for k in ["Problem", "Why", "Solution"])):
                    raise ValueError(f"Missing expected keys or invalid JSON structure from Claude response.")

                # Enhance the solution with SOP details if it's generic
                if len(result.get("Solution", "")) < 10 or "refer" in result.get("Solution", "").lower():
                    result["Solution"] = generate_sop_based_solution(classification, sop_details, chunk)

                return result
            except Exception as e:
                print(f"⚠️ Chunk {idx} attempt {attempt} failed: {e}")
                if attempt == max_retries:
                    # Return SOP-based solution as fallback
                    return {
                        "Problem": "Error extracting problem",
                        "Why": "Processing error",
                        "Solution": generate_sop_based_solution(classification, sop_details, text)
                    }
                time.sleep(1)
        return {"Problem": "", "Why": "", "Solution": ""}

    print("Summarizing chunks in parallel...")
    partials = [None] * len(chunks)
    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = {executor.submit(summarize_chunk, chunk, i): i - 1 for i, chunk in enumerate(chunks, 1)}
        for fut in as_completed(futures):
            idx = futures[fut]
            try:
                partials[idx] = fut.result()
            except Exception as e:
                print(f"Thread error in chunk summarization: {e}")
                partials[idx] = {"Problem": "", "Why": "", "Solution": ""}

    valid = [p for p in partials if p and any(p[f].strip() for f in ("Problem", "Why", "Solution"))]
    if not valid:
        return {
            "Problem": "(no content)",
            "Why": "(no content)",
            "Solution": generate_sop_based_solution(classification, sop_details, text)
        }
    if len(valid) == 1:
        return valid[0]

    # Enhanced merge prompt with SOP context
    user_prompt_tpl_merge = (
        f"Merge the following JSON summaries for a {classification} ticket into a final one. "
        "Combine all 'Problem', 'Why', and 'Solution' fields, preserving distinct insights. "
        "For the Solution field, ensure it aligns with InsuranceDekho SOPs for this ticket type. "
        "Respond only with a single JSON object with these keys: Problem, Why, Solution.\n\n"
        f"Here are the summaries to merge:\n{json.dumps(valid, indent=2)}"
    )

    def merge_jsons(fname, max_retries=2):
        """Merges multiple JSON summaries using Claude."""
        out = ""
        for attempt in range(1, max_retries + 1):
            try:
                messages = [
                    {"role": "user", "content": user_prompt_tpl_merge}
                ]
                res = anthropic_client.messages.create(
                    model="claude-3-haiku-20240307",
                    max_tokens=1000,
                    messages=messages,
                    system=CLAUDE_SYSTEM_PROMPT,
                )
                out = res.content[0].text.strip()

                try:
                    result = json.loads(out)
                except json.JSONDecodeError:
                    if out.startswith("```json"):
                        out = out.strip("```json").strip("` \n")
                    result = json.loads(out)

                if isinstance(result, list):
                    flat = {"Problem": [], "Why": [], "Solution": []}
                    for item in result:
                        flat["Problem"].append(item.get("Problem", ""))
                        flat["Why"].append(item.get("Why", ""))
                        flat["Solution"].append(item.get("Solution", ""))
                    
                    merged = {
                        "Problem": ", ".join(filter(None, sorted(list(set(s.strip() for s in flat["Problem"] if s.strip()))))),
                        "Why": ", ".join(filter(None, sorted(list(set(s.strip() for s in flat["Why"] if s.strip()))))),
                        "Solution": ", ".join(filter(None, sorted(list(set(s.strip() for s in flat["Solution"] if s.strip()))))),
                    }
                    
                    # Ensure solution is SOP-compliant
                    if not merged["Solution"] or len(merged["Solution"]) < 10:
                        merged["Solution"] = generate_sop_based_solution(classification, sop_details, text)
                    
                    return merged
                
                cleaned_result = {}
                for key in ["Problem", "Why", "Solution"]:
                    value = result.get(key, "").strip()
                    cleaned_result[key] = value if value else ""
                
                # Ensure solution is SOP-compliant
                if not cleaned_result["Solution"] or len(cleaned_result["Solution"]) < 10:
                    cleaned_result["Solution"] = generate_sop_based_solution(classification, sop_details, text)
                
                return cleaned_result

            except Exception as e:
                print(f"⚠️ Merge attempt {attempt} failed: {e}")
                if attempt == max_retries:
                    return {
                        "Problem": "Merge failed",
                        "Why": "Processing error",
                        "Solution": generate_sop_based_solution(classification, sop_details, text)
                    }
                time.sleep(1)
        return {"Problem": "", "Why": "", "Solution": ""}

    return merge_jsons("merge_failed.txt")

# ========== Enhanced NLP Query Processing ==========

def process_nlp_query(ticket_id, user_question):
    """
    Enhanced version that uses SOP knowledge to answer questions.
    """
    if not anthropic_client:
        return "Error: Claude AI service is not available. Please check the API key."

    print(f"Processing NLP query for ticket {ticket_id}: '{user_question}'")

    ticket = fetch_ticket_by_id(ticket_id)
    if not ticket:
        return f"Could not retrieve details for ticket ID {ticket_id} from Freshdesk."

    conversations = fetch_all_ticket_conversations(ticket_id)
    ticket_content, _ = extract_email_content_and_attachments(ticket, conversations)
    
    if not ticket_content.strip():
        return f"No detailed content found for ticket ID {ticket_id} to answer your question."

    # Classify the ticket to provide SOP context
    classification, sop_details = classify_ticket_with_sop(ticket_content)
    
    # Enhanced prompt with SOP context
    prompt_content = f"""
    Here is the full content of Freshdesk Ticket ID: {ticket_id}
    <ticket_content>
    {ticket_content}
    </ticket_content>

    This ticket has been classified as: {classification}
    
    Relevant SOP information:
    {json.dumps(sop_details, indent=2) if sop_details else "No specific SOP details available"}

    Based on the ticket content and applicable SOPs, please answer the following question:
    <user_question>
    {user_question}
    </user_question>

    When answering:
    1. First check if the answer is in the ticket content
    2. If asking about process/next steps, refer to the relevant SOP
    3. If asking about TAT/timeline, provide specific SOP-defined timelines
    4. If asking about escalation, provide the escalation matrix details
    
    If the information needed is not in the ticket or SOPs, state that clearly.
    """

    try:
        messages = [
            {"role": "user", "content": prompt_content}
        ]
        
        response = anthropic_client.messages.create(
            model="claude-3-haiku-20240307",
            max_tokens=1500,
            messages=messages,
            system=CLAUDE_CHATBOT_SYSTEM_PROMPT,
            temperature=0.2,
        )
        
        claude_answer = response.content[0].text.strip()
        print(f"Claude's response: {claude_answer}")
        return claude_answer

    except anthropic.APIError as e:
        print(f"❌ Claude API error: {e}")
        return f"An error occurred while communicating with the AI. Details: {e}"
    except Exception as e:
        print(f"❌ Unexpected error during NLP query processing: {e}")
        return f"An unexpected error occurred: {e}"

def calculate_workflow_progress(ticket_data, sop_steps=None):
    """
    Calculate workflow progress based on ticket status and SOP completion
    
    Args:
        ticket_data: Dict containing ticket information
        sop_steps: Optional list of SOP steps for the ticket category
    
    Returns:
        Dict with progress percentage and details
    """
    status = ticket_data.get('status', 2)  # Default to Open
    status_name = status_map.get(status, 'Unknown')
    
    # Base progress by status
    status_progress = {
        1: 10,   # New - 10%
        2: 20,   # Open - 20%
        3: 50,   # Pending - 50%
        4: 90,   # Resolved - 90%
        5: 100   # Closed - 100%
    }
    
    base_progress = status_progress.get(status, 0)
    
    # If ticket is closed or resolved, return the base progress
    if status in [4, 5]:
        return {
            'progress': base_progress,
            'status': status_name,
            'completed_steps': sop_steps if sop_steps else [],
            'remaining_steps': [],
            'current_step': 'Complete' if status == 5 else 'Resolution verification'
        }
    
    # For open tickets, calculate based on SOP steps if available
    if sop_steps and status in [1, 2, 3]:
        # Extract completed actions from ticket
        actions_taken = ticket_data.get('actions_taken', [])
        if isinstance(actions_taken, str):
            actions_taken = [actions_taken]
        
        # Estimate completed steps based on actions and keywords
        completed_count = 0
        current_step = None
        remaining_steps = []
        
        for i, step in enumerate(sop_steps):
            step_completed = False
            
            # Check if this step appears to be completed
            for action in actions_taken:
                if isinstance(action, str):
                    # Simple keyword matching - you can enhance this
                    keywords = step.lower().split()[:3]  # First 3 words
                    if any(keyword in action.lower() for keyword in keywords):
                        step_completed = True
                        break
            
            if step_completed:
                completed_count += 1
            elif not current_step:
                current_step = step
                remaining_steps = sop_steps[i:]
        
        # Calculate progress within the status range
        if len(sop_steps) > 0:
            sop_completion = (completed_count / len(sop_steps)) * 100
            
            # Blend status progress with SOP progress
            if status == 1:  # New
                progress = 10 + (sop_completion * 0.1)  # 10-20%
            elif status == 2:  # Open
                progress = 20 + (sop_completion * 0.5)  # 20-70%
            elif status == 3:  # Pending
                progress = 50 + (sop_completion * 0.4)  # 50-90%
            else:
                progress = base_progress
        else:
            progress = base_progress
        
        return {
            'progress': min(progress, 100),
            'status': status_name,
            'completed_steps': completed_count,
            'total_steps': len(sop_steps),
            'remaining_steps': remaining_steps,
            'current_step': current_step or 'Awaiting action'
        }
    
    return {
        'progress': base_progress,
        'status': status_name,
        'completed_steps': [],
        'remaining_steps': [],
        'current_step': 'Processing'
    }
def format_predictions(predictions):
    """
    Format prediction values for better display
    
    Args:
        predictions: Dict containing prediction data
    
    Returns:
        Formatted predictions dict
    """
    if not predictions:
        return {}
    
    formatted = {}
    
    # Format escalation risk
    if 'escalation_risk' in predictions:
        # Ensure it's a percentage between 0-100
        risk = float(predictions['escalation_risk'])
        if risk < 1:  # Assume it's a decimal if less than 1
            risk = risk * 100
        formatted['escalation_risk'] = round(risk, 1)
    
    # Format resolution time
    if 'estimated_resolution_time' in predictions:
        resolution = predictions['estimated_resolution_time']
        if isinstance(resolution, dict) and 'hours' in resolution:
            hours = float(resolution['hours'])
            # Round to reasonable precision
            if hours < 1:
                formatted['estimated_resolution_time'] = {'hours': round(hours, 2), 'display': f'{round(hours * 60)} mins'}
            elif hours < 24:
                formatted['estimated_resolution_time'] = {'hours': round(hours, 1), 'display': f'{round(hours, 1)} hrs'}
            else:
                days = hours / 24
                formatted['estimated_resolution_time'] = {'hours': round(hours, 1), 'display': f'{round(days, 1)} days'}
        elif isinstance(resolution, (int, float)):
            hours = float(resolution)
            if hours < 1:
                formatted['estimated_resolution_time'] = {'hours': round(hours, 2), 'display': f'{round(hours * 60)} mins'}
            elif hours < 24:
                formatted['estimated_resolution_time'] = {'hours': round(hours, 1), 'display': f'{round(hours, 1)} hrs'}
            else:
                days = hours / 24
                formatted['estimated_resolution_time'] = {'hours': round(hours, 1), 'display': f'{round(days, 1)} days'}
    
    # Format satisfaction risk
    if 'customer_satisfaction_risk' in predictions:
        risk = float(predictions['customer_satisfaction_risk'])
        if risk < 1:  # Assume it's a decimal if less than 1
            risk = risk * 100
        formatted['customer_satisfaction_risk'] = round(risk, 1)
    
    # Format automation potential
    if 'automation_potential' in predictions:
        potential = float(predictions['automation_potential'])
        if potential > 1:  # If it's already a percentage
            potential = potential / 100
        formatted['automation_potential'] = potential
    
    return formatted

def get_sop_steps_for_category(category):
    """
    Get SOP steps for a given ticket category
    
    Args:
        category: The ticket category/classification
    
    Returns:
        List of SOP steps
    """
    # This is a simplified example - in real implementation, 
    # this would fetch from a database or configuration
    sop_database = {
        'password_reset': [
            'Verify user identity',
            'Check account status',
            'Send password reset link',
            'Confirm email sent',
            'Wait for user confirmation',
            'Verify password changed',
            'Update ticket status',
            'Send satisfaction survey'
        ],
        'billing_inquiry': [
            'Review account details',
            'Check billing history',
            'Identify billing issue',
            'Calculate adjustment if needed',
            'Apply correction',
            'Document changes',
            'Notify customer',
            'Follow up confirmation'
        ],
        'technical_support': [
            'Gather system information',
            'Reproduce issue',
            'Check known issues database',
            'Perform diagnostics',
            'Apply fix or workaround',
            'Test solution',
            'Document resolution',
            'Customer confirmation',
            'Close ticket'
        ],
        'general': [
            'Acknowledge ticket',
            'Categorize issue',
            'Assign to specialist',
            'Investigation',
            'Solution implementation',
            'Quality check',
            'Customer communication',
            'Closure'
        ]
    }
    
    # Try to match category
    category_lower = str(category).lower() if category else ''
    
    for key in sop_database:
        if key in category_lower:
            return sop_database[key]
    
    # Default SOP steps
    return sop_database['general']

class WorkflowEngine:
    def get_workflow_status(self, workflow_id):
        """
        Get the current status of a workflow
        """
        # In real implementation, this would fetch from a database
        # For now, return mock data based on the workflow
        return {
            'progress': 0,  # Will be overridden by calculate_workflow_progress
            'next_action': {
                'action': 'Processing'
            }
        }

# ========== Enhanced Process Ticket Function ==========

def process_ticket_id_orignal(ticket_id):
    """
    Enhanced version that includes SOP-based classification and solutions.
    Now uses subject line for better classification.
    """
    initialize_excel_if_needed()

    rec = search_ticket_in_excel(ticket_id)
    
    ticket_data = None

    if rec:
        print(f"Ticket {ticket_id} already processed and found in Excel.")
        ticket_data = fetch_ticket_by_id(ticket_id)
        if not ticket_data:
            print(f"[Error] Even though in Excel, ticket {ticket_id} could not be fetched from Freshdesk.")
            rec["raw_ticket_content"] = "Could not fetch raw content for existing ticket."
            rec["status"] = 0
            rec["Assignee"] = "N/A"
            rec["Actions Taken"] = "N/A"
            rec["ticket_url"] = f"https://{FRESHDESK_DOMAIN}.freshdesk.com/a/tickets/{ticket_id}"
            rec["sop_category"] = "Unknown"
            return rec

        conversations = fetch_all_ticket_conversations(ticket_id)
        raw_ticket_content, actions_taken = extract_email_content_and_attachments(ticket_data, conversations)
        
        # Classify with SOP using subject priority
        classification, _ = classify_ticket_with_subject_priority(ticket_data, conversations)
        
        rec["raw_ticket_content"] = raw_ticket_content
        rec["status"] = ticket_data.get("status", 0)
        rec["agent_id"] = ticket_data.get("responder_id")
        rec["Assignee"] = get_agent_name_from_id(rec["agent_id"])
        rec["Actions Taken"] = actions_taken
        rec["ticket_url"] = f"https://{FRESHDESK_DOMAIN}.freshdesk.com/a/tickets/{ticket_id}"
        rec["sop_category"] = classification
        return rec

    print(f"Processing new ticket: {ticket_id}...")
    ticket_data = fetch_ticket_by_id(ticket_id)
    if not ticket_data:
        print(f"[Error] Ticket ID {ticket_id} could not be fetched from Freshdesk.")
        return None
    
    # Log subject for debugging
    print(f"Ticket Subject: {ticket_data.get('subject', 'N/A')}")
    
    conversations = fetch_all_ticket_conversations(ticket_id)
    txt, actions_taken = extract_email_content_and_attachments(ticket_data, conversations)
    
    if not txt.strip():
        print(f"[Error] No extractable text content found for Ticket ID {ticket_id}.")
        agent_id = ticket_data.get("responder_id")
        assignee_name = get_agent_name_from_id(agent_id)
        
        # Try to classify based on subject alone
        classification, _ = classify_ticket_with_subject_priority(ticket_data)
        
        return {
            "Ticket ID": str(ticket_id),
            "Subject": ticket_data.get("subject", ""),
            "Problem": "No extractable text content for summary.",
            "Why": "N/A",
            "Solution": "N/A",
            "Classification": classification,
            "Cluster": "No Content",
            "Assignee": assignee_name,
            "Actions Taken": "No specific actions logged.",
            "ticket_url": f"https://{FRESHDESK_DOMAIN}.freshdesk.com/a/tickets/{ticket_id}",
            "raw_ticket_content": txt,
            "status": ticket_data.get("status", 0),
            "agent_id": agent_id,
            "sop_category": classification
        }

    # Get enhanced summary with SOP context
    summ = get_claude_summary(txt)
    if not isinstance(summ, dict):
        print(f"[Error] Summary for Ticket ID {ticket_id} is not a valid dictionary.")
        agent_id = ticket_data.get("responder_id")
        assignee_name = get_agent_name_from_id(agent_id)
        classification, _ = classify_ticket_with_subject_priority(ticket_data, conversations)
        return {
            "Ticket ID": str(ticket_id),
            "Subject": ticket_data.get("subject", ""),
            "Problem": "Failed to get Claude summary.",
            "Why": "AI Error",
            "Solution": "AI Error",
            "Classification": classification,
            "Cluster": "AI Error",
            "Assignee": assignee_name,
            "Actions Taken": "AI summary failed, no actions extracted.",
            "ticket_url": f"https://{FRESHDESK_DOMAIN}.freshdesk.com/a/tickets/{ticket_id}",
            "raw_ticket_content": txt,
            "status": ticket_data.get("status", 0),
            "agent_id": agent_id,
            "sop_category": classification
        }

    prob = summ.get("Problem", "Problem not found.")
    why = summ.get("Why", "Why not found.")
    sol = summ.get("Solution", "Solution not found.")
    
    # Enhanced classification with SOP using subject priority
    classification, _ = classify_ticket_with_subject_priority(ticket_data, conversations)
    
    clus = "No Cluster"
    if cluster_model and embedding_model:
        try:
            # Use subject + summary for clustering
            cluster_text = f"{ticket_data.get('subject', '')} {prob} {why} {sol}"
            emb = embedding_model.encode(cluster_text, convert_to_tensor=False)
            clus = cluster_model.predict([emb])[0]
        except Exception as e:
            print(f"⚠️ Error during clustering for ticket {ticket_id}: {e}")
            clus = "Clustering Error"

    agent_id = ticket_data.get("responder_id")
    assignee_name = get_agent_name_from_id(agent_id)

    row = [ticket_id, ticket_data.get("subject", ""), prob, why, sol, classification, clus]
    append_to_excel(row)
    print(f"Successfully processed and saved ticket {ticket_id}.")

    return {
        "Ticket ID": str(ticket_id),
        "Subject": ticket_data.get("subject", ""),
        "Problem": prob,
        "Why": why,
        "Solution": sol,
        "Classification": classification,
        "Cluster": clus,
        "Assignee": assignee_name,
        "Actions Taken": actions_taken,
        "ticket_url": f"https://{FRESHDESK_DOMAIN}.freshdesk.com/a/tickets/{ticket_id}",
        "raw_ticket_content": txt,
        "status": ticket_data.get("status", 0),
        "agent_id": agent_id,
        "sop_category": classification
    }
# Add these new functions to Freshdeskintegration9.py

# ========== AUTONOMOUS ACTION SYSTEM ==========

class AutonomousActionSystem:
    """System for generating and executing autonomous actions based on SOPs"""
    
    def __init__(self, freshdesk_domain, freshdesk_api_key):
        self.freshdesk_domain = freshdesk_domain
        self.freshdesk_api_key = freshdesk_api_key
        self.action_log = []
    
    def analyze_ticket_for_actions(self, ticket_data: dict, classification: str, sop_details: dict) -> list:
        """Analyze ticket and generate recommended actions based on SOPs"""
        actions = []
        
        # Extract ticket details
        ticket_id = ticket_data.get('Ticket ID')
        status = ticket_data.get('status', 0)
        created_at = ticket_data.get('created_at')
        subject = ticket_data.get('Subject', '')
        
        # Check ticket age for TAT compliance
        if created_at:
            ticket_age = self._calculate_ticket_age(created_at)
        else:
            ticket_age = 0
        
        # Generate actions based on classification
        if classification.startswith("Claims"):
            actions.extend(self._generate_claims_actions(ticket_data, classification, sop_details, ticket_age))
        elif classification.startswith("Endorsement"):
            actions.extend(self._generate_endorsement_actions(ticket_data, classification, sop_details, ticket_age))
        elif classification.startswith("Support"):
            actions.extend(self._generate_support_actions(ticket_data, classification, sop_details, ticket_age))
        
        # Check for escalation needs
        escalation_action = self._check_escalation_requirements(classification, sop_details, ticket_age, status)
        if escalation_action:
            actions.append(escalation_action)
        
        # Check for auto-closure eligibility
        closure_action = self._check_auto_closure(ticket_data, classification, ticket_age)
        if closure_action:
            actions.append(closure_action)
        
        return actions
    
    def _calculate_ticket_age(self, created_at: str) -> float:
        """Calculate ticket age in hours"""
        try:
            created_time = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
            current_time = datetime.now(created_time.tzinfo)
            age_hours = (current_time - created_time).total_seconds() / 3600
            return age_hours
        except:
            return 0
    
    def _generate_claims_actions(self, ticket_data: dict, classification: str, sop_details: dict, ticket_age: float) -> list:
        """Generate actions specific to claims tickets"""
        actions = []
        status = ticket_data.get('status', 0)
        
        # Check if initial customer contact is needed
        if status == 1 and ticket_age < 1:  # New ticket, less than 1 hour old
            actions.append({
                'type': 'IMMEDIATE_CONTACT',
                'priority': 'HIGH',
                'action': 'Call customer within 1 hour',
                'reason': 'Claims SOP requires contact within 1 hour',
                'auto_executable': True,
                'execution_method': 'schedule_call',
                'parameters': {
                    'customer_phone': ticket_data.get('requester_phone'),
                    'ticket_id': ticket_data.get('Ticket ID'),
                    'deadline': datetime.now() + timedelta(hours=1)
                }
            })
        
        # Check if manual revert is needed
        if status in [1, 2] and ticket_age < 2:
            actions.append({
                'type': 'MANUAL_REVERT',
                'priority': 'HIGH',
                'action': 'Send manual acknowledgment within 2 hours',
                'reason': 'Claims SOP requires manual revert within 2 hours',
                'auto_executable': True,
                'execution_method': 'send_acknowledgment',
                'template': 'claims_acknowledgment',
                'parameters': {
                    'ticket_id': ticket_data.get('Ticket ID'),
                    'deadline': datetime.now() + timedelta(hours=2)
                }
            })
        
        # Check for document requirements
        if "motor" in classification.lower() and "psu" in ticket_data.get('Subject', '').lower():
            actions.append({
                'type': 'DOCUMENT_REQUEST',
                'priority': 'MEDIUM',
                'action': 'Request PSU-specific documents',
                'reason': 'Motor claims for PSU insurers require specific documents',
                'auto_executable': True,
                'execution_method': 'send_document_request',
                'required_docs': ['Filled claim form', 'PAN card', 'Cancelled cheque'],
                'parameters': {
                    'ticket_id': ticket_data.get('Ticket ID'),
                    'template': 'psu_document_request'
                }
            })
        
        # NC reminder check
        if status == 3 and "BM Dealer" in str(ticket_data.get('custom_status', '')):
            reminder_schedule = self._calculate_nc_reminder_schedule(ticket_data)
            if reminder_schedule:
                actions.append(reminder_schedule)
        
        return actions
    
    def _generate_endorsement_actions(self, ticket_data: dict, classification: str, sop_details: dict, ticket_age: float) -> list:
        """Generate actions specific to endorsement tickets"""
        actions = []
        
        # Extract endorsement type details
        parts = classification.split("-")
        if len(parts) >= 3:
            endorsement_type = parts[1].lower()
            sub_type = parts[2].lower()
            
            # Get TAT for this specific endorsement
            tat_hours = self._get_endorsement_tat(endorsement_type, sub_type, sop_details)
            
            if tat_hours and ticket_age > tat_hours * 0.75:  # 75% of TAT reached
                actions.append({
                    'type': 'TAT_WARNING',
                    'priority': 'HIGH',
                    'action': f'Process endorsement urgently - 75% of {tat_hours} hour TAT reached',
                    'reason': f'{endorsement_type.title()} {sub_type} endorsement TAT is {tat_hours} hours',
                    'auto_executable': True,
                    'execution_method': 'send_tat_alert',
                    'parameters': {
                        'ticket_id': ticket_data.get('Ticket ID'),
                        'tat_hours': tat_hours,
                        'current_age': ticket_age
                    }
                })
        
        # Check for document verification needs
        if "financial" in classification.lower():
            actions.append({
                'type': 'VERIFICATION_REQUIRED',
                'priority': 'MEDIUM',
                'action': 'Verify financial impact and get approval if needed',
                'reason': 'Financial endorsements require verification',
                'auto_executable': False,
                'checklist': ['Verify premium change', 'Check underwriting guidelines', 'Get manager approval if >10% change']
            })
        
        return actions
    
    def _generate_support_actions(self, ticket_data: dict, classification: str, sop_details: dict, ticket_age: float) -> list:
        """Generate actions specific to support tickets"""
        actions = []
        
        if "PDPNR" in classification:
            actions.append({
                'type': 'TECHNICAL_CHECK',
                'priority': 'HIGH',
                'action': 'Check payment status with gateway and insurer API',
                'reason': 'Payment successful but policy not received',
                'auto_executable': True,
                'execution_method': 'check_payment_status',
                'steps': [
                    'Verify payment gateway status',
                    'Check insurer API response',
                    'Initiate manual policy generation if confirmed'
                ],
                'parameters': {
                    'ticket_id': ticket_data.get('Ticket ID'),
                    'transaction_id': self._extract_transaction_id(ticket_data)
                }
            })
        
        elif "PLNG" in classification:
            actions.append({
                'type': 'API_DIAGNOSIS',
                'priority': 'HIGH',
                'action': 'Diagnose payment link generation failure',
                'reason': 'Payment link not generated - API or UW issue',
                'auto_executable': True,
                'execution_method': 'diagnose_plng',
                'checks': [
                    'Check API status',
                    'Verify underwriting criteria',
                    'Review validation errors'
                ]
            })
        
        return actions
    
    def _check_escalation_requirements(self, classification: str, sop_details: dict, ticket_age: float, status: int) -> dict:
        """Check if ticket needs escalation based on SOP matrix"""
        if status in [4, 5]:  # Resolved or Closed
            return None
        
        escalation_matrix = None
        
        # Get appropriate escalation matrix
        if classification.startswith("Claims"):
            claim_type = classification.split("-")[1].lower() if "-" in classification else "general"
            escalation_matrix = SOP_KNOWLEDGE_BASE["claims"].get(claim_type, {}).get("escalation_matrix", {})
        elif classification.startswith("Endorsement"):
            endorsement_type = classification.split("-")[1].lower() if "-" in classification else "general"
            escalation_matrix = SOP_KNOWLEDGE_BASE["endorsement"].get(endorsement_type, {}).get("escalation_matrix", {})
        
        if not escalation_matrix:
            return None
        
        # Determine escalation level based on ticket age
        escalation_action = None
        
        if ticket_age > 6 and "level1" in escalation_matrix:
            level = "level1"
            details = escalation_matrix[level]
        elif ticket_age > 12 and "level2" in escalation_matrix:
            level = "level2"
            details = escalation_matrix[level]
        elif ticket_age > 18 and "level3" in escalation_matrix:
            level = "level3"
            details = escalation_matrix[level]
        else:
            return None
        
        if details:
            escalation_action = {
                'type': 'ESCALATION_REQUIRED',
                'priority': 'URGENT',
                'action': f'Escalate to {details.get("name", "Manager")} - {details.get("designation", "")}',
                'reason': f'Ticket age ({ticket_age:.1f} hours) exceeds escalation threshold',
                'auto_executable': True,
                'execution_method': 'escalate_ticket',
                'parameters': {
                    'escalation_level': level,
                    'escalate_to': details.get("email"),
                    'escalate_to_name': details.get("name"),
                    'wait_time': details.get("wait_time")
                }
            }
        
        return escalation_action
    
    def _check_auto_closure(self, ticket_data: dict, classification: str, ticket_age: float) -> dict:
        """Check if ticket is eligible for auto-closure"""
        status = ticket_data.get('status', 0)
        
        # NC auto-closure check for claims
        if classification.startswith("Claims") and status == 3:
            custom_status = str(ticket_data.get('custom_status', ''))
            if "third_reminder_sent" in custom_status and ticket_age > 168:  # 7 days
                return {
                    'type': 'AUTO_CLOSURE',
                    'priority': 'LOW',
                    'action': 'Auto-close ticket - NC reminders exhausted',
                    'reason': 'Third NC reminder sent 48 hours ago with no response',
                    'auto_executable': True,
                    'execution_method': 'close_ticket',
                    'parameters': {
                        'ticket_id': ticket_data.get('Ticket ID'),
                        'closure_reason': 'No customer response after 3 reminders'
                    }
                }
        
        return None
    
    def _calculate_nc_reminder_schedule(self, ticket_data: dict) -> dict:
        """Calculate NC reminder schedule for claims"""
        # This would check the ticket's reminder history and determine next reminder
        return {
            'type': 'NC_REMINDER',
            'priority': 'MEDIUM',
            'action': 'Send NC reminder',
            'reason': 'Customer not contactable',
            'auto_executable': True,
            'execution_method': 'send_nc_reminder',
            'schedule': {
                'first': '72 hours after NC status',
                'second': '48 hours after first',
                'third_and_close': '48 hours after second'
            }
        }
    
    def _get_endorsement_tat(self, endorsement_type: str, sub_type: str, sop_details: dict) -> int:
        """Get TAT in hours for specific endorsement type"""
        tat_mapping = {
            'motor': {
                'financial': {
                    'ownership_transfer': 5 * 24,
                    'policy_cancellation': 15 * 24,
                    'pa_other_addon_idv': 7 * 24
                },
                'non_financial': {
                    'customer_details': 3 * 24,
                    'hypothecation': 1 * 24,
                    'gstin': 3 * 24
                }
            },
            'health': {
                'financial': {
                    'ped_updation': 15 * 24,
                    'member_addition_newborn': 15 * 24,
                    'dob_change': 8 * 24
                },
                'non_financial': {
                    'gstin': 5 * 24,
                    'name_correction': 5 * 24,
                    'address_updation': 5 * 24
                }
            }
        }
        
        return tat_mapping.get(endorsement_type, {}).get(sub_type, {}).get('default', 72)
    
    def _extract_transaction_id(self, ticket_data: dict) -> str:
        """Extract transaction ID from ticket content"""
        content = ticket_data.get('raw_ticket_content', '')
        # Implement regex or pattern matching to extract transaction ID
        import re
        match = re.search(r'transaction[_\s]?id[:\s]+([A-Z0-9]+)', content, re.IGNORECASE)
        return match.group(1) if match else None
    
    def execute_action(self, action: dict) -> dict:
        """Execute an autonomous action"""
        result = {
            'action_type': action['type'],
            'success': False,
            'message': '',
            'timestamp': datetime.now().isoformat()
        }
        
        if not action.get('auto_executable', False):
            result['message'] = 'Action requires manual execution'
            return result
        
        method_name = action.get('execution_method')
        if hasattr(self, method_name):
            try:
                method = getattr(self, method_name)
                execution_result = method(action.get('parameters', {}))
                result['success'] = execution_result.get('success', False)
                result['message'] = execution_result.get('message', '')
                result['details'] = execution_result.get('details', {})
            except Exception as e:
                result['message'] = f'Execution failed: {str(e)}'
        else:
            result['message'] = f'Execution method {method_name} not implemented'
        
        # Log the action
        self.action_log.append(result)
        return result
    
    # Implementation methods for autonomous actions
    def send_acknowledgment(self, params: dict) -> dict:
        """Send automated acknowledgment email"""
        ticket_id = params.get('ticket_id')
        template = params.get('template', 'default')
        
        # In real implementation, this would send an email via Freshdesk API
        try:
            # Simulate API call
            print(f"Sending acknowledgment for ticket {ticket_id} using template {template}")
            return {
                'success': True,
                'message': f'Acknowledgment sent for ticket {ticket_id}',
                'details': {'template_used': template, 'sent_at': datetime.now().isoformat()}
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}
    
    def send_document_request(self, params: dict) -> dict:
        """Send document request to customer"""
        ticket_id = params.get('ticket_id')
        required_docs = params.get('required_docs', [])
        
        # Implementation would use Freshdesk API to send templated email
        return {
            'success': True,
            'message': f'Document request sent for {len(required_docs)} documents',
            'details': {'documents_requested': required_docs}
        }
    
    def check_payment_status(self, params: dict) -> dict:
        """Check payment status with gateway and insurer"""
        transaction_id = params.get('transaction_id')
        
        # This would integrate with payment gateway and insurer APIs
        return {
            'success': True,
            'message': 'Payment verified with gateway',
            'details': {
                'gateway_status': 'SUCCESS',
                'insurer_status': 'PENDING',
                'recommended_action': 'Initiate manual policy generation'
            }
        }
    
    def escalate_ticket(self, params: dict) -> dict:
        """Escalate ticket to higher level"""
        escalate_to = params.get('escalate_to')
        level = params.get('escalation_level')
        
        # Would update ticket and send escalation email
        return {
            'success': True,
            'message': f'Ticket escalated to {params.get("escalate_to_name")}',
            'details': {'escalation_level': level, 'escalated_at': datetime.now().isoformat()}
        }

# ========== ENHANCED SUMMARY GENERATION WITH ACTIONS ==========

def get_enhanced_claude_summary_with_actions(text, classification, sop_details):
    """Enhanced version that generates both summary and recommended actions"""
    if not anthropic_client:
        return None, []
    
    enhanced_prompt = f"""
    Analyze this {classification} ticket and provide:
    1. Standard RCA summary (Problem, Why, Solution)
    2. Specific action items based on InsuranceDekho SOPs
    
    Ticket content:
    {text}
    
    Relevant SOP:
    {json.dumps(sop_details, indent=2)}
    
    Return JSON with this structure:
    {{
        "summary": {{
            "Problem": "...",
            "Why": "...",
            "Solution": "..."
        }},
        "recommended_actions": [
            {{
                "action": "specific action to take",
                "priority": "HIGH/MEDIUM/LOW",
                "reason": "why this action is needed per SOP",
                "deadline": "when this should be completed"
            }}
        ],
        "automation_suggestions": [
            {{
                "process": "what can be automated",
                "benefit": "expected improvement"
            }}
        ]
    }}
    """
    
    try:
        response = anthropic_client.messages.create(
            model="claude-3-haiku-20240307",
            max_tokens=1500,
            messages=[{"role": "user", "content": enhanced_prompt}],
            system=CLAUDE_SYSTEM_PROMPT,
            temperature=0.2
        )
        
        result = json.loads(response.content[0].text.strip())
        return result.get("summary"), result.get("recommended_actions", [])
    except Exception as e:
        print(f"Error in enhanced summary: {e}")
        return None, []

# ========== SMART RESPONSE GENERATOR ==========

class SmartResponseGenerator:
    """Generate intelligent, context-aware responses based on email trail with zero hallucination."""

    def __init__(self, anthropic_client):
        self.client = anthropic_client
        self.company_signature = """
Girnar Insurance Brokers Private Limited
CIN: U66010RJ2016PTC054811
Registered Office - Girnar 21, Govind Marg, Moti Doongari Road, Dharam Singh Circle, Jaipur, Rajasthan- 302004
Corporate Office - Plot no.301, Phase-2, Udyog Vihar, Gurugram-122022, Haryana, India
IRDAI License no 588 | CIN: U66010RJ2016PTC054811 | Composite Broker Valid Till: 19/03/2026"""

    def _analyze_sentiment(self, ticket_data: dict) -> str:
        """Analyze customer sentiment from ticket content."""
        content = ticket_data.get('raw_ticket_content', '').lower()
        negative_words = ['frustrated', 'angry', 'disappointed', 'unacceptable', 'terrible', 'worst', 'pathetic', 'horrible', 'disgusted', 'furious']
        positive_words = ['thank', 'appreciate', 'grateful', 'pleased', 'satisfied', 'happy', 'excellent', 'wonderful']
        urgent_words = ['urgent', 'asap', 'immediately', 'emergency', 'critical', 'priority', 'escalate', 'now']
        
        # Count occurrences with context
        negative_count = sum(content.count(word) for word in negative_words)
        positive_count = sum(content.count(word) for word in positive_words)
        urgent_count = sum(content.count(word) for word in urgent_words)
        
        # Check for intensity modifiers
        if any(phrase in content for phrase in ['extremely urgent', 'very urgent', 'highly critical']):
            urgent_count += 3
        
        # Weighted sentiment analysis
        if urgent_count >= 2 or 'urgent' in ticket_data.get('tags', []):
            return "urgent"
        elif negative_count > positive_count * 1.5:
            return "negative"
        elif positive_count > negative_count:
            return "positive"
        return "neutral"

    def _extract_last_customer_message(self, ticket_data: dict) -> dict:
        """Extract the last customer message from the email trail with enhanced parsing."""
        raw_content = ticket_data.get('raw_ticket_content', '')
        
        # Enhanced email separators
        email_separators = [
            '\n\nOn ', '\n\n----', '\n\nFrom:', '\n\n_____',
            '\n\n--- Original Message ---', '\n\nSent from',
            '\r\n\r\nOn ', '\r\n\r\n----', '\n\n> ', 
            '\n\nForwarded message', '--- Reply above this line ---'
        ]
        
        # Extract last message
        last_message = raw_content
        for separator in email_separators:
            if separator in raw_content:
                parts = raw_content.split(separator)
                if parts and parts[0].strip():
                    last_message = parts[0].strip()
                    break
        
        # Enhanced extraction of questions and requests
        questions = []
        requests = []
        concerns = []
        
        lines = last_message.split('\n')
        for line in lines:
            line = line.strip()
            
            # Questions - enhanced detection
            if line.endswith('?') or any(q_word in line.lower() for q_word in ['what', 'when', 'where', 'why', 'how', 'can you', 'could you', 'will you']):
                questions.append(line)
            
            # Requests - enhanced detection
            request_indicators = ['please', 'kindly', 'request', 'need', 'want', 'require', 'expecting', 'waiting for', 'looking for', 'help me']
            if any(indicator in line.lower() for indicator in request_indicators):
                requests.append(line)
            
            # Concerns - new category
            concern_indicators = ['concerned', 'worried', 'problem', 'issue', 'trouble', 'difficulty', 'stuck', 'not working', 'failed']
            if any(indicator in line.lower() for indicator in concern_indicators):
                concerns.append(line)
        
        # Extract mentioned specifics (policy numbers, claim numbers, etc.)
        import re
        mentioned_items = {
            'policy_numbers': re.findall(r'\b(?:policy|policy no|policy number)[\s:]*([A-Z0-9/-]+)\b', last_message, re.IGNORECASE),
            'claim_numbers': re.findall(r'\b(?:claim|claim no|claim number)[\s:]*([A-Z0-9/-]+)\b', last_message, re.IGNORECASE),
            'amounts': re.findall(r'₹\s*(\d+(?:,\d+)*(?:\.\d+)?)', last_message),
            'dates': re.findall(r'\b\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\b', last_message)
        }
        
        return {
            'content': last_message,
            'questions': questions,
            'requests': requests,
            'concerns': concerns,
            'mentioned_items': mentioned_items,
            'word_count': len(last_message.split()),
            'has_attachments': 'attached' in last_message.lower() or 'attachment' in last_message.lower()
        }

    def _extract_case_status(self, ticket_data: dict) -> dict:
        """Extract current case status with enhanced accuracy and no assumptions."""
        raw_content = ticket_data.get('raw_ticket_content', '')
        actions_taken = ticket_data.get('Actions Taken', '')
        
        # Initialize with facts only - no assumptions
        status_info = {
            'documents_received': [],
            'documents_pending': [],
            'claim_submitted': False,
            'claim_number': None,
            'insurer_name': None,
            'pending_from': 'unknown',
            'last_action': '',
            'last_action_date': None,
            'next_steps': [],
            'awaiting_response': [],
            'factual_status': [],
            'escalation_level': 0
        }
        
        # Extract factual information only
        content_lower = raw_content.lower()
        
        # Document status - be specific about what was received
        doc_received_patterns = [
            (r'received\s+(\w+\s*\w*)\s*documents?', 'documents'),
            (r'(\w+\s*\w*)\s*documents?\s+received', 'documents'),
            (r'received.*?:\s*([^\.]+)', 'list'),
            (r'documents received.*?:\s*([^\.]+)', 'list')
        ]
        
        for pattern, type in doc_received_patterns:
            matches = re.findall(pattern, content_lower, re.IGNORECASE)
            if matches:
                if type == 'list':
                    status_info['documents_received'].extend([m.strip() for m in matches[0].split(',') if m.strip()])
                else:
                    status_info['documents_received'].extend([m.strip() for m in matches if m.strip()])
        
        # Claim submission status - extract specific details
        claim_patterns = [
            r'claim\s+(?:submitted|raised|forwarded).*?(?:claim\s*(?:no|number|id)[\s:]*)?([A-Z0-9/-]+)',
            r'(?:claim\s*(?:no|number|id)[\s:]*)?([A-Z0-9/-]+).*?submitted\s+to\s+(\w+)',
            r'forwarded\s+to\s+(\w+)\s+(?:insurance|insurer)'
        ]
        
        for pattern in claim_patterns:
            match = re.search(pattern, content_lower, re.IGNORECASE)
            if match:
                status_info['claim_submitted'] = True
                if len(match.groups()) >= 1 and match.group(1):
                    potential_claim = match.group(1).upper()
                    if len(potential_claim) > 3:  # Basic validation
                        status_info['claim_number'] = potential_claim
                if len(match.groups()) >= 2 and match.group(2):
                    status_info['insurer_name'] = match.group(2).title()
                break
        
        # Extract last action with date if available
        if actions_taken:
            actions_list = [a.strip() for a in actions_taken.split('\n') if a.strip()]
            if actions_list:
                last_action = actions_list[-1]
                status_info['last_action'] = last_action
                
                # Try to extract date from last action
                date_match = re.search(r'(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})', last_action)
                if date_match:
                    status_info['last_action_date'] = date_match.group(1)
        
        # Determine pending from - based on facts only
        if status_info['claim_submitted'] and status_info['insurer_name']:
            status_info['pending_from'] = f"{status_info['insurer_name']} Insurance"
        elif 'waiting for documents' in content_lower or 'pending documents' in content_lower:
            status_info['pending_from'] = 'customer'
        elif 'under review' in content_lower or 'processing' in content_lower:
            status_info['pending_from'] = 'internal team'
        
        # Extract specific waiting items
        waiting_patterns = [
            r'waiting\s+for\s+([^\.]+)',
            r'awaiting\s+([^\.]+)',
            r'pending\s+([^\.]+)',
            r'require\s+([^\.]+)'
        ]
        
        for pattern in waiting_patterns:
            matches = re.findall(pattern, content_lower, re.IGNORECASE)
            for match in matches:
                if len(match) < 100:  # Reasonable length
                    status_info['awaiting_response'].append(match.strip())
        
        # Build factual status statements
        if status_info['documents_received']:
            status_info['factual_status'].append(f"Received: {', '.join(status_info['documents_received'][:3])}")
        
        if status_info['claim_submitted']:
            claim_status = "Claim submitted"
            if status_info['claim_number']:
                claim_status += f" (#{status_info['claim_number']})"
            if status_info['insurer_name']:
                claim_status += f" to {status_info['insurer_name']}"
            status_info['factual_status'].append(claim_status)
        
        # Determine next steps based on current status - no assumptions
        if not status_info['documents_received']:
            status_info['next_steps'].append("Document submission pending")
        elif status_info['claim_submitted']:
            status_info['next_steps'].append("Awaiting insurer response")
        else:
            status_info['next_steps'].append("Processing in progress")
        
        # Check escalation level
        if 'escalat' in content_lower or 'manager' in content_lower or 'supervisor' in content_lower:
            status_info['escalation_level'] = 1
        if 'legal' in content_lower or 'ombudsman' in content_lower:
            status_info['escalation_level'] = 2
        
        return status_info

    def _validate_response_facts(self, response: str, ticket_data: dict) -> str:
        """Validate response against known facts to prevent hallucination."""
        
        # Extract all numbers, dates, and specific claims from response
        import re
        
        # Things to validate
        validations = {
            'claim_numbers': re.findall(r'claim\s*#?\s*([A-Z0-9/-]+)', response, re.IGNORECASE),
            'policy_numbers': re.findall(r'policy\s*#?\s*([A-Z0-9/-]+)', response, re.IGNORECASE),
            'amounts': re.findall(r'₹\s*(\d+(?:,\d+)*(?:\.\d+)?)', response),
            'dates': re.findall(r'\b\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\b', response),
            'insurer_names': re.findall(r'(?:submitted to|forwarded to)\s+(\w+)\s+(?:insurance|insurer)', response, re.IGNORECASE)
        }
        
        # Get facts from ticket data
        ticket_content = ticket_data.get('raw_ticket_content', '')
        custom_fields = ticket_data.get('custom_fields', {})
        
        # Validate each type
        corrections_needed = []
        
        # Validate claim numbers
        for claim_num in validations['claim_numbers']:
            if claim_num not in ticket_content and claim_num not in str(custom_fields):
                corrections_needed.append(('claim_number', claim_num))
        
        # Remove any hallucinated information
        if corrections_needed:
            for item_type, item_value in corrections_needed:
                # Remove specific mentions that aren't in the source data
                response = response.replace(f"#{item_value}", "")
                response = response.replace(item_value, "your claim")
            
            # Clean up any double spaces
            response = re.sub(r'\s+', ' ', response)
        
        return response.strip()

    def generate_response(self, ticket_data: dict, context: str, response_type: str = 'general', agent_name: str = 'Support Team') -> str:
        """Generate intelligent response with zero hallucination and maximum accuracy."""
        
        classification = ticket_data.get('Classification', '')
        sentiment = self._analyze_sentiment(ticket_data)
        case_status = self._extract_case_status(ticket_data)
        last_message = self._extract_last_customer_message(ticket_data)
        
        # Convert datetime objects to strings
        safe_ticket_data = _sanitize_for_json(ticket_data)
        
        # Build fact sheet for the AI
        fact_sheet = f"""
VERIFIED FACTS ONLY - DO NOT ASSUME OR INVENT INFORMATION:
- Customer Sentiment: {sentiment}
- Last Message Word Count: {last_message['word_count']}
- Specific Questions Asked: {len(last_message['questions'])}
- Documents Actually Received: {', '.join(case_status['documents_received']) if case_status['documents_received'] else 'None confirmed'}
- Claim Submission Status: {'Confirmed' if case_status['claim_submitted'] else 'Not confirmed'}
- Claim Number: {case_status['claim_number'] if case_status['claim_number'] else 'Not found in records'}
- Insurer Name: {case_status['insurer_name'] if case_status['insurer_name'] else 'Not specified'}
- Last Recorded Action: {case_status['last_action'] if case_status['last_action'] else 'No action recorded'}
- Currently Pending From: {case_status['pending_from']}
- Escalation Level: {case_status['escalation_level']}

MENTIONED IN LAST MESSAGE:
- Policy Numbers: {', '.join(last_message['mentioned_items']['policy_numbers']) if last_message['mentioned_items']['policy_numbers'] else 'None'}
- Claim Numbers: {', '.join(last_message['mentioned_items']['claim_numbers']) if last_message['mentioned_items']['claim_numbers'] else 'None'}
- Amounts: {', '.join(last_message['mentioned_items']['amounts']) if last_message['mentioned_items']['amounts'] else 'None'}
- Has Attachments: {'Yes' if last_message['has_attachments'] else 'No'}
"""
        
        prompt = f"""
You are a professional insurance service agent. Generate a response following these STRICT RULES:

CRITICAL RULES FOR ACCURACY:
1. ONLY state facts that are explicitly mentioned in the VERIFIED FACTS section
2. NEVER invent claim numbers, policy numbers, dates, or amounts
3. NEVER assume document names if not explicitly listed
4. NEVER mention specific timelines or TAT
5. If a fact is marked as "None" or "Not confirmed", do not make up information
6. Address the customer's actual questions from their last message
7. Be empathetic but factual - no false promises
8. Do NOT include any signature or company details

{fact_sheet}

CUSTOMER'S LAST MESSAGE:
{last_message['content'][:500]}

THEIR SPECIFIC QUESTIONS:
{chr(10).join(['- ' + q for q in last_message['questions'][:5]])}

THEIR REQUESTS:
{chr(10).join(['- ' + r for r in last_message['requests'][:3]])}

THEIR CONCERNS:
{chr(10).join(['- ' + c for c in last_message['concerns'][:3]])}

CLASSIFICATION: {classification}
CONTEXT: {context}

Generate a response that:
1. Directly addresses each question with facts only
2. Acknowledges their concerns without making assumptions
3. States only confirmed next steps
4. Uses "your claim" instead of inventing claim numbers
5. Says "our team" instead of inventing department names

Keep under 150 words. Be helpful but stick to facts.
"""

        try:
            response = self.client.messages.create(
                model="claude-3-haiku-20240307",
                max_tokens=300,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.2  # Lower temperature for more consistent, factual responses
            )
            
            generated_response = response.content[0].text.strip()
            
            # Validate response for accuracy
            generated_response = self._validate_response_facts(generated_response, ticket_data)
            
            # Post-process to ensure no TAT mentions
            tat_patterns = [
                r'\d+\s*(hours?|days?|business\s*hours?|business\s*days?|working\s*days?)',
                r'within\s*\d+',
                r'in\s*\d+\s*(hours?|days?)',
                r'TAT',
                r'turnaround time',
                r'by\s+(monday|tuesday|wednesday|thursday|friday|saturday|sunday)',
                r'by\s+(tomorrow|today|end of)',
            ]
            
            import re
            for pattern in tat_patterns:
                generated_response = re.sub(pattern, 'soon', generated_response, flags=re.IGNORECASE)
            
            # Additional hallucination prevention
            hallucination_phrases = [
                r'claim\s*#\s*[A-Z0-9]{10,}',  # Remove made-up long claim numbers
                r'reference\s*#\s*[A-Z0-9]{10,}',  # Remove made-up reference numbers
                r'₹\s*\d{7,}',  # Remove unrealistic amounts (7+ digits)
                r'\b\d{2,3}%\s*(?:processed|completed|approved)',  # Remove made-up percentages
            ]
            
            for pattern in hallucination_phrases:
                generated_response = re.sub(pattern, '', generated_response, flags=re.IGNORECASE)
            
            # Clean up any artifacts from removal
            generated_response = re.sub(r'\s+', ' ', generated_response).strip()
            
            # Add agent signature and company details
            final_response = f"{generated_response}\n\nBest regards,\n{agent_name}\n{self.company_signature}"
            
            return final_response
            
        except Exception as e:
            # Enhanced fallback response based on verified facts
            opening = ""
            
            # Sentiment-based opening
            if sentiment == "urgent":
                opening = f"I understand the urgency of your {classification.lower()}. "
            elif sentiment == "negative":
                opening = f"I understand your concerns regarding your {classification.lower()}. "
            elif last_message['questions']:
                opening = "Thank you for your questions. "
            else:
                opening = "Thank you for your message. "
            
            # Status-based body
            body = ""
            if case_status['claim_submitted'] and case_status['claim_number']:
                body = f"Your claim (#{case_status['claim_number']}) has been submitted to {case_status['insurer_name'] or 'the insurer'}. "
            elif case_status['claim_submitted']:
                body = "Your claim has been submitted to the insurer. "
            elif case_status['documents_received']:
                body = f"We have received your {', '.join(case_status['documents_received'][:2])}. "
            else:
                body = "We are processing your request. "
            
            # Add next steps
            if case_status['awaiting_response']:
                body += f"We are currently {case_status['awaiting_response'][0]}. "
            
            closing = "Our team is actively working on this and will update you as soon as we have more information."
            
            # Combine all parts
            base_response = f"{opening}{body}{closing}"
            
            # Add agent signature and company details
            return f"{base_response}\n\nBest regards,\n{agent_name}\n{self.company_signature}"

# Enhanced Smart Response Generator with Complete Context
class EnhancedSmartResponseGenerator(SmartResponseGenerator):
    """Extended response generator that uses complete ticket context"""

    def __init__(self, anthropic_client, routing_analyzer):
        super().__init__(anthropic_client)
        self.routing_analyzer = routing_analyzer

    def generate_contextual_response(self, ticket_id, response_type='update', agent_name='Support Team'):
        """
        Generate response based on complete ticket analysis including parent-child relationships
        """
        # Get complete ticket context
        complete_context = self.routing_analyzer.analyze_complete_ticket_context(ticket_id)

        if not complete_context:
            return "Unable to fetch ticket details for response generation."

        # Fetch ticket data for base analysis
        ticket_data = fetch_ticket_by_id(ticket_id)
        conversations = fetch_all_ticket_conversations(ticket_id)
        ticket_content, _ = extract_email_content_and_attachments(ticket_data, conversations)

        # Build enhanced context for response generation
        enhanced_context = f"""
COMPLETE TICKET ANALYSIS:

Current Status: {complete_context['current_status']['display']}
Status Category: {complete_context['current_status']['category']}
Primary Routing: {complete_context['routing_intent']}
Current Action: {complete_context['current_action']['primary_action']}

ROUTING HISTORY:
{self._format_routing_history(complete_context['routing_history'])}

PENDING ITEMS:
{self._format_pending_items(complete_context['pending_items'])}

KEY ENTITIES:
- Claim Numbers: {', '.join(complete_context['key_entities']['claim_numbers']) or 'None'}
- Policy Numbers: {', '.join(complete_context['key_entities']['policy_numbers']) or 'None'}
- Insurers Mentioned: {', '.join(complete_context['key_entities']['insurer_names']) or 'None'}

PRIORITY INDICATORS:
{self._format_priority_indicators(complete_context['priority_indicators'])}

PARENT-CHILD ANALYSIS:
{self._format_parent_child_analysis(complete_context['parent_child_analysis'])}

Ticket Age: {complete_context['ticket_age_hours']:.1f} hours
Last Update: {complete_context['last_update_hours']:.1f} hours ago

CURRENT BLOCKERS:
{self._format_blockers(complete_context['current_action'].get('blockers', []))}
"""

        # Determine response strategy based on context
        response_strategy = self._determine_response_strategy(complete_context)

        # Generate response using enhanced prompt
        prompt = f"""
You are a professional insurance service agent. Generate a response based on this complete ticket analysis.

{enhanced_context}

RESPONSE TYPE: {response_type}
RESPONSE STRATEGY: {response_strategy}

INSTRUCTIONS:
1. Address the current routing destination ({complete_context['routing_intent']})
2. Acknowledge the current action status
3. Reference specific entities (claim/policy numbers) if available
4. Address any pending items clearly
5. If parent-child tickets exist, reference the relationship
6. Be specific about next steps based on routing

Generate a professional response that:
- Directly addresses the current situation
- Provides clear next steps
- Uses appropriate tone for the routing destination
- Includes relevant reference numbers
- Maintains accuracy without assumptions

Keep the response concise and action-oriented.
"""

        try:
            response = self.client.messages.create(
                model="claude-3-haiku-20240307",
                max_tokens=400,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.3
            )

            generated_response = response.content[0].text.strip()

            # Add appropriate signature based on routing
            signature = self._get_contextual_signature(complete_context['routing_intent'], agent_name)

            return f"{generated_response}\n\n{signature}\n{self.company_signature}"

        except Exception as e:
            # Fallback response with context
            return self._generate_fallback_response(complete_context, agent_name)

    def _format_routing_history(self, routing_history):
        """Format routing history for display"""
        if not routing_history:
            return "No routing history available"

        formatted = []
        for event in routing_history[-3:]:  # Show last 3 routing events
            formatted.append(f"- {event['routed_to']}: {event['reason']} ({event.get('timestamp', 'Unknown time')})")

        return '\n'.join(formatted)

    def _format_pending_items(self, pending_items):
        """Format pending items for display"""
        if not pending_items:
            return "No specific items pending"

        unique_items = []
        seen = set()
        for item in pending_items[-5:]:  # Show last 5 items
            if item['item'] not in seen:
                unique_items.append(f"- {item['item']}")
                seen.add(item['item'])

        return '\n'.join(unique_items)

    def _format_priority_indicators(self, indicators):
        """Format priority indicators"""
        if not indicators:
            return "Normal priority"

        return '\n'.join([f"- {ind['description']}" for ind in indicators[:3]])

    def _format_parent_child_analysis(self, analysis):
        """Format parent-child relationship analysis"""
        if analysis['is_parent']:
            return f"Parent ticket with {len(analysis['child_tickets'])} child tickets"
        elif analysis['is_child']:
            return f"Child ticket of #{analysis['parent_ticket']}"
        return "Standalone ticket"

    def _format_blockers(self, blockers):
        """Format current blockers"""
        if not blockers:
            return "No blockers identified"

        return '\n'.join([f"- {blocker['description']}" for blocker in blockers])

    def _determine_response_strategy(self, context):
        """Determine the best response strategy based on context"""
        status_category = context['current_status']['category']
        routing = context['routing_intent']

        # Critical tickets need immediate action
        if status_category == 'critical':
            return 'immediate_action_required'
        # Waiting tickets need follow-up
        if status_category == 'waiting':
            if routing == 'customer':
                return 'gentle_reminder_to_customer'
            elif routing == 'insurer':
                return 'professional_follow_up_to_insurer'
            elif routing == 'dealer':
                return 'partner_coordination'
            else:
                return 'status_update'

        # Active tickets need progress update
        if status_category == 'active':
            return 'progress_update'

        # Complete tickets need closure confirmation
        if status_category == 'complete':
            return 'closure_confirmation'

        return 'general_update'

    def _get_contextual_signature(self, routing_intent, agent_name):
        """Get appropriate signature based on routing context"""
        signatures = {
            'insurer': f"Best regards,\n{agent_name}\nInsurer Relations Team",
            'customer': f"Best regards,\n{agent_name}\nCustomer Success Team",
            'dealer': f"Best regards,\n{agent_name}\nPartner Support Team",
            'inspection': f"Best regards,\n{agent_name}\nClaims Processing Team",
            'tech_support': f"Best regards,\n{agent_name}\nTechnical Support Team",
            'internal': f"Best regards,\n{agent_name}\nOperations Team",
            'critical_escalation': f"Best regards,\n{agent_name}\nEscalation Team"
        }

        return signatures.get(routing_intent, f"Best regards,\n{agent_name}")

    def _generate_fallback_response(self, context, agent_name):
        """Generate fallback response when AI fails"""
        routing = context['routing_intent']
        status = context['current_status']['display']
        action = context['current_action']['primary_action']

        responses = {
            'customer': f"Thank you for your patience. Your request is currently {status}. {action}. We will update you as soon as we have more information.",
            'insurer': f"This ticket (#{{context['ticket_id']}}) is currently {status}. {action}. Please provide an update at your earliest convenience.",
            'dealer': f"Partner attention required for ticket #{{context['ticket_id']}}. Current status: {status}. {action}.",
            'internal': f"Internal update: Ticket #{{context['ticket_id']}} is {status}. {action}."
        }

        base_response = responses.get(routing, f"Ticket #{{context['ticket_id']}} status: {status}. {action}.")
        signature = self._get_contextual_signature(routing, agent_name)

        return f"{base_response}\n\n{signature}\n{self.company_signature}"

# Integration function to add routing analysis to existing ticket processing
def enhance_ticket_with_routing_context(ticket_data):
   """
   Enhance ticket data with complete routing and context analysis
   """
   # Initialize analyzer
   routing_analyzer = EnhancedContextualRoutingAnalyzer(FRESHDESK_DOMAIN, FRESHDESK_API_KEY)
   
   # Get ticket ID
   ticket_id = ticket_data.get('Ticket ID') or ticket_data.get('id')
   
   if not ticket_id:
       return ticket_data
   
   # Analyze complete context
   complete_context = routing_analyzer.analyze_complete_ticket_context(ticket_id)
   
   if complete_context:
       # Add routing analysis to ticket data
       ticket_data['routing_analysis'] = {
           'primary_intent': complete_context['routing_intent'],
           'current_routing': complete_context['current_status']['routing'],
           'routing_history': complete_context['routing_history'],
           'routing_summary': _generate_routing_summary(complete_context)
       }
       
       # Add action analysis
       ticket_data['action_analysis'] = {
           'current_action': complete_context['current_action']['primary_action'],
           'action_details': complete_context['current_action']['details'],
           'blockers': complete_context['current_action']['blockers'],
           'next_steps': _determine_next_steps(complete_context)
       }
       
       # Add enhanced status information
       ticket_data['enhanced_status'] = {
           'display_name': complete_context['current_status']['display'],
           'category': complete_context['current_status']['category'],
           'routing_implication': complete_context['current_status']['routing'],
           'action_context': complete_context['current_status']['action']
       }
       
       # Add relationship analysis
       ticket_data['relationship_analysis'] = complete_context['parent_child_analysis']
       
       # Add priority assessment
       ticket_data['priority_assessment'] = {
           'indicators': complete_context['priority_indicators'],
           'urgency_level': _calculate_urgency_level(complete_context),
           'recommended_priority': _recommend_priority(complete_context)
       }
       
       # Add key entities for quick reference
       ticket_data['key_entities'] = complete_context['key_entities']
       
       # Add pending items summary
       ticket_data['pending_summary'] = {
           'items': complete_context['pending_items'],
           'count': len(complete_context['pending_items']),
           'primary_pending': complete_context['pending_items'][0] if complete_context['pending_items'] else None
       }
       
       # Generate intelligent response suggestion
       if smart_response_generator:
           enhanced_generator = EnhancedSmartResponseGenerator(anthropic_client, routing_analyzer)
           ticket_data['suggested_response'] = enhanced_generator.generate_contextual_response(
               ticket_id,
               response_type='update'
           )
       
       # Add actionable insights
       ticket_data['actionable_insights'] = _generate_actionable_insights(complete_context, ticket_data)
   
   return ticket_data

def _generate_routing_summary(context):
   """Generate human-readable routing summary"""
   current_routing = context['current_status']['routing'] or 'internal'
   routing_history = context['routing_history']
   
   if not routing_history:
       return f"Currently with {current_routing} team for initial processing"
   
   # Get unique routing destinations
   destinations = []
   for event in routing_history:
       if event['routed_to'] not in [d['destination'] for d in destinations]:
           destinations.append({
               'destination': event['routed_to'],
               'timestamp': event['timestamp'],
               'reason': event['reason']
           })
   
   # Build summary
   if len(destinations) == 1:
       return f"Routed to {destinations[0]['destination']} for {destinations[0]['reason']}"
   else:
       journey = " → ".join([d['destination'] for d in destinations[-3:]])
       return f"Routing journey: {journey}"

def _determine_next_steps(context):
   """Determine next steps based on current context"""
   next_steps = []
   
   routing = context['routing_intent']
   status_category = context['current_status']['category']
   blockers = context['current_action'].get('blockers', [])
   
   # Based on routing destination
   if routing == 'customer' and status_category == 'waiting':
       next_steps.append({
           'action': 'Follow up with customer',
           'priority': 'high',
           'deadline': '24 hours'
       })
   elif routing == 'insurer' and status_category == 'waiting':
       next_steps.append({
           'action': 'Check insurer portal for updates',
           'priority': 'medium',
           'deadline': '4 hours'
       })
       next_steps.append({
           'action': 'Send reminder to insurer if no response',
           'priority': 'medium',
           'deadline': '48 hours'
       })
   elif routing == 'dealer':
       next_steps.append({
           'action': 'Coordinate with channel partner',
           'priority': 'medium',
           'deadline': '24 hours'
       })
   
   # Based on blockers
   for blocker in blockers:
       if 'document' in blocker.get('description', '').lower():
           next_steps.append({
               'action': 'Verify and collect pending documents',
               'priority': 'high',
               'deadline': 'immediate'
           })
       elif 'kyc' in blocker.get('description', '').lower():
           next_steps.append({
               'action': 'Complete KYC verification',
               'priority': 'high',
               'deadline': '24 hours'
           })
   
   # Based on priority indicators
   if context.get('priority_indicators'):
       next_steps.append({
           'action': 'Escalate to supervisor for priority handling',
           'priority': 'urgent',
           'deadline': 'immediate'
       })
   
   return next_steps

def _calculate_urgency_level(context):
   """Calculate urgency level based on multiple factors"""
   score = 0
   
   # Age factor
   age_hours = context['ticket_age_hours']
   if age_hours > 72:
       score += 3
   elif age_hours > 48:
       score += 2
   elif age_hours > 24:
       score += 1
   
   # Status category factor
   category = context['current_status']['category']
   if category == 'critical':
       score += 5
   elif category == 'waiting' and age_hours > 24:
       score += 2
   
   # Priority indicators
   score += len(context['priority_indicators']) * 2
   
   # Determine level
   if score >= 8:
       return 'critical'
   elif score >= 5:
       return 'high'
   elif score >= 3:
       return 'medium'
   else:
       return 'normal'

def _recommend_priority(context):
   """Recommend priority handling based on context"""
   urgency = _calculate_urgency_level(context)
   routing = context['routing_intent']
   
   recommendations = {
       'critical': {
           'priority': 'P1',
           'sla': '2 hours',
           'action': 'Immediate escalation required'
       },
       'high': {
           'priority': 'P2',
           'sla': '4 hours',
           'action': 'Priority handling needed'
       },
       'medium': {
           'priority': 'P3',
           'sla': '24 hours',
           'action': 'Standard priority with monitoring'
       },
       'normal': {
           'priority': 'P4',
           'sla': '48 hours',
           'action': 'Regular processing'
       }
   }
   
   base_recommendation = recommendations[urgency]
   
   # Adjust based on routing
   if routing == 'critical_escalation':
       base_recommendation['priority'] = 'P1'
       base_recommendation['sla'] = '1 hour'
   elif routing == 'customer' and urgency in ['high', 'critical']:
       base_recommendation['action'] += ' with customer callback'
   
   return base_recommendation

def _generate_actionable_insights(context, ticket_data):
   """Generate actionable insights for the support team"""
   insights = []
   
   # Routing insights
   routing = context['routing_intent']
   if routing == 'insurer' and context['ticket_age_hours'] > 48:
       insights.append({
           'type': 'escalation',
           'insight': 'Ticket pending with insurer for >48 hours',
           'action': 'Escalate to insurer relationship manager',
           'impact': 'high'
       })
   
   # Pattern insights
   if len(context['routing_history']) > 3:
       insights.append({
           'type': 'complexity',
           'insight': 'Multiple routing changes indicate complex issue',
           'action': 'Assign to senior agent for resolution',
           'impact': 'medium'
       })
   
   # Entity insights
   entities = context['key_entities']
   if entities['claim_numbers'] and not entities['insurer_names']:
       insights.append({
           'type': 'missing_info',
           'insight': 'Claim number present but insurer not identified',
           'action': 'Verify insurer details from policy number',
           'impact': 'medium'
       })
   
   # Status insights
   if context['current_status']['category'] == 'waiting' and context['last_update_hours'] > 24:
       insights.append({
           'type': 'stale',
           'insight': 'No updates in last 24 hours while waiting for response',
           'action': 'Send follow-up communication',
           'impact': 'high'
       })
   
   # Parent-child insights
   if context['parent_child_analysis']['is_parent'] and not context['parent_child_analysis']['child_tickets']:
       insights.append({
           'type': 'process',
           'insight': 'Parent ticket without child tickets',
           'action': 'Create necessary child tickets for proper tracking',
           'impact': 'medium'
       })
   
   return insights
def _generate_executive_summary(ticket_data):
   """Generate executive summary of ticket status"""
   routing = ticket_data.get('routing_analysis', {})
   action = ticket_data.get('action_analysis', {})
   status = ticket_data.get('enhanced_status', {})
   priority = ticket_data.get('priority_assessment', {})
   
   summary = {
       'one_line_status': f"{status.get('display_name', 'Unknown')} - {action.get('current_action', 'Processing')}",
       'routing_status': f"Currently with {routing.get('primary_intent', 'internal')} team",
       'urgency': priority.get('urgency_level', 'normal'),
       'blockers': len(action.get('blockers', [])),
       'days_open': round(ticket_data.get('ticket_age_hours', 0) / 24, 1),
       'last_activity': round(ticket_data.get('last_update_hours', 0), 1),
       'key_numbers': {
           'claims': ticket_data.get('key_entities', {}).get('claim_numbers', []),
           'policies': ticket_data.get('key_entities', {}).get('policy_numbers', [])
       }
   }
   
   return summary

# ========== WORKFLOW AUTOMATION ENGINE ==========

class WorkflowAutomationEngine:
    """Handle complex multi-step workflows"""
    
    def __init__(self, freshdesk_domain, freshdesk_api_key):
        self.freshdesk_domain = freshdesk_domain
        self.freshdesk_api_key = freshdesk_api_key
        self.active_workflows = {}
    
    def create_workflow(self, ticket_id: str, classification: str, sop_details: dict) -> dict:
        """Create automated workflow based on ticket type"""
        workflow_id = f"WF_{ticket_id}_{datetime.now().timestamp()}"
        
        if classification.startswith("Claims"):
            workflow = self._create_claims_workflow(ticket_id, classification, sop_details)
        elif classification.startswith("Endorsement"):
            workflow = self._create_endorsement_workflow(ticket_id, classification, sop_details)
        else:
            workflow = self._create_generic_workflow(ticket_id, classification)
        
        workflow['id'] = workflow_id
        workflow['created_at'] = datetime.now().isoformat()
        workflow['status'] = 'ACTIVE'
        
        self.active_workflows[workflow_id] = workflow
        return workflow
    
    def _create_claims_workflow(self, ticket_id: str, classification: str, sop_details: dict) -> dict:
        """Create claims-specific workflow"""
        return {
            'type': 'CLAIMS_WORKFLOW',
            'ticket_id': ticket_id,
            'steps': [
                {
                    'id': 'step_1',
                    'name': 'Initial Contact',
                    'action': 'Contact customer within 1 hour',
                    'deadline': datetime.now() + timedelta(hours=1),
                    'status': 'PENDING',
                    'automated': True,
                    'dependencies': []
                },
                {
                    'id': 'step_2',
                    'name': 'Document Collection',
                    'action': 'Collect required claim documents',
                    'deadline': datetime.now() + timedelta(days=2),
                    'status': 'PENDING',
                    'automated': False,
                    'dependencies': ['step_1']
                },
                {
                    'id': 'step_3',
                    'name': 'Claim Submission',
                    'action': 'Submit claim to insurer',
                    'deadline': datetime.now() + timedelta(days=3),
                    'status': 'PENDING',
                    'automated': True,
                    'dependencies': ['step_2']
                },
                {
                    'id': 'step_4',
                    'name': 'Follow-up',
                    'action': 'Follow up with insurer for updates',
                    'deadline': datetime.now() + timedelta(days=7),
                    'status': 'PENDING',
                    'automated': True,
                    'dependencies': ['step_3']
                }
            ],
            'escalation_triggers': [
                {'condition': 'step_overdue', 'hours': 2, 'action': 'escalate_to_manager'},
                {'condition': 'customer_complaint', 'action': 'immediate_escalation'}
            ]
        }
    
    def _create_endorsement_workflow(self, ticket_id: str, classification: str, sop_details: dict) -> dict:
        """Create endorsement-specific workflow"""
        is_financial = 'financial' in classification.lower()
        
        return {
            'type': 'ENDORSEMENT_WORKFLOW',
            'ticket_id': ticket_id,
            'is_financial': is_financial,
            'steps': [
                {
                    'id': 'step_1',
                    'name': 'Document Verification',
                    'action': 'Verify submitted documents',
                    'deadline': datetime.now() + timedelta(hours=4),
                    'status': 'PENDING',
                    'automated': False,
                    'dependencies': []
                },
                {
                    'id': 'step_2',
                    'name': 'Impact Assessment',
                    'action': 'Assess financial impact' if is_financial else 'Verify details',
                    'deadline': datetime.now() + timedelta(days=1),
                    'status': 'PENDING',
                    'automated': is_financial,
                    'dependencies': ['step_1']
                },
                {
                    'id': 'step_3',
                    'name': 'Approval',
                    'action': 'Get approval if needed' if is_financial else 'Process update',
                    'deadline': datetime.now() + timedelta(days=2),
                    'status': 'PENDING',
                    'automated': False,
                    'dependencies': ['step_2']
                },
                {
                    'id': 'step_4',
                    'name': 'Implementation',
                    'action': 'Implement endorsement in system',
                    'deadline': datetime.now() + timedelta(days=3),
                    'status': 'PENDING',
                    'automated': True,
                    'dependencies': ['step_3']
                }
            ]
        }
    
    def execute_workflow_step(self, workflow_id: str, step_id: str) -> dict:
        """Execute a specific workflow step"""
        workflow = self.active_workflows.get(workflow_id)
        if not workflow:
            return {'success': False, 'message': 'Workflow not found'}
        
        step = next((s for s in workflow['steps'] if s['id'] == step_id), None)
        if not step:
            return {'success': False, 'message': 'Step not found'}
        
        # Check dependencies
        for dep_id in step['dependencies']:
            dep_step = next((s for s in workflow['steps'] if s['id'] == dep_id), None)
            if dep_step and dep_step['status'] != 'COMPLETED':
                return {'success': False, 'message': f'Dependency {dep_id} not completed'}
        
        # Execute step (in real implementation, this would perform actual actions)
        if step['automated']:
            # Simulate automated execution
            step['status'] = 'COMPLETED'
            step['completed_at'] = datetime.now().isoformat()
            return {'success': True, 'message': f'Step {step_id} completed automatically'}
        else:
            # Mark as ready for manual execution
            step['status'] = 'IN_PROGRESS'
            return {'success': True, 'message': f'Step {step_id} ready for manual execution'}
    
    def get_workflow_status(self, workflow_id: str) -> dict:
        """Get current workflow status"""
        workflow = self.active_workflows.get(workflow_id)
        if not workflow:
            return None
        
        total_steps = len(workflow['steps'])
        completed_steps = sum(1 for s in workflow['steps'] if s['status'] == 'COMPLETED')
        
        return {
            'workflow_id': workflow_id,
            'type': workflow['type'],
            'progress': (completed_steps / total_steps) * 100 if total_steps > 0 else 0,
            'status': workflow['status'],
            'steps': workflow['steps'],
            'next_action': self._get_next_action(workflow)
        }
    
    def _get_next_action(self, workflow: dict) -> dict:
        """Get the next action required in workflow"""
        for step in workflow['steps']:
            if step['status'] == 'PENDING':
                # Check if dependencies are met
                deps_met = all(
                    any(s['id'] == dep and s['status'] == 'COMPLETED' 
                        for s in workflow['steps']) 
                    for dep in step['dependencies']
                )
                if deps_met:
                    return {
                        'step_id': step['id'],
                        'action': step['action'],
                        'deadline': step['deadline'],
                        'automated': step['automated']
                    }
        return None
    
    def _create_generic_workflow(self, ticket_id: str, classification: str) -> dict:
        """Create generic workflow for other ticket types"""
        return {
            'type': 'GENERIC_WORKFLOW',
            'ticket_id': ticket_id,
            'steps': [
                {
                    'id': 'step_1',
                    'name': 'Initial Assessment',
                    'action': 'Review and categorize request',
                    'deadline': datetime.now() + timedelta(hours=2),
                    'status': 'PENDING',
                    'automated': False,
                    'dependencies': []
                },
                {
                    'id': 'step_2',
                    'name': 'Process Request',
                    'action': 'Take appropriate action based on request type',
                    'deadline': datetime.now() + timedelta(days=1),
                    'status': 'PENDING',
                    'automated': False,
                    'dependencies': ['step_1']
                },
                {
                    'id': 'step_3',
                    'name': 'Resolution',
                    'action': 'Complete request and notify customer',
                    'deadline': datetime.now() + timedelta(days=2),
                    'status': 'PENDING',
                    'automated': True,
                    'dependencies': ['step_2']
                }
            ]
        }

# ========== PREDICTIVE ANALYTICS ENGINE ==========

class PredictiveAnalyticsEngine:
    """Predict ticket outcomes and suggest proactive actions"""
    
    def __init__(self):
        self.patterns = self._load_historical_patterns()
    
    def _load_historical_patterns(self):
        """Load patterns from historical data"""
        return {
            'escalation_predictors': {
                'keywords': ['urgent', 'frustrated', 'disappointed', 'legal', 'complaint'],
                'age_threshold': 24,  # hours
                'interaction_count': 3
            },
            'resolution_time_factors': {
                'claims': {'motor': 72, 'health': 96, 'life': 120},
                'endorsement': {'financial': 48, 'non_financial': 24},
                'support': {'pdpnr': 12, 'plng': 6, 'general': 24}
            },
            'customer_satisfaction_indicators': {
                'positive': ['thank you', 'appreciated', 'helpful', 'resolved'],
                'negative': ['poor service', 'disappointed', 'unacceptable', 'delay']
            }
        }
    
    def predict_ticket_outcome(self, ticket_data: dict, classification: str) -> dict:
        """Predict likely outcome and timeline for ticket"""
        predictions = {
            'escalation_risk': self._calculate_escalation_risk(ticket_data),
            'estimated_resolution_time': self._estimate_resolution_time(classification, ticket_data),
            'customer_satisfaction_risk': self._assess_satisfaction_risk(ticket_data),
            'automation_potential': self._assess_automation_potential(classification),
            'recommendations': []
        }
        
        # Generate recommendations based on predictions
        if predictions['escalation_risk'] > 0.7:
            predictions['recommendations'].append({
                'action': 'Proactive escalation',
                'reason': 'High escalation risk detected',
                'priority': 'HIGH'
            })
        
        if predictions['customer_satisfaction_risk'] > 0.6:
            predictions['recommendations'].append({
                'action': 'Priority handling with senior agent',
                'reason': 'Customer dissatisfaction risk',
                'priority': 'HIGH'
            })
        
        return predictions
    
    def _calculate_escalation_risk(self, ticket_data: dict) -> float:
        """Calculate probability of escalation"""
        risk_score = 0.0
        content = ticket_data.get('raw_ticket_content', '').lower()
        
        # Check for escalation keywords
        escalation_keywords = self.patterns['escalation_predictors']['keywords']
        keyword_matches = sum(1 for k in escalation_keywords if k in content)
        risk_score += min(keyword_matches * 0.2, 0.6)
        
        # Check ticket age
        created_at = ticket_data.get('created_at')
        if created_at:
            age_hours = (datetime.now() - datetime.fromisoformat(created_at.replace('Z', '+00:00'))).total_seconds() / 3600
            if age_hours > self.patterns['escalation_predictors']['age_threshold']:
                risk_score += 0.3
        
        # Check interaction count
        interaction_count = len(ticket_data.get('conversations', []))
        if interaction_count > self.patterns['escalation_predictors']['interaction_count']:
            risk_score += 0.2
        
        return min(risk_score, 1.0)
    
    def _estimate_resolution_time(self, classification: str, ticket_data: dict) -> dict:
        """Estimate time to resolution"""
        base_times = self.patterns['resolution_time_factors']
        
        # Parse classification
        parts = classification.lower().split('-')
        category = parts[0] if parts else 'general'
        sub_category = parts[1] if len(parts) > 1 else 'general'
        
        # Get base time
        base_time = base_times.get(category, {}).get(sub_category, 48)
        
        # Adjust based on complexity
        complexity_factor = self._assess_complexity(ticket_data)
        estimated_hours = base_time * complexity_factor
        
        return {
            'hours': estimated_hours,
            'confidence': 0.75,  # Could be calculated based on historical accuracy
            'factors': {
                'base_time': base_time,
                'complexity_factor': complexity_factor
            }
        }
    
    def _assess_satisfaction_risk(self, ticket_data: dict) -> float:
        """Assess risk of customer dissatisfaction"""
        risk_score = 0.0
        content = ticket_data.get('raw_ticket_content', '').lower()
        
        # Check sentiment indicators
        positive_count = sum(1 for word in self.patterns['customer_satisfaction_indicators']['positive'] if word in content)
        negative_count = sum(1 for word in self.patterns['customer_satisfaction_indicators']['negative'] if word in content)
        
        if negative_count > positive_count:
            risk_score += 0.5 * (negative_count / (negative_count + positive_count + 1))
        
        # Check response time
        if ticket_data.get('status') == 1:  # New ticket
            risk_score += 0.3
        
        return min(risk_score, 1.0)
    
    def _assess_complexity(self, ticket_data: dict) -> float:
        """Assess ticket complexity"""
        complexity = 1.0
        content = ticket_data.get('raw_ticket_content', '')
        
        # Factors that increase complexity
        if len(content) > 2000:
            complexity += 0.2
        if 'attachment' in content.lower():
            complexity += 0.1
        if any(word in content.lower() for word in ['multiple', 'several', 'complex']):
            complexity += 0.3
        
        return min(complexity, 2.0)
    
    def _assess_automation_potential(self, classification: str) -> float:
        """Assess how much of the process can be automated"""
        automation_scores = {
            'Claims-Cashless-Garage': 0.9,
            'Claims-Cashless-Hospital': 0.9,
            'Claims-PI-Request': 0.7,
            'Endorsement-Motor-NonFinancial': 0.8,
            'Endorsement-Health-NonFinancial': 0.8,
            'Support-PDPNR': 0.6,
            'Support-PLNG': 0.5,
            'General-Document-Request': 0.9,
            'General-Information-Request': 0.8
        }
        
        for key, score in automation_scores.items():
            if classification.startswith(key):
                return score
        
        return 0.3  # Default low automation potential
    
workflow_engine = WorkflowEngine()
predictive_engine = PredictiveAnalyticsEngine()
smart_response_generator = SmartResponseGenerator(anthropic_client) if anthropic_client else None
autonomous_action_system = AutonomousActionSystem(FRESHDESK_DOMAIN, FRESHDESK_API_KEY) if FRESHDESK_DOMAIN and FRESHDESK_API_KEY else None

# ========== INTEGRATION WITH MAIN PROCESS ==========

# Initialize global instances
autonomous_action_system = None
smart_response_generator = None
workflow_engine = None
predictive_engine = None

def initialize_autonomous_systems():
    """Initialize all autonomous systems"""
    global autonomous_action_system, smart_response_generator, workflow_engine, predictive_engine
    
    if FRESHDESK_DOMAIN and FRESHDESK_API_KEY:
        autonomous_action_system = AutonomousActionSystem(FRESHDESK_DOMAIN, FRESHDESK_API_KEY)
        workflow_engine = WorkflowAutomationEngine(FRESHDESK_DOMAIN, FRESHDESK_API_KEY)
    
    if anthropic_client:
        smart_response_generator = SmartResponseGenerator(anthropic_client)
    
    predictive_engine = PredictiveAnalyticsEngine()

# Call initialization
initialize_autonomous_systems()

def process_ticket_id_enhanced(ticket_id):
    """
    Enhanced ticket processing with autonomous features and advanced pending status detection
    """
    routing_analyzer = EnhancedContextualRoutingAnalyzer(FRESHDESK_DOMAIN, FRESHDESK_API_KEY)
    
    # ========== NEW: Initialize Enhanced Ticket Analyzer ==========
    enhanced_analyzer = EnhancedTicketAnalyzer(FRESHDESK_DOMAIN, FRESHDESK_API_KEY)
    
    # Get the original ticket data
    result_data = process_ticket_id_orignal(ticket_id)
    
    if not result_data or 'error' in result_data:
        return result_data
    
    # Get classification and SOP details
    classification = result_data.get('Classification', 'Unknown')
    _, sop_details = classify_ticket_with_sop(result_data.get('raw_ticket_content', ''))
    
    # ========== NEW: Comprehensive Child Ticket Analysis with Pending Status ==========
    if result_data and not 'error' in result_data:
        try:
            print(f"Starting comprehensive analysis for ticket {ticket_id}")
            
            # Use the enhanced analyzer for complete analysis
            comprehensive_analysis = enhanced_analyzer.analyze_ticket_with_children(ticket_id)
            
            if 'error' not in comprehensive_analysis:
                # Add the comprehensive analysis to result_data
                result_data['comprehensive_ticket_analysis'] = comprehensive_analysis
                
                # Extract key information for backward compatibility
                main_ticket = comprehensive_analysis['main_ticket']
                child_tickets = comprehensive_analysis['child_tickets']
                
                # Update existing fields with enhanced data
                result_data['pending_from'] = main_ticket['pending_from']['content_based']
                result_data['pending_confidence'] = main_ticket['pending_from']['confidence']
                result_data['pending_evidence'] = main_ticket['pending_from']['evidence']
                
                # Add enhanced status information
                result_data['enhanced_status'] = {
                    'display': main_ticket['status']['display'],
                    'category': main_ticket['status']['category'],
                    'action': main_ticket['status']['action']
                }
                
                # Add timing information
                result_data['timing_info'] = main_ticket['timing']
                
                # Add key extracted information
                result_data['key_information'] = main_ticket['key_information']
                
                # Add next expected action
                result_data['next_expected_action'] = main_ticket['next_expected_action']
                
                # Process child tickets if any
                if child_tickets:
                    print(f"Found {len(child_tickets)} child tickets")
                    result_data['child_tickets'] = child_tickets
                    
                    # Create child ticket summary for easy access
                    child_summary = []
                    for child in child_tickets:
                        child_summary.append({
                            'ticket_id': child['ticket_id'],
                            'subject': child['subject'],
                            'status': child['status']['display'],
                            'pending_from': child['pending_from']['content_based'],
                            'confidence': child['pending_from']['confidence'],
                            'key_info': child['key_information']
                        })
                    result_data['child_summary'] = child_summary
                
                # Add overall pending summary
                result_data['pending_summary'] = comprehensive_analysis['pending_summary']
                
                # Add relationship analysis
                result_data['relationship_analysis'] = comprehensive_analysis['relationship_analysis']
                
                # Add actionable insights
                result_data['actionable_insights'] = enhanced_analyzer.generate_actionable_insights(comprehensive_analysis)
                
                print(f"Comprehensive analysis completed successfully")
                
            else:
                print(f"Error in comprehensive analysis: {comprehensive_analysis['error']}")
                # Fall back to basic analysis if enhanced fails
                result_data['comprehensive_analysis_error'] = comprehensive_analysis['error']
                
        except Exception as e:
            print(f"Error in enhanced ticket analysis: {e}")
            result_data['enhanced_analysis_error'] = str(e)
            
            # Fall back to your existing basic child ticket analysis
            status = result_data.get('status', 0)
            if status in [10, 11, 12]:  # Parent ticket statuses
                print(f"Falling back to basic child ticket analysis...")
                
                child_tickets = fetch_child_tickets(ticket_id)
                result_data['child_tickets'] = child_tickets
                
                child_analyses = []
                for child in child_tickets:
                    child_id = child.get('id')
                    print(f"Analyzing child ticket {child_id}...")
                    
                    child_data = fetch_ticket_by_id(child_id)
                    if child_data:
                        child_conversations = fetch_all_ticket_conversations(child_id)
                        
                        child_analysis = {
                            'ticket_id': child_id,
                            'subject': child.get('subject'),
                            'status': status_map.get(child.get('status'), 'Unknown'),
                            'created_at': child.get('created_at'),
                            'conversations': len(child_conversations),
                            'last_update': child.get('updated_at')
                        }
                        
                        child_content = ""
                        for conv in child_conversations:
                            child_content += conv.get('body_text', '') + " "
                        
                        if 'claim' in child_content.lower():
                            import re
                            claim_match = re.search(r'claim\s*#?\s*([A-Z0-9/-]+)', child_content, re.IGNORECASE)
                            if claim_match:
                                child_analysis['claim_number'] = claim_match.group(1)
                        
                        if any(insurer in child_content.lower() for insurer in ['hdfc', 'icici', 'bajaj', 'tata']):
                            child_analysis['insurer_involved'] = True
                        
                        child_analyses.append(child_analysis)
                        
                result_data['child_analyses'] = child_analyses
    
    # ========== NEW: Quick Pending Status for Non-Parent Tickets ==========
    if not result_data.get('pending_from'):  # If not set by comprehensive analysis
        try:
            pending_summary = enhanced_analyzer.get_pending_status_summary(ticket_id)
            if 'error' not in pending_summary:
                result_data['pending_from'] = pending_summary['pending_from_analysis']
                result_data['pending_confidence'] = pending_summary['confidence']
                result_data['pending_evidence'] = pending_summary['evidence']
                result_data['next_expected_action'] = pending_summary['recommendation']
        except Exception as e:
            print(f"Error getting pending status summary: {e}")
    
    # Add autonomous actions
    if autonomous_action_system:
        autonomous_actions = autonomous_action_system.analyze_ticket_for_actions(
            result_data, classification, sop_details
        )
        result_data['autonomous_actions'] = autonomous_actions
    else:
        result_data['autonomous_actions'] = []
    
    # Add predictions
    if predictive_engine:
        predictions = predictive_engine.predict_ticket_outcome(result_data, classification)
        result_data['predictions'] = format_predictions(predictions)
    else:
        result_data['predictions'] = {}
    
    # Add workflow
    if workflow_engine:
        workflow = workflow_engine.create_workflow(
            result_data.get('Ticket ID'),
            classification,
            sop_details
        )
        result_data['workflow'] = workflow
    else:
        result_data['workflow'] = {}
    
    # Generate suggested response
    if smart_response_generator:
        suggested_response = smart_response_generator.generate_response(
            result_data,
            "Initial response",
            "general"
        )
        result_data['suggested_response'] = suggested_response
    else:
        result_data['suggested_response'] = ""
    
    # Get SOP steps based on category
    category = result_data.get('Classification') or result_data.get('sop_category', 'general')
    sop_steps = get_sop_steps_for_category(category)
    
    # Calculate workflow progress
    workflow_progress_data = calculate_workflow_progress(result_data, sop_steps)
    
    # Update workflow with progress
    if 'workflow' not in result_data:
        result_data['workflow'] = {}
    
    result_data['workflow'].update({
        'progress': workflow_progress_data['progress'],
        'status': workflow_progress_data['status'],
        'current_step': workflow_progress_data['current_step'],
        'completed_steps': workflow_progress_data.get('completed_steps', 0),
        'total_steps': workflow_progress_data.get('total_steps', len(sop_steps)),
        'remaining_steps': workflow_progress_data.get('remaining_steps', []),
        'sop_steps': sop_steps
    })

    # Process attachments if available
    if result_data.get('attachments'):
        try:
            # Analyze documents using Vision API
            result_data = process_ticket_attachments_enhanced(result_data)
            
            # Get document suggestions
            suggestion_engine = DocumentSuggestionEngine()
            result_data['suggested_documents'] = suggestion_engine.suggest_documents(result_data)
            
            # Process document workflow
            doc_workflow = DocumentWorkflowAutomation(DocumentAnalyzer())
            result_data['document_workflow'] = doc_workflow.process_document_workflow(result_data)
            
            # Add claims document automation
            if classification.startswith("Claims"):
                claims_automation_result = automated_claims_workflow(ticket_id)
                result_data['claims_automation'] = claims_automation_result
    
                # If documents are missing, add the generated response
                if claims_automation_result.get('success') and claims_automation_result.get('result'):
                     if claims_automation_result['result'].get('documents', {}).get('missing_required'):
                         result_data['suggested_response'] = claims_automation_result['result']['generated_response']
    
        except Exception as e:
            print(f"Error in document analysis: {e}")
            result_data['attachment_analysis'] = {'error': str(e)}
    
    return result_data

# ========== NEW: Additional Helper Functions ==========

def get_pending_status_summary(ticket_id):
    """Quick function to get just the pending status information"""
    analyzer = EnhancedTicketAnalyzer(FRESHDESK_DOMAIN, FRESHDESK_API_KEY)
    
    ticket_data = fetch_ticket_by_id(ticket_id)
    if not ticket_data:
        return {'error': f'Could not fetch ticket {ticket_id}'}
    
    conversations = fetch_all_ticket_conversations(ticket_id)
    raw_content, _ = extract_email_content_and_attachments(ticket_data, conversations)
    
    status = ticket_data.get('status')
    status_info = analyzer.status_mappings.get(status, {})
    
    actual_pending = analyzer.determine_actual_pending_from(
        raw_content, 
        conversations, 
        status_info.get('pending_from', 'unknown')
    )
    
    return {
        'ticket_id': ticket_id,
        'status_name': status_info.get('display', 'Unknown'),
        'pending_from_status': status_info.get('pending_from', 'unknown'),
        'pending_from_analysis': actual_pending['primary'],
        'confidence': actual_pending['confidence'],
        'evidence': actual_pending['evidence'],
        'recommendation': analyzer.determine_next_action(
            status_info, 
            actual_pending, 
            analyzer.extract_key_information(raw_content, conversations)
        )
    }

def analyze_ticket_comprehensively(ticket_id):
    """Main function to analyze ticket with children and pending status"""
    analyzer = EnhancedTicketAnalyzer(FRESHDESK_DOMAIN, FRESHDESK_API_KEY)
    
    try:
        result = analyzer.analyze_ticket_with_children(ticket_id)
        if 'error' in result:
            return result
        formatted_result = analyzer.format_analysis_for_display(result)
        formatted_result['raw_analysis'] = result
        return formatted_result
    except Exception as e:
        return {
            'error': f'Error analyzing ticket {ticket_id}: {str(e)}',
            'ticket_id': ticket_id
        }

def print_ticket_summary(ticket_id):
    """Print a nice summary of ticket status"""
    analysis = analyze_ticket_comprehensively(ticket_id)
    
    if 'error' in analysis:
        print(f"Error: {analysis['error']}")
        return
    
    summary = analysis['summary']
    pending = analysis['pending_analysis']
    
    print(f"\n=== TICKET {summary['ticket_id']} ANALYSIS ===")
    print(f"Subject: {summary['subject']}")
    print(f"Status: {summary['current_status']}")
    print(f"Age: {summary['age_days']} days")
    print(f"Type: {summary['relationship_type'].title()} ticket")
    
    if summary['child_count'] > 0:
        print(f"Child Tickets: {summary['child_count']}")
    
    print(f"\n--- PENDING STATUS ---")
    print(f"Currently Pending From: {pending['primary_source'].title()}")
    print(f"Confidence: {pending['confidence']:.2%}")
    print(f"Evidence: {', '.join(pending['evidence'][:3])}")
    
    if analysis['next_actions']:
        print(f"\n--- NEXT ACTIONS ---")
        for action in analysis['next_actions']:
            print(f"• {action['action']}: {action['details']}")
            print(f"  Priority: {action['priority']}, Timeline: {action['timeline']}")
    
    if analysis['child_summary']:
        print(f"\n--- CHILD TICKETS ---")
        for child in analysis['child_summary']:
            print(f"• #{child['ticket_id']}: {child['status']} (pending from {child['pending_from']})")
    
    if analysis['insights']:
        print(f"\n--- INSIGHTS ---")
        for insight in analysis['insights']:
            print(f"• [{insight['type'].upper()}] {insight['message']}")
            print(f"  Action: {insight['action']}")
            
# Ensure all systems are initialized
initialize_autonomous_systems()

# Export all necessary items
__all__ = [
    # Main functions
    'process_ticket_id_enhanced',
    'get_enhanced_claude_answer',
    'get_claude_answer',  # Alias
    
    # Autonomous systems
    'autonomous_action_system',
    'workflow_engine',
    'predictive_engine',
    'smart_response_generator',
    
    # Helper functions
    'calculate_workflow_progress',
    'format_predictions',
    'get_sop_steps_for_category',
    
    # Classes (if you want to export them)
    'AutonomousActionSystem',
    'WorkflowAutomationEngine',
    'PredictiveAnalyticsEngine',
    'SmartResponseGenerator',
    # Add these new exports
    'EnhancedContextualRoutingAnalyzer',
    'EnhancedSmartResponseGenerator',
    'enhance_ticket_with_routing_context',
    'process_ticket_id_with_routing',
    'routing_analyzer',
    'enhanced_response_generator',
    'DocumentRequirementEngine',
    'extract_claim_type_from_ticket',
    'extract_insurer_from_ticket',
    'check_existing_attachments',
    'generate_document_request_response',
    'process_claims_ticket_with_documents',
    'automated_claims_workflow',
    'send_automated_response',
    'update_ticket_tags',
    'update_ticket_status',
    
    # Original functions (keep for compatibility)
    'process_ticket_id_orignal',
    'process_ticket_id_for_gui',
    'classify_ticket_with_sop',
    'fetch_ticket_by_id',
    'fetch_all_ticket_conversations',
    'get_claude_summary',
    'download_attachment',
    'download_all_ticket_attachments',
    'analyze_downloaded_attachments',
]

# Create an alias for backward compatibility
process_ticket_id = process_ticket_id_enhanced

# ========== ENHANCED CHATBOT FUNCTION ==========
def _format_received_documents(ticket_data):
    """Return a formatted string of received documents for the response."""
    docs = []
    if ticket_data:
        if 'attachment_analysis' in ticket_data and ticket_data['attachment_analysis'].get('analyzed'):
            docs = [doc.get('document_type', doc.get('filename', 'Document')) for doc in ticket_data['attachment_analysis']['analyzed']]
        elif 'attachments' in ticket_data:
            docs = [att.get('name', att.get('filename', 'Document')) for att in ticket_data['attachments']]
    if not docs:
        return "No documents received yet."
    return "\n- " + "\n- ".join(docs)

def get_enhanced_claude_answer(ticket_content_text: str, user_question: str, ticket_data: dict = None) -> str:
    """Enhanced insurance broker agent that provides contextual responses based on complete ticket analysis"""
    if not anthropic_client:
        return "Error: Claude client not initialized. Please check your API configuration."
    
    # Enhanced system prompt for insurance broker specialization
    INSURANCE_BROKER_SYSTEM_PROMPT = """You are an expert insurance broker agent specializing in claims and endorsements for InsuranceDekho. 
    You have deep knowledge of:
    - Motor, Health, Life, and MSME insurance claims processing
    - Policy endorsements (financial and non-financial)
    - Insurer-specific requirements and processes
    - Document requirements and verification
    - Escalation procedures and TATs
    - Customer service excellence in insurance

    When responding:
    1. Always consider the complete ticket context including parent-child relationships
    2. Provide specific, actionable guidance based on the ticket status and pending items
    3. Reference relevant policy numbers, claim numbers, and insurer details when available
    4. Suggest next steps based on current ticket status and SOP requirements
    5. Be empathetic but professional, understanding customer concerns
    6. Provide realistic timelines without making false promises
    7. Escalate appropriately when needed

    Act as a knowledgeable insurance professional who can solve complex claims and endorsement issues."""

    try:
        # Extract comprehensive context from ticket data
        context_info = _extract_comprehensive_context(ticket_data) if ticket_data else {}
        
        # Build enhanced prompt with full context
        enhanced_prompt = f"""
TICKET ANALYSIS CONTEXT:
{_format_ticket_context(ticket_data, context_info)}

CUSTOMER QUESTION: {user_question}

TICKET CONTENT:
{ticket_content_text[:3000]}...

Based on the complete ticket analysis above, provide a comprehensive response as an insurance broker agent. 

Consider:
- Current ticket status and what's pending
- Parent-child ticket relationships if any
- Specific claim/policy numbers mentioned
- Required documents and their status
- Appropriate next steps and timelines
- Any escalation needs

Provide a helpful, professional response that addresses the customer's question while considering the full context of their case.
"""

        response = anthropic_client.messages.create(
            model="claude-3-haiku-20240307",
            max_tokens=1500,
            messages=[{"role": "user", "content": enhanced_prompt}],
            system=INSURANCE_BROKER_SYSTEM_PROMPT,
            temperature=0.6
        )
        
        return response.content[0].text.strip()
        
    except Exception as e:
        print(f"Error in enhanced Claude response: {e}")
        # Fallback to basic response
        return _generate_fallback_insurance_response(user_question, ticket_data)

def _extract_comprehensive_context(ticket_data: dict) -> dict:
    """Extract comprehensive context from ticket data for better responses"""
    if not ticket_data:
        return {}
    
    context = {
        'ticket_id': ticket_data.get('Ticket ID'),
        'classification': ticket_data.get('Classification', 'Unknown'),
        'status': ticket_data.get('status'),
        'pending_from': ticket_data.get('pending_from', 'Unknown'),
        'pending_confidence': ticket_data.get('pending_confidence', 0),
        'key_entities': ticket_data.get('key_entities', {}),
        'child_tickets': ticket_data.get('child_summary', []),
        'timing_info': ticket_data.get('timing_info', {}),
        'next_action': ticket_data.get('next_expected_action', {}),
        'autonomous_actions': ticket_data.get('autonomous_actions', []),
        'document_status': ticket_data.get('attachment_analysis', {}),
        'comprehensive_analysis': ticket_data.get('comprehensive_ticket_analysis', {})
    }
    
    return context

def _format_ticket_context(ticket_data: dict, context_info: dict) -> str:
    """Format ticket context for AI prompt"""
    if not ticket_data:
        return "No ticket context available."
    
    formatted_context = []
    
    # Basic ticket info
    formatted_context.append(f"Ticket ID: {context_info.get('ticket_id', 'Unknown')}")
    formatted_context.append(f"Classification: {context_info.get('classification', 'Unknown')}")
    formatted_context.append(f"Current Status: {status_map.get(context_info.get('status'), 'Unknown')}")
    
    # Pending status with confidence
    pending_from = context_info.get('pending_from', 'Unknown')
    pending_confidence = context_info.get('pending_confidence', 0)
    formatted_context.append(f"Currently Pending From: {pending_from} (Confidence: {pending_confidence:.0%})")
    
    # Key entities (claim numbers, policy numbers, etc.)
    key_entities = context_info.get('key_entities', {})
    if key_entities.get('claim_numbers'):
        formatted_context.append(f"Claim Numbers: {', '.join(key_entities['claim_numbers'])}")
    if key_entities.get('policy_numbers'):
        formatted_context.append(f"Policy Numbers: {', '.join(key_entities['policy_numbers'])}")
    if key_entities.get('insurer_names'):
        formatted_context.append(f"Insurers Involved: {', '.join(key_entities['insurer_names'])}")
    
    # Child tickets if any
    child_tickets = context_info.get('child_tickets', [])
    if child_tickets:
        formatted_context.append(f"Child Tickets: {len(child_tickets)} related tickets")
        for child in child_tickets[:3]:  # Show first 3
            formatted_context.append(f"  - #{child.get('ticket_id')}: {child.get('status')} (pending from {child.get('pending_from')})")
    
    # Timing information
    timing_info = context_info.get('timing_info', {})
    if timing_info.get('age_days'):
        formatted_context.append(f"Ticket Age: {timing_info['age_days']} days")
    if timing_info.get('hours_since_last_update'):
        formatted_context.append(f"Last Update: {timing_info['hours_since_last_update']:.1f} hours ago")
    
    # Next expected action
    next_action = context_info.get('next_action', {})
    if next_action.get('action'):
        formatted_context.append(f"Next Expected Action: {next_action['action']} (Priority: {next_action.get('priority', 'Medium')})")
    
    # Document status
    doc_status = context_info.get('document_status', {})
    if doc_status.get('analyzed'):
        analyzed_docs = len(doc_status['analyzed'])
        formatted_context.append(f"Documents Analyzed: {analyzed_docs}")
    if doc_status.get('missing_documents'):
        missing_docs = len(doc_status['missing_documents'])
        formatted_context.append(f"Missing Documents: {missing_docs}")
    
    # Autonomous actions available
    autonomous_actions = context_info.get('autonomous_actions', [])
    if autonomous_actions:
        high_priority_actions = [a for a in autonomous_actions if a.get('priority') == 'HIGH']
        if high_priority_actions:
            formatted_context.append(f"High Priority Actions Available: {len(high_priority_actions)}")
    
    # Comprehensive analysis summary
    comprehensive = context_info.get('comprehensive_analysis', {})
    if comprehensive.get('main_ticket'):
        main_ticket = comprehensive['main_ticket']
        if main_ticket.get('key_information'):
            key_info = main_ticket['key_information']
            if key_info.get('claim_numbers'):
                formatted_context.append(f"Extracted Claim Info: {len(key_info['claim_numbers'])} claims identified")
            if key_info.get('document_types'):
                formatted_context.append(f"Document Types Found: {', '.join(key_info['document_types'][:3])}")
    
    return '\n'.join(formatted_context)

def _generate_fallback_insurance_response(user_question: str, ticket_data: dict) -> str:
    """Generate fallback response when AI fails"""
    if not ticket_data:
        return """I understand you have a query about your insurance case. To provide you with the most accurate assistance, I would need to review your complete ticket details. 

As your insurance broker, I'm here to help with:
- Claims processing and status updates
- Policy endorsements and modifications  
- Document requirements and verification
- Insurer coordination and follow-ups
- Escalation when needed

Please ensure your ticket ID is properly loaded so I can provide specific guidance for your case."""

    ticket_id = ticket_data.get('Ticket ID', 'your case')
    classification = ticket_data.get('Classification', 'insurance request')
    status = status_map.get(ticket_data.get('status'), 'being processed')
    
    # Generate contextual fallback based on available data
    response_parts = []
    
    # Acknowledge the question
    if 'status' in user_question.lower():
        response_parts.append(f"Regarding the status of your {classification.lower()}")
    elif 'document' in user_question.lower():
        response_parts.append(f"Regarding the documents for your {classification.lower()}")
    elif 'claim' in user_question.lower():
        response_parts.append(f"Regarding your claim inquiry")
    else:
        response_parts.append(f"Thank you for your question about {classification.lower()}")
    
    # Add current status
    response_parts.append(f"Your case (#{ticket_id}) is currently {status}.")
    
    # Add pending information if available
    pending_from = ticket_data.get('pending_from')
    if pending_from and pending_from != 'Unknown':
        if pending_from == 'customer':
            response_parts.append("We are currently waiting for some information or documents from you.")
        elif pending_from == 'insurer':
            response_parts.append("We have submitted your case to the insurance company and are awaiting their response.")
        elif pending_from == 'internal_team':
            response_parts.append("Our team is actively processing your request.")
        else:
            response_parts.append(f"Your case is currently with {pending_from} for further action.")
    
    # Add next steps
    next_action = ticket_data.get('next_expected_action', {})
    if next_action.get('action'):
        response_parts.append(f"Next step: {next_action['action']}")
    else:
        response_parts.append("Our team will update you as soon as we have more information.")
    
    # Professional closing
    response_parts.append("As your insurance broker, I'm committed to ensuring your case is resolved efficiently. If you have any specific concerns, please let me know and I'll address them immediately.")
    
    return ' '.join(response_parts)




# ========== Main function for testing ==========
if __name__ == "__main__":
    print("\n--- Starting Enhanced Freshdeskintegration with SOP Support ---")
# Test attachment download
    test_ticket_with_attachments = "5163041"  # Replace with a ticket ID that has attachments
    print(f"\n--- Testing Attachment Download for Ticket {test_ticket_with_attachments} ---")
    
    # Test 1: Just download attachments
    print("\nTest 1: Downloading attachments...")
    attachments = download_all_ticket_attachments(test_ticket_with_attachments, "test_downloads/")
    print(f"Downloaded {len(attachments)} attachments")
    
    # Test 2: Download and analyze
    print("\nTest 2: Analyzing attachments...")
    analysis_results = analyze_downloaded_attachments(test_ticket_with_attachments)
    for result in analysis_results:
        print(f"- {result.get('filename')}: {result.get('document_type')} (confidence: {result.get('confidence', 0):.2f})")
    if not FRESHDESK_DOMAIN or not FRESHDESK_API_KEY or not CLAUDE_API_KEY:
        print("Please set FRESHDESK_DOMAIN, FRESHDESK_API_KEY, and CLAUDE_API_KEY in your .env file.")
        print("Exiting test.")
    else:
        test_ticket_id = "1"  # Replace with a real ticket ID

        print(f"\n--- Testing Enhanced RCA Summary with SOP Context for Ticket ID: {test_ticket_id} ---")
        rca_result = process_ticket_id_for_gui(test_ticket_id)
        
        raw_ticket_content_for_nlp = ""

        if "error" in rca_result:
            print(f"RCA Summary Error: {rca_result['error']}")
            print("Cannot proceed with NLP Chatbot Query without successful RCA processing.")
        else:
            print(f"RCA Summary Result: {json.dumps(rca_result, indent=2)}")
            raw_ticket_content_for_nlp = rca_result.get('raw_ticket_content', '')
            
            print(f"\n--- SOP Category: {rca_result.get('sop_category', 'Unknown')} ---")
            print(f"Raw Ticket Content Length: {len(raw_ticket_content_for_nlp)} characters")

            print(f"\n--- Testing NLP Chatbot Query with SOP Context for Ticket ID: {test_ticket_id} ---")
            
            if not raw_ticket_content_for_nlp.strip():
                print("No raw ticket content available for NLP queries. Skipping chatbot test.")
            else:
                # Test questions that would benefit from SOP knowledge
                questions = [
                    "What is the main problem described in this ticket?",
                    "What is the SOP for handling this type of ticket?",
                    "What is the TAT for this request?",
                    "Who should I escalate this to if not resolved?",
                    "What documents are required for this case?",
                    "What are the next steps according to the SOP?",
                    "Is this a financial or non-financial endorsement?",
                    "What is the process for handling this claim?"
                ]

                for q in questions:
                    print(f"\nUser: {q}")
                    response = get_claude_answer(raw_ticket_content_for_nlp, q)
                    print(f"AI: {response}")
                    print("-" * 30)
    # Update the main process function to include routing analysis
def process_ticket_id_with_routing(ticket_id):
   """
   Enhanced ticket processing with complete routing analysis
   """
   # Get base ticket data using existing function
   ticket_data = process_ticket_id_enhanced(ticket_id)
   
   if not ticket_data or 'error' in ticket_data:
       return ticket_data
   
   # Enhance with routing context
   enhanced_data = enhance_ticket_with_routing_context(ticket_data)
   
   # Add summary of what's happening
   enhanced_data['executive_summary'] = _generate_executive_summary(enhanced_data)
   
   return enhanced_data
    
print("\n--- End of Enhanced Freshdeskintegration Test ---")
